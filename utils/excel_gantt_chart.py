"""
ガントチャート作成関数
- セルの塗りつぶしでバーを表現（ブラウザ版と同じ見た目）
- Y軸左側にユニット名を明確表示
- X軸上部に日付ヘッダー
- 今日のラインを赤で強調
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta


# ユニットごとの色パレット（ブラウザ版と合わせる）
UNIT_COLORS = [
    "4472C4",  # 青
    "ED7D31",  # オレンジ
    "70AD47",  # 緑
    "FFC000",  # 黄
    "5B9BD5",  # 水色
    "FF6384",  # ピンク
    "A855F7",  # 紫
    "44BBA4",  # ティール
    "FF9F40",  # 薄オレンジ
    "36A2EB",  # ライト青
    "C45850",  # 赤茶
    "8B5CF6",  # バイオレット
    "059669",  # エメラルド
    "DC2626",  # 赤
    "7C3AED",  # インディゴ
]


def create_gantt_chart_sheet(wb, seiban, orders):
    """
    セルベースのガントチャートシートを作成（ブラウザ版と同じ見た目）
    """
    try:
        ws = wb.create_sheet("納期ガントチャート", 0)

        # ========================================
        # 1. データ収集
        # ========================================
        unit_data = []
        all_dates = []

        for order in orders:
            unit_name = order.unit if order.unit else 'ユニット名無し'

            if not order.details:
                continue

            dates = []
            for detail in order.details:
                if detail.delivery_date:
                    try:
                        date_str = detail.delivery_date.strip()
                        if not date_str or date_str == '-':
                            continue
                        if '/' in date_str:
                            parts = date_str.split('/')
                            if len(parts) == 3:
                                year = int(parts[0])
                                if year < 100:
                                    year = 2000 + year if year < 50 else 1900 + year
                                date_obj = datetime(year, int(parts[1]), int(parts[2]))
                                dates.append(date_obj)
                        elif '-' in date_str:
                            parts = date_str.split('-')
                            if len(parts) == 3:
                                date_obj = datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                                dates.append(date_obj)
                    except Exception:
                        continue

            if dates:
                min_date = min(dates)
                max_date = max(dates)
                all_dates.extend(dates)
                unit_data.append({
                    'unit': unit_name,
                    'min_date': min_date,
                    'max_date': max_date,
                    'count': len(order.details),
                    'detail_count': len(dates)
                })

        if not unit_data or not all_dates:
            ws['A3'] = '納期データがありません'
            ws['A3'].font = Font(size=14, color="FF0000", bold=True)
            return

        # 日付範囲（7日前～7日後に余裕を持たせる）
        global_min = min(all_dates) - timedelta(days=3)
        global_max = max(all_dates) + timedelta(days=3)
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        total_days = (global_max - global_min).days + 1

        # ユニットを最早納期順にソート
        unit_data.sort(key=lambda x: x['min_date'])

        # ========================================
        # 2. レイアウト定数
        # ========================================
        HEADER_ROW = 1       # タイトル
        INFO_ROW = 2         # 期間情報
        DATE_HEADER_ROW = 3  # 日付ヘッダー（月表示）
        DAY_HEADER_ROW = 4   # 日付ヘッダー（日表示）
        DATA_START_ROW = 5   # データ開始
        UNIT_COL = 1         # ユニット名列
        INFO_COL = 2         # 情報列（期間・件数）
        BAR_START_COL = 3    # バー開始列

        # ========================================
        # 3. タイトル・情報
        # ========================================
        ws.cell(row=HEADER_ROW, column=UNIT_COL, value=f"{seiban} 納期ガントチャート")
        ws.cell(row=HEADER_ROW, column=UNIT_COL).font = Font(size=16, bold=True, color="1F4E78")
        ws.merge_cells(start_row=HEADER_ROW, start_column=UNIT_COL,
                        end_row=HEADER_ROW, end_column=min(BAR_START_COL + 15, BAR_START_COL + total_days - 1))
        ws.row_dimensions[HEADER_ROW].height = 28

        period_text = f"期間: {global_min.strftime('%Y/%m/%d')} ～ {global_max.strftime('%Y/%m/%d')}  今日: {today.strftime('%Y/%m/%d')}  ユニット数: {len(unit_data)}"
        ws.cell(row=INFO_ROW, column=UNIT_COL, value=period_text)
        ws.cell(row=INFO_ROW, column=UNIT_COL).font = Font(size=10, italic=True, color="666666")
        ws.row_dimensions[INFO_ROW].height = 18

        # ========================================
        # 4. 日付ヘッダー
        # ========================================
        thin_border = Border(
            left=Side(style='thin', color="DDDDDD"),
            right=Side(style='thin', color="DDDDDD"),
            top=Side(style='thin', color="DDDDDD"),
            bottom=Side(style='thin', color="DDDDDD")
        )

        # ユニット列・情報列ヘッダー
        for row in [DATE_HEADER_ROW, DAY_HEADER_ROW]:
            cell = ws.cell(row=row, column=UNIT_COL, value="ユニット" if row == DATE_HEADER_ROW else "")
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            cell2 = ws.cell(row=row, column=INFO_COL, value="納期" if row == DATE_HEADER_ROW else "")
            cell2.font = Font(bold=True, size=9, color="FFFFFF")
            cell2.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            cell2.alignment = Alignment(horizontal='center', vertical='center')
            cell2.border = thin_border

        # 日付列ヘッダー
        weekend_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
        today_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        prev_month = None
        month_start_col = None

        for day_offset in range(total_days):
            col = BAR_START_COL + day_offset
            current_date = global_min + timedelta(days=day_offset)
            is_weekend = current_date.weekday() >= 5
            is_today = current_date.date() == today.date()

            # 月ヘッダー（月が変わったらマージ）
            current_month = current_date.strftime('%Y/%m')
            if current_month != prev_month:
                if prev_month is not None and month_start_col is not None:
                    if col - 1 > month_start_col:
                        ws.merge_cells(start_row=DATE_HEADER_ROW, start_column=month_start_col,
                                        end_row=DATE_HEADER_ROW, end_column=col - 1)
                month_start_col = col
                cell = ws.cell(row=DATE_HEADER_ROW, column=col, value=current_date.strftime('%Y年%m月'))
                cell.font = Font(bold=True, size=8, color="333333")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                prev_month = current_month

            # 日ヘッダー
            weekday_names = ['月', '火', '水', '木', '金', '土', '日']
            day_label = f"{current_date.day}"
            cell = ws.cell(row=DAY_HEADER_ROW, column=col, value=day_label)
            cell.font = Font(size=7, color="CC0000" if is_weekend else ("333333"))
            if is_today:
                cell.fill = today_fill
                cell.font = Font(size=7, bold=True, color="856404")
            elif is_weekend:
                cell.fill = weekend_fill
            else:
                cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 最後の月のマージ
        if month_start_col is not None:
            last_col = BAR_START_COL + total_days - 1
            if last_col > month_start_col:
                ws.merge_cells(start_row=DATE_HEADER_ROW, start_column=month_start_col,
                                end_row=DATE_HEADER_ROW, end_column=last_col)

        # ========================================
        # 5. データ行（ユニットごとのバー）
        # ========================================
        for idx, data in enumerate(unit_data):
            row = DATA_START_ROW + idx
            color = UNIT_COLORS[idx % len(UNIT_COLORS)]
            bar_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # ユニット名
            cell = ws.cell(row=row, column=UNIT_COL, value=data['unit'])
            cell.font = Font(size=9, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border

            # 納期情報
            period_info = f"{data['min_date'].strftime('%m/%d')}～{data['max_date'].strftime('%m/%d')} ({data['detail_count']}件)"
            cell2 = ws.cell(row=row, column=INFO_COL, value=period_info)
            cell2.font = Font(size=8, color="666666")
            cell2.alignment = Alignment(horizontal='center', vertical='center')
            cell2.border = thin_border

            # バーを描画
            bar_start = (data['min_date'] - global_min).days
            bar_end = (data['max_date'] - global_min).days

            for day_offset in range(total_days):
                col = BAR_START_COL + day_offset
                current_date = global_min + timedelta(days=day_offset)
                is_weekend = current_date.weekday() >= 5
                is_today = current_date.date() == today.date()
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

                if bar_start <= day_offset <= bar_end:
                    # バー範囲内
                    cell.fill = bar_fill
                    # バーの中央付近にユニット名を表示
                    mid_offset = (bar_start + bar_end) // 2
                    if day_offset == mid_offset:
                        cell.value = data['unit']
                        cell.font = Font(size=7, bold=True, color="FFFFFF")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                elif is_today:
                    cell.fill = PatternFill(start_color="FFEEEE", end_color="FFEEEE", fill_type="solid")
                elif is_weekend:
                    cell.fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

            ws.row_dimensions[row].height = 22

        # ========================================
        # 6. 今日の列を赤枠で強調
        # ========================================
        today_offset = (today - global_min).days
        if 0 <= today_offset < total_days:
            today_col = BAR_START_COL + today_offset
            today_border = Border(
                left=Side(style='medium', color="FF0000"),
                right=Side(style='medium', color="FF0000"),
                top=Side(style='thin', color="DDDDDD"),
                bottom=Side(style='thin', color="DDDDDD")
            )
            for r in range(DATE_HEADER_ROW, DATA_START_ROW + len(unit_data)):
                ws.cell(row=r, column=today_col).border = today_border

        # ========================================
        # 7. 凡例行
        # ========================================
        legend_row = DATA_START_ROW + len(unit_data) + 1
        ws.cell(row=legend_row, column=UNIT_COL, value="凡例:").font = Font(size=9, bold=True)
        for idx, data in enumerate(unit_data):
            col = INFO_COL + idx
            color = UNIT_COLORS[idx % len(UNIT_COLORS)]
            cell = ws.cell(row=legend_row, column=col, value=data['unit'])
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(size=8, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # ========================================
        # 8. 列幅・行高さ調整
        # ========================================
        ws.column_dimensions[get_column_letter(UNIT_COL)].width = 18
        ws.column_dimensions[get_column_letter(INFO_COL)].width = 16
        for day_offset in range(total_days):
            col_letter = get_column_letter(BAR_START_COL + day_offset)
            ws.column_dimensions[col_letter].width = 3.5

        ws.row_dimensions[DATE_HEADER_ROW].height = 18
        ws.row_dimensions[DAY_HEADER_ROW].height = 16

        # ========================================
        # 9. 印刷設定（横向き）
        # ========================================
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # ウィンドウ枠の固定（ユニット名と日付ヘッダーを固定）
        ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=BAR_START_COL)

        print(f"✅ ガントチャート作成完了: {len(unit_data)}ユニット、{total_days}日間")

    except Exception as e:
        print(f"⚠️  ガントチャート作成エラー: {e}")
        import traceback
        traceback.print_exc()
