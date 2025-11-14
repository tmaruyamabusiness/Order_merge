"""
Excel装飾ユーティリティモジュール
Untitled.py の198行目～234行目から抽出
"""

from openpyxl.styles import Font, PatternFill
from .constants import Constants


class ExcelStyler:
    """Excel装飾ユーティリティ"""
    
    @staticmethod
    def get_fill(is_received, is_even_row, is_child=False):
        """
        背景色取得
        
        Args:
            is_received: 受入済みかどうか
            is_even_row: 偶数行かどうか
            is_child: 子要素（部品）かどうか
            
        Returns:
            PatternFill: 背景色のPatternFillオブジェクト
        """
        if is_received:
            # 受入済みは行番号で交互に色を変える
            if is_even_row:
                return PatternFill(start_color=Constants.COLOR_ACCEPT_DARK, 
                                 end_color=Constants.COLOR_ACCEPT_DARK, fill_type="solid")
            else:
                return PatternFill(start_color=Constants.COLOR_ACCEPT_LIGHT, 
                                 end_color=Constants.COLOR_ACCEPT_LIGHT, fill_type="solid")
        if is_child:
            return PatternFill(start_color=Constants.COLOR_CHILD, 
                             end_color=Constants.COLOR_CHILD, fill_type="solid")
        if is_even_row:
            return PatternFill(start_color=Constants.COLOR_GRAY, 
                             end_color=Constants.COLOR_GRAY, fill_type="solid")
        return PatternFill(start_color=Constants.COLOR_WHITE, 
                         end_color=Constants.COLOR_WHITE, fill_type="solid")
    
    @staticmethod
    def get_font(is_blank=False, is_mekki=False):
        """
        フォント取得
        
        Args:
            is_blank: ブランク（加工用）かどうか
            is_mekki: メッキ出かどうか
            
        Returns:
            Font or None: 赤色太字フォント（該当する場合）またはNone
        """
        if is_mekki or is_blank:
            return Font(color=Constants.COLOR_RED, bold=True)
        return None
    
    @staticmethod
    def apply_column_widths(ws):
        """
        列幅適用
        
        Args:
            ws: openpyxlのワークシートオブジェクト
        """
        for col_letter, width in Constants.COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width


# テスト用コード
if __name__ == '__main__':
    from openpyxl import Workbook
    
    print("=== ExcelStyler テスト ===")
    
    # get_fillのテスト
    print("\n--- get_fill テスト ---")
    fill1 = ExcelStyler.get_fill(is_received=True, is_even_row=True)
    print(f"受入済み・偶数行 → {fill1.start_color.rgb}")
    
    fill2 = ExcelStyler.get_fill(is_received=True, is_even_row=False)
    print(f"受入済み・奇数行 → {fill2.start_color.rgb}")
    
    fill3 = ExcelStyler.get_fill(is_received=False, is_even_row=True, is_child=True)
    print(f"未受入・子要素 → {fill3.start_color.rgb}")
    
    # get_fontのテスト
    print("\n--- get_font テスト ---")
    font1 = ExcelStyler.get_font(is_blank=True)
    print(f"ブランク → color={font1.color.rgb if font1 else None}, bold={font1.bold if font1 else None}")
    
    font2 = ExcelStyler.get_font(is_mekki=True)
    print(f"メッキ → color={font2.color.rgb if font2 else None}, bold={font2.bold if font2 else None}")
    
    font3 = ExcelStyler.get_font()
    print(f"通常 → {font3}")
    
    # apply_column_widthsのテスト
    print("\n--- apply_column_widths テスト ---")
    wb = Workbook()
    ws = wb.active
    ExcelStyler.apply_column_widths(ws)
    print(f"A列の幅: {ws.column_dimensions['A'].width}")
    print(f"F列の幅: {ws.column_dimensions['F'].width}")
