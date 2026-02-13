"""
Flask Web Application for 手配発注マージシステム
"""

from flask import Flask, render_template, request, jsonify, send_file, session
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta, timezone
import pandas as pd
import os
import sys
import tempfile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment ,Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
import hashlib
import json
from werkzeug.utils import secure_filename
import re
import qrcode
import io
from io import BytesIO
import base64
import threading
import time
import shutil
import pyodbc
from pathlib import Path
import pytz
from datetime import datetime, timedelta, timezone
import subprocess
import win32com.client as win32
from threading import Thread
import pythoncom
from flask_cors import CORS
from openpyxl.worksheet.page import PageMargins
from openpyxl.chart import BarChart, Reference
import glob
from PIL import Image
from utils import Constants, DataUtils, MekkiUtils, ExcelStyler, generate_qr_code, create_gantt_chart_sheet, EmailSender, DeliveryUtils


app = Flask(__name__)

# Load configuration
# config.pyがある場合はそちらを使用、なければデフォルト設定
try:
    from config import get_config
    app.config.from_object(get_config())
except ImportError:
    # デフォルト設定
    app.config['SECRET_KEY'] = 'your-secret-key-here-change-in-production'
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///order_management.db'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    # コネクションプール設定（枯渇エラー対策）
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_size': 10,
        'max_overflow': 20,
        'pool_pre_ping': True,
        'pool_recycle': 300,  # 5分で接続をリサイクル
    }
    app.config['UPLOAD_FOLDER'] = 'uploads'
    app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size for large Excel files

    # Network path configuration
    app.config['HISTORY_EXCEL_PATH'] = r'\\server3\Share-data\Document\仕入れ\002_手配リスト\手配発注マージリスト発行履歴.xlsx'
    app.config['SEIBAN_LIST_PATH'] = r'\\server3\share-data\Document\Acrossデータ\製番一覧表.xlsx'
    app.config['EXPORT_EXCEL_PATH'] = r'\\SERVER3\Share-data\Document\仕入れ\002_手配リスト\手配発注リスト'
    app.config['USE_ODBC'] = False  # ODBCを使用する場合はTrue
    app.config['ODBC_CONNECTION_STRING'] = ''  # ODBC接続文字列（必要に応じて設定）

db = SQLAlchemy(app)

# リクエスト後のDBセッションクリーンアップ（コネクションプール枯渇対策）
@app.teardown_appcontext
def shutdown_session(exception=None):
    db.session.remove()

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('exports', exist_ok=True)
os.makedirs('cache', exist_ok=True)

# Global variables for background tasks
last_refresh_time = None
cached_file_info = {}


# Database Models

class Order(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    seiban = db.Column(db.String(50), nullable=False)
    unit = db.Column(db.String(100))
    status = db.Column(db.String(50), default='受入準備前')
    location = db.Column(db.String(50), default='未定')
    remarks = db.Column(db.Text)
    product_name = db.Column(db.String(200))
    customer_abbr = db.Column(db.String(100))
    memo2 = db.Column(db.String(200))
    pallet_number = db.Column(db.String(50))
    floor = db.Column(db.String(10))
    image_path = db.Column(db.String(500))
    is_archived = db.Column(db.Boolean, default=False)
    archived_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))


class OrderDetail(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.Integer, db.ForeignKey('order.id'), nullable=False)
    delivery_date = db.Column(db.String(20))
    supplier = db.Column(db.String(100))
    supplier_cd = db.Column(db.String(50))
    order_number = db.Column(db.String(50))
    quantity = db.Column(db.Integer)
    unit_measure = db.Column(db.String(20))
    item_name = db.Column(db.String(200))
    spec1 = db.Column(db.String(200))
    spec2 = db.Column(db.String(200))
    item_code = db.Column(db.String(50))
    order_type_code = db.Column(db.String(20))
    order_type = db.Column(db.String(50))
    maker = db.Column(db.String(100))
    remarks = db.Column(db.Text)
    member_count = db.Column(db.Integer)
    required_count = db.Column(db.Integer)
    seiban = db.Column(db.String(50))
    material = db.Column(db.String(100))
    is_received = db.Column(db.Boolean, default=False)
    received_at = db.Column(db.DateTime)
    received_quantity = db.Column(db.Integer)  # 実際に受け入れた数量（Noneの場合は全数受入）
    has_internal_processing = db.Column(db.Boolean, default=False)  # 社内加工フラグ
    parent_id = db.Column(db.Integer, db.ForeignKey('order_detail.id'), nullable=True)# 🔥 親子関係フィールド
    part_number = db.Column(db.String(50))
    page_number = db.Column(db.String(20))
    row_number = db.Column(db.String(20))
    hierarchy = db.Column(db.Integer)
    reply_delivery_date = db.Column(db.String(20))  # 回答納期
    # (children関係）
    children = db.relationship('OrderDetail', 
                            backref=db.backref('parent', remote_side=[id]),
                            lazy='dynamic')
    
    order = db.relationship('Order', backref=db.backref('details', lazy=True))


class ReceivedHistory(db.Model):
    """受入履歴テーブル - 発注番号をキーに受入情報を永続保存"""
    id = db.Column(db.Integer, primary_key=True)
    order_number = db.Column(db.String(50), nullable=False, index=True)  # 発注番号（キー）
    item_name = db.Column(db.String(200))  # 品名
    spec1 = db.Column(db.String(200))  # 仕様1
    quantity = db.Column(db.Integer)  # 手配数量
    received_quantity = db.Column(db.Integer)  # 実際に受け入れた数量
    is_received = db.Column(db.Boolean, default=True)  # 受入状態（True=受入、False=キャンセル）
    received_at = db.Column(db.DateTime)  # 受入日時
    cancelled_at = db.Column(db.DateTime)  # キャンセル日時
    received_by = db.Column(db.String(100))  # 受入者（IPアドレス）
    cancelled_by = db.Column(db.String(100))  # キャンセル者（IPアドレス）
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))

    @classmethod
    def record_receive(cls, order_number, item_name, spec1, quantity, client_ip, received_quantity=None):
        """受入を記録"""
        # 既存レコードを検索（発注番号+品名+仕様1+数量で一意）
        existing = cls.query.filter_by(
            order_number=order_number,
            item_name=item_name,
            spec1=spec1,
            quantity=quantity
        ).first()

        if existing:
            # 既存レコードを更新
            existing.is_received = True
            existing.received_at = datetime.now(timezone.utc)
            existing.received_by = client_ip
            existing.received_quantity = received_quantity if received_quantity is not None else quantity
            existing.cancelled_at = None
            existing.cancelled_by = None
        else:
            # 新規レコードを作成
            history = cls(
                order_number=order_number,
                item_name=item_name,
                spec1=spec1,
                quantity=quantity,
                received_quantity=received_quantity if received_quantity is not None else quantity,
                is_received=True,
                received_at=datetime.now(timezone.utc),
                received_by=client_ip
            )
            db.session.add(history)

        db.session.commit()

    @classmethod
    def record_cancel(cls, order_number, item_name, spec1, quantity, client_ip):
        """受入キャンセルを記録"""
        existing = cls.query.filter_by(
            order_number=order_number,
            item_name=item_name,
            spec1=spec1,
            quantity=quantity
        ).first()

        if existing:
            existing.is_received = False
            existing.cancelled_at = datetime.now(timezone.utc)
            existing.cancelled_by = client_ip
            db.session.commit()

    @classmethod
    def is_received_in_history(cls, order_number, item_name, spec1, quantity):
        """履歴から受入状態を確認"""
        existing = cls.query.filter_by(
            order_number=order_number,
            item_name=item_name,
            spec1=spec1,
            quantity=quantity,
            is_received=True
        ).first()
        return existing is not None

    @classmethod
    def get_received_info(cls, order_number, item_name, spec1, quantity):
        """履歴から受入情報を取得"""
        return cls.query.filter_by(
            order_number=order_number,
            item_name=item_name,
            spec1=spec1,
            quantity=quantity,
            is_received=True
        ).first()


class EditLog(db.Model):
    """編集ログテーブル"""
    id = db.Column(db.Integer, primary_key=True)
    detail_id = db.Column(db.Integer, db.ForeignKey('order_detail.id'))
    action = db.Column(db.String(50))  # 'receive', 'unreceive'
    ip_address = db.Column(db.String(45))
    timestamp = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    user_agent = db.Column(db.String(500))


class PartCategory(db.Model):
    """部品分類記号マスタテーブル - N**-形式の部品番号の分類情報"""
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(10), unique=True, nullable=False, index=True)  # NAA, NAB, NFA等（3文字）
    major_category = db.Column(db.String(50))   # 大分類
    minor_category = db.Column(db.String(50))   # 小分類
    note = db.Column(db.String(500))            # 補足

    @classmethod
    def get_category_info(cls, part_code):
        """部品コードから分類情報を取得（NAA-00123-01-00 → NAA で検索）"""
        if not part_code or len(part_code) < 3:
            return None
        category_code = part_code[:3].upper()
        return cls.query.filter_by(code=category_code).first()

    @classmethod
    def parse_part_number(cls, part_code):
        """部品番号をパースして詳細情報を返す
        形式: NAA-00000-00-00
              ^^^-^^^^^-^^-^^
              分類-シリアル-派生-リビジョン
        """
        if not part_code:
            return None

        parts = part_code.split('-')
        if len(parts) < 1:
            return None

        result = {
            'category_code': parts[0] if len(parts) > 0 else None,
            'serial': parts[1] if len(parts) > 1 else None,
            'derivative': parts[2] if len(parts) > 2 else None,
            'revision': parts[3] if len(parts) > 3 else None,
        }

        # 分類情報を取得
        if result['category_code']:
            category = cls.query.filter_by(code=result['category_code']).first()
            if category:
                result['major_category'] = category.major_category
                result['minor_category'] = category.minor_category
                result['note'] = category.note

        return result


class UserSettings(db.Model):
    """ユーザー設定テーブル - クライアントIPアドレスをキーに設定を保存"""
    id = db.Column(db.Integer, primary_key=True)
    client_ip = db.Column(db.String(45), unique=True, nullable=False, index=True)  # IPv6対応

    # 受入モード設定
    simple_mode = db.Column(db.Boolean, default=False)  # シンプルモード（箱QRスキャン時に未受入部品リストを表示）

    # 表示設定
    view_mode = db.Column(db.String(20), default='card')  # card / table

    # その他の設定（JSON形式で拡張可能）
    settings_json = db.Column(db.Text, default='{}')

    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))

    @classmethod
    def get_or_create(cls, client_ip):
        """クライアントIPで設定を取得、なければ作成"""
        settings = cls.query.filter_by(client_ip=client_ip).first()
        if not settings:
            settings = cls(client_ip=client_ip)
            db.session.add(settings)
            db.session.commit()
        return settings

    @classmethod
    def get_settings(cls, client_ip):
        """設定を辞書形式で取得"""
        settings = cls.get_or_create(client_ip)
        extra_settings = {}
        if settings.settings_json:
            try:
                extra_settings = json.loads(settings.settings_json)
            except:
                pass
        return {
            'simple_mode': settings.simple_mode,
            'view_mode': settings.view_mode,
            **extra_settings
        }

    @classmethod
    def update_settings(cls, client_ip, **kwargs):
        """設定を更新"""
        settings = cls.get_or_create(client_ip)

        # 基本設定を更新
        if 'simple_mode' in kwargs:
            settings.simple_mode = kwargs.pop('simple_mode')
        if 'view_mode' in kwargs:
            settings.view_mode = kwargs.pop('view_mode')

        # 残りの設定はJSONに保存
        if kwargs:
            try:
                extra = json.loads(settings.settings_json or '{}')
            except:
                extra = {}
            extra.update(kwargs)
            settings.settings_json = json.dumps(extra)

        db.session.commit()
        return settings


# 分類記号マスタの初期データ
PART_CATEGORY_INITIAL_DATA = [
    ('NAA', '角ブロック', 'スペーサブロック', '主に角型ブロック（円筒形状中心穴はカラー）'),
    ('NAB', '角ブロック', 'スライドブロック', ''),
    ('NAC', '角ブロック', '本体ブロック', '角ストッパー含む'),
    ('NAD', '角ブロック', 'ねじブロック', ''),
    ('NAE', '角ブロック', '充填機シール板', ''),
    ('NAF', '角ブロック', 'ブロックダイ', ''),
    ('NBA', 'ばね', '板', ''),
    ('NBB', 'ばね', 'コイル', ''),
    ('NBC', 'ばね', 'トーション', ''),
    ('NBD', 'ばね', 'さらばね', ''),
    ('NCA', '架台／フレーム', '製缶', 'ユニット用サブ架台含む'),
    ('NCB', '架台／フレーム', '鋳物', ''),
    ('NCC', '架台／フレーム', '板金', ''),
    ('NCD', '架台／フレーム', '型鋼・パイプ', ''),
    ('NDA', 'ベース板/プレート', 'ｔ６以上', '一枚板で曲げ加工等無い物で、溶接等で他の部品が付いても良い'),
    ('NDB', 'ベース板/プレート', 't６未満', ''),
    ('NDC', 'ベース板/プレート', 'ｔ６以上', '一枚板で曲げ加工した物、溶接等で他の部品が付いても良い'),
    ('NDD', 'ベース板/プレート', 't６未満', ''),
    ('NDE', 'ベース板/プレート', 'フレキシブル刃', ''),
    ('NDF', 'ベース板/プレート', 'トムソン刃', ''),
    ('NEA', '配管／ホッパー', '一般配管', '空油圧'),
    ('NEB', '配管／ホッパー', 'サニタリー配管', ''),
    ('NEC', '配管／ホッパー', '配管保護', ''),
    ('NED', '配管／ホッパー', 'ホッパー', ''),
    ('NEE', '配管／ホッパー', 'ダクト', ''),
    ('NFA', 'シリンダ', 'チューブ', 'ヘッド側・ロッド側カバー一体含む'),
    ('NFB', 'シリンダ', 'ヘッド側・ロッド側カバー', '単体'),
    ('NFC', 'シリンダ', 'ピストン', 'ロッド一体含む'),
    ('NFD', 'シリンダ', 'エアピッカー', ''),
    ('NGA', 'カップリング', '製作物', ''),
    ('NGB', 'カップリング', '購入品追加工', ''),
    ('NGC', 'カップリング', '熱電対', ''),
    ('NHA', 'マニホールド', '', '空油圧用'),
    ('NIA', 'カバー・ガイド類', '金属', ''),
    ('NIB', 'カバー・ガイド類', '樹脂', ''),
    ('NIC', 'カバー・ガイド類', 'ハンドル', ''),
    ('NID', 'カバー・ガイド類', '内袋製品図', ''),
    ('NIE', 'カバー・ガイド類', 'PID外容器', '紙容器・パウチ・ホルダー・etc'),
    ('NIF', 'カバー・ガイド類', '印刷データ', 'フィルム・紙容器・パウチ・ホルダー・etc'),
    ('NJA', 'カム（ドグ類含む）', '円板カム', '円筒径方向がカム形状'),
    ('NJB', 'カム（ドグ類含む）', '円筒カム', '円筒軸方向がカム形状'),
    ('NJC', 'カム（ドグ類含む）', '板カム', '非回転型カム'),
    ('NKA', '取付金具', '単品取付け用', 'モーター・センサー等'),
    ('NKB', '取付金具', 'ユニット取付け用', ''),
    ('NLA', 'キー', '固定キー', ''),
    ('NLB', 'キー', 'すべりキー', ''),
    ('NMA', 'ギア', '平。はすば歯車', '製作物・購入品追加工'),
    ('NMB', 'ギア', '内歯車', ''),
    ('NMC', 'ギア', 'ラック', ''),
    ('NMD', 'ギア', 'マイタ・かさ・ハイポイド歯車', ''),
    ('NME', 'ギア', 'ウォーム・ウォームホイール', ''),
    ('NNA', '軸（円筒・多角形）', '支点ピン・固定ピン', '短尺物'),
    ('NNB', '軸（円筒・多角形）', 'ストレート円筒軸', 'ピストンロッド含む(外径公差精度級）'),
    ('NNC', '軸（円筒・多角形）', '回転用多段軸', '取付物回転含む'),
    ('NND', '軸（円筒・多角形）', '丸棒・六角棒加工品', ''),
    ('NNE', '軸（円筒・多角形）', 'テンションロッド', ''),
    ('NNF', '軸（円筒・多角形）', 'パンチ', ''),
    ('NOA', 'ねじ類', 'すべりねじ', 'JIS規格品含む'),
    ('NOB', 'ねじ類', 'ボールねじ', ''),
    ('NOC', 'ねじ類', 'スプライン', 'ボールスプライン含む'),
    ('NPA', 'ナット', 'すべりねじ', 'JIS規格品含む'),
    ('NPB', 'ナット', 'ボールねじナット（単独）', ''),
    ('NPC', 'ナット', 'スプラインナット（単体）', 'ボールスプライン含む'),
    ('NPD', 'ナット', '板ナット（複数穴含）', ''),
    ('NQA', 'カラー/座金', '座金', '板厚＜外径'),
    ('NQB', 'カラー/座金', 'カラー', '板厚≧外径'),
    ('NRA', '軸受け関連', 'ケース', ''),
    ('NRB', '軸受け関連', 'すべり軸受け', ''),
    ('NRC', '軸受け関連', '転がり軸受け', ''),
    ('NRD', '軸受け関連', 'ベアリングフタ／固定部品', ''),
    ('NSA', 'プーリー／スプロケット', 'スプロケット', ''),
    ('NSB', 'プーリー／スプロケット', '平プーリー', '面長＜外径'),
    ('NSC', 'プーリー／スプロケット', 'タイミングプーリー', ''),
    ('NSD', 'プーリー／スプロケット', 'Vプーリー', '丸ベルト用含む'),
    ('NSE', 'プーリー／スプロケット', 'ローラー', '金属・ゴム・樹脂ローラー　面長＞外径'),
    ('NSF', 'プーリー／スプロケット', 'コンベア', 'コンベア追加工'),
    ('NTA', 'レバー・リンク・ロープ', 'レバー', ''),
    ('NTB', 'レバー・リンク・ロープ', 'リンク', ''),
    ('NTC', 'レバー・リンク・ロープ', 'ロープ', ''),
    ('NUA', 'アーム類', 'アーム', ''),
    ('NVA', 'フランジ類', 'フランジ', ''),
    ('NWA', 'ブラシ類', '円筒形状', ''),
    ('NWB', 'ブラシ類', '平、角形状', ''),
    ('NXA', '銘板、刻印', '銘板', ''),
    ('NXB', '銘板、刻印', '表示シール', ''),
    ('NXC', '銘板、刻印', '刻印', ''),
    ('NYA', '空圧・油圧回路図', '', ''),
    ('NZA', '電気回路図', '', ''),
    ('AAA', 'Type-VＡｓｓｙ図', '', ''),
    ('AAB', 'Type-3アンプルカット改造', '', ''),
    ('AAC', '紙エコパックシール機', '', ''),
    ('AAD', 'Type-3アンプルカット改造', '', ''),
    ('AAE', 'ＰＩＤ自動充填機ライン（久原）', '', ''),
    ('AAF', 'PID小口生産システム', '', ''),
    ('AAG', 'PID注出試験機', '', ''),
    ('AAH', 'PID充填機W600(NEO)', '', ''),
    ('AAI', 'Type-G大容量アンプルカット改造', '', ''),
    ('AAJ', 'PIDパウチ半自動シール機', '', ''),
    ('AAK', '袋切断機（W600)', '', ''),
    ('AAL', 'CTカートン半自動シール機', '', ''),
    ('AAM', 'FSS(充填支援システム)', '', ''),
    ('AAN', 'CTパウチ半自動シール機', '', ''),
    ('AAO', 'CTカートン自動化検証機', '', ''),
    ('AAP', '小袋包装機', '', ''),
    ('AAQ', 'オートスプライサー', '', ''),
    ('AAR', 'SPG', '', ''),
    ('AAS', 'ドラム冷却装置', '', ''),
    ('AAT', '充填機用ポンプユニット', '', ''),
    ('AAU', '配管洗浄機（STC）', '', ''),
]


# ===== 枝番関連ヘルパー関数 =====
def get_parent_seiban(seiban):
    """製番から親製番を抽出（枝番を除去）
    MHT0620-001 → MHT0620, 620-008 → 620
    枝番でない場合はNoneを返す
    """
    if not seiban:
        return None
    match = re.match(r'^(.+?)-\d+$', seiban)
    return match.group(1) if match else None


def get_seiban_family(seiban):
    """製番とその枝番ファミリーをすべて取得
    親製番を入力 → 親 + すべての枝番を返す
    枝番を入力 → 親 + すべての枝番を返す
    """
    if not seiban:
        return []

    # 枝番の場合、親製番を取得
    parent = get_parent_seiban(seiban)
    if parent:
        base_seiban = parent
    else:
        base_seiban = seiban

    # 親製番 + 枝番パターンで検索
    pattern = f"{base_seiban}%"

    # DBから該当する製番を取得
    orders = Order.query.filter(
        (Order.seiban == base_seiban) |
        (Order.seiban.like(f"{base_seiban}-%"))
    ).filter(
        Order.is_archived == False
    ).all()

    # 製番リストを作成（重複除去）
    seibans = list(set([o.seiban for o in orders]))

    # ソート（親製番が先、枝番は番号順）
    def sort_key(s):
        if s == base_seiban:
            return (0, 0)  # 親製番は最初
        branch_match = re.match(rf'^{re.escape(base_seiban)}-(\d+)$', s)
        if branch_match:
            return (1, int(branch_match.group(1)))
        return (2, s)

    return sorted(seibans, key=sort_key)


# Initialize database
with app.app_context():
    db.create_all()
    # マイグレーション: reply_delivery_date カラム追加
    try:
        with db.engine.connect() as conn:
            conn.execute(db.text("ALTER TABLE order_detail ADD COLUMN reply_delivery_date VARCHAR(20)"))
            conn.commit()
        print("✓ reply_delivery_date カラムを追加しました")
    except Exception:
        pass  # 既に存在する場合は無視

    # マイグレーション: received_quantity カラム追加 (OrderDetail)
    try:
        with db.engine.connect() as conn:
            conn.execute(db.text("ALTER TABLE order_detail ADD COLUMN received_quantity INTEGER"))
            conn.commit()
        print("✓ order_detail.received_quantity カラムを追加しました")
    except Exception:
        pass  # 既に存在する場合は無視

    # マイグレーション: received_quantity カラム追加 (ReceivedHistory)
    try:
        with db.engine.connect() as conn:
            conn.execute(db.text("ALTER TABLE received_history ADD COLUMN received_quantity INTEGER"))
            conn.commit()
        print("✓ received_history.received_quantity カラムを追加しました")
    except Exception:
        pass  # 既に存在する場合は無視

    # 分類記号マスタの初期データ投入
    try:
        if PartCategory.query.count() == 0:
            for code, major, minor, note in PART_CATEGORY_INITIAL_DATA:
                category = PartCategory(
                    code=code,
                    major_category=major,
                    minor_category=minor,
                    note=note
                )
                db.session.add(category)
            db.session.commit()
            print(f"✓ 分類記号マスタに{len(PART_CATEGORY_INITIAL_DATA)}件の初期データを投入しました")
    except Exception as e:
        print(f"分類記号マスタ初期化エラー: {e}")
        db.session.rollback()

def to_jst(utc_dt):
    """UTC時刻をJSTに変換"""
    if utc_dt is None:
        return None
    if utc_dt.tzinfo is None:
        # naive datetimeの場合、UTCとして扱う
        utc_dt = pytz.utc.localize(utc_dt)
    jst = pytz.timezone('Asia/Tokyo')
    return utc_dt.astimezone(jst)

# Utility Functions
def get_cad_file_info(spec1):
    """仕様1からCADファイル情報を取得"""
    if not spec1 or not spec1.startswith('N'):
        return None
    
    # NKA-00437-00-00 → K を抽出
    parts = spec1.split('-')
    if len(parts) < 2 or len(parts[0]) < 2:
        return None
    
    # 2文字目のアルファベットを取得
    folder_letter = parts[0][1].upper()
    
    # CADフォルダパス
    cad_folder = f"\\\\SERVER3\\Share-data\\CadData\\Parts\\{folder_letter}"
    
    # ファイルを検索（ワイルドカード）
    # 例: NKA-00437-00-00*.mx2 または NKA-00437-00-00*.pdf
    mx2_pattern = os.path.join(cad_folder, f"{spec1}*.mx2")
    pdf_pattern = os.path.join(cad_folder, f"{spec1}*.pdf")
    
    mx2_files = glob.glob(mx2_pattern)
    pdf_files = glob.glob(pdf_pattern)
    
    if not mx2_files and not pdf_files:
        return None
    
    return {
        'folder': cad_folder,
        'letter': folder_letter,
        'spec1': spec1,
        'mx2_files': mx2_files,
        'pdf_files': pdf_files,
        'has_mx2': len(mx2_files) > 0,
        'has_pdf': len(pdf_files) > 0
    }

def load_seiban_info():
    """製番情報を取得（V_D受注DBから取得、フォールバックでExcel）"""
    try:
        # まずDBから取得を試みる
        result = across_db.get_seiban_list_from_db()
        if result.get('success') and result.get('items'):
            # DB結果を辞書形式に変換
            seiban_info = {}
            for item in result['items']:
                seiban = item.get('seiban', '')
                if seiban:
                    seiban_info[seiban] = {
                        'product_name': item.get('product_name', ''),
                        'customer_abbr': item.get('customer_name', ''),  # customer_name → customer_abbr
                        'memo2': item.get('memo2', '')
                    }
            print(f"製番情報をDBから取得: {len(seiban_info)}件")
            return seiban_info
    except Exception as e:
        print(f"DB取得エラー、Excelにフォールバック: {str(e)}")

    # DBが使えない場合はExcelから読み込み（フォールバック）
    try:
        seiban_file = app.config.get('SEIBAN_LIST_PATH', r'\\server3\share-data\Document\Acrossデータ\製番一覧表.xlsx')
        seiban_path = Path(seiban_file)

        if not seiban_path.exists():
            print(f"製番一覧表が見つかりません: {seiban_path}")
            return {}

        # Excelファイルを読み込み
        df = pd.read_excel(str(seiban_path), sheet_name='製番')

        # 製番と情報の辞書を作成
        seiban_info = {}
        for _, row in df.iterrows():
            if pd.notna(row.get('製番')):
                seiban_info[str(row['製番'])] = {
                    'product_name': str(row.get('品名', '')) if pd.notna(row.get('品名')) else '',
                    'customer_abbr': str(row.get('得意先略称', '')) if pd.notna(row.get('得意先略称')) else '',
                    'memo2': str(row.get('メモ２', '')) if pd.notna(row.get('メモ２')) else ''
                }

        print(f"製番情報をExcelから取得: {len(seiban_info)}件")
        return seiban_info
    except Exception as e:
        print(f"製番一覧表読み込みエラー: {str(e)}")
        return {}


def extract_seiban_from_filename(filename):
    """Extract seiban (MHTxxxx) from filename"""
    pattern = r'(MHT\d{4})'
    match = re.search(pattern, filename)
    return match.group(1) if match else None

def process_excel_file(file_path, sheet1_name, sheet2_name, seiban_prefix, order_date_from=None, order_date_to=None):
    """Process Excel file and merge data"""
    try:
        df1 = pd.read_excel(file_path, sheet_name=sheet1_name, header=0)
        df2 = pd.read_excel(file_path, sheet_name=sheet2_name, header=0)
        
        # デバッグ情報
        print(f"=== デバッグ情報 ===")
        print(f"検索製番: {seiban_prefix}")
        print(f"シート1名: {sheet1_name}, 件数: {len(df1)}件")
        print(f"シート2名: {sheet2_name}, 件数: {len(df2)}件")
        
        # 製番の前処理（空白除去）
        df1['製番'] = df1['製番'].astype(str).str.strip()
        df2['製番'] = df2['製番'].astype(str).str.strip()
        
        # 製番でフィルタリング
        df1 = df1[df1['製番'].str.startswith(seiban_prefix, na=False)]
        df2 = df2[df2['製番'].str.startswith(seiban_prefix, na=False)]
        
        print(f"製番フィルタ({seiban_prefix})後: シート1={len(df1)}件, シート2={len(df2)}件")
        print(f"===================")
        
        return process_excel_file_from_dataframes(df1, df2, seiban_prefix, order_date_from, order_date_to)
        
    except Exception as e:
        print(f"エラー詳細: {e}")
        import traceback
        traceback.print_exc()
        raise Exception(f"Error processing Excel: {str(e)}")

def process_excel_file_from_dataframes(df1, df2, seiban_prefix, order_date_from=None, order_date_to=None):
    """Process dataframes and merge data"""
    try:
        df1 = df1.fillna('')
        df2 = df2.fillna('')
        
        if (order_date_from or order_date_to) and '発注日' in df2.columns:
            print(f"発注日フィルタ適用:")
            df2['発注日'] = pd.to_datetime(df2['発注日'], errors='coerce')
            before_count = len(df2)
            
            if order_date_from:
                filter_date_from = pd.to_datetime(order_date_from)
                df2 = df2[df2['発注日'] >= filter_date_from]
                print(f"  開始日: {order_date_from}以降")
            
            if order_date_to:
                filter_date_to = pd.to_datetime(order_date_to)
                df2 = df2[df2['発注日'] <= filter_date_to]
                print(f"  終了日: {order_date_to}まで")
            
            after_count = len(df2)
            print(f"発注日フィルタ: {before_count}件 → {after_count}件 ({before_count - after_count}件除外)")
        
        if '発注番号' in df1.columns:
            df1['発注番号'] = df1['発注番号'].apply(lambda x: str(int(float(x))) if isinstance(x, (int, float)) and x != '' else str(x))
        if '発注番号' in df2.columns:
            df2['発注番号'] = df2['発注番号'].apply(lambda x: str(int(float(x))) if isinstance(x, (int, float)) and x != '' else str(x))
        
        if '手配区分CD' in df1.columns:
            df1['手配区分CD'] = df1['手配区分CD'].apply(lambda x: str(int(float(x))) if isinstance(x, (int, float)) and x != '' else str(x))
        if '手配区分CD' in df2.columns:
            df2['手配区分CD'] = df2['手配区分CD'].apply(lambda x: str(int(float(x))) if isinstance(x, (int, float)) and x != '' else str(x))
        
        # 仕入先CDも文字列に変換
        if '仕入先CD' in df2.columns:
            df2['仕入先CD'] = df2['仕入先CD'].apply(lambda x: str(int(float(x))) if isinstance(x, (int, float)) and x != '' else str(x))
        
        if '納期' in df2.columns:
            df2.loc[df2['納期'] != '', '納期'] = pd.to_datetime(
                df2.loc[df2['納期'] != '', '納期'], 
                errors='coerce'
            )
        
        # 仕入先CD用の列を初期化
        if '仕入先CD' not in df1.columns:
            df1['仕入先CD'] = ''
        
        # Merge data
        for i, row in df1.iterrows():
            matched = False
            
            if all(col in row for col in ['材質', '仕様１', '製番']):
                if row['材質'] and row['仕様１'] and row['製番']:
                    cond = ((df2['材質'] == row['材質']) &
                           (df2['仕様１'] == row['仕様１']) &
                           (df2['製番'] == row['製番']))
                    match = df2[cond]
                    if not match.empty:
                        matched = True
                        if '発注番号' in match.columns:
                            df1.at[i, '発注番号'] = str(match['発注番号'].iloc[0])
                        if '仕入先略称' in match.columns:
                            df1.at[i, '仕入先略称'] = match['仕入先略称'].iloc[0]
                        if '仕入先CD' in match.columns:
                            df1.at[i, '仕入先CD'] = str(match['仕入先CD'].iloc[0])
                        if '納期' in match.columns:
                            try:
                                if pd.notna(match['納期'].iloc[0]) and match['納期'].iloc[0] != '':
                                    df1.at[i, '納期'] = match['納期'].dt.strftime('%y/%m/%d').iloc[0]
                            except:
                                df1.at[i, '納期'] = str(match['納期'].iloc[0])
            
            if not matched and all(col in row for col in ['製番', '仕様１', '手配区分']):
                if row['製番'] and row['仕様１']:
                    cond = ((df2['製番'] == row['製番']) &
                           (df2['仕様１'] == row['仕様１']))
                    if row['手配区分'] and '手配区分' in df2.columns:
                        cond = cond & (df2['手配区分'] == row['手配区分'])
                    
                    match = df2[cond]
                    if not match.empty:
                        if '発注番号' in match.columns:
                            df1.at[i, '発注番号'] = str(match['発注番号'].iloc[0])
                        if '仕入先略称' in match.columns:
                            df1.at[i, '仕入先略称'] = match['仕入先略称'].iloc[0]
                        if '仕入先CD' in match.columns:
                            df1.at[i, '仕入先CD'] = str(match['仕入先CD'].iloc[0])
                        if '納期' in match.columns:
                            try:
                                if pd.notna(match['納期'].iloc[0]) and match['納期'].iloc[0] != '':
                                    df1.at[i, '納期'] = match['納期'].dt.strftime('%y/%m/%d').iloc[0]
                            except:
                                df1.at[i, '納期'] = str(match['納期'].iloc[0])
        
        # Reorder columns（仕入先CDを含める - DB保存用）
        cols = ['納期', '仕入先略称', '仕入先CD', '発注番号', '手配数', '単位', '品名', '仕様１', '仕様２',
                '品目CD', '手配区分CD', '手配区分', 'メーカー', '備考', '員数', '必要数', '製番', '材質']
        cols = [c for c in cols if c in df1.columns]
        df1 = df1[cols]
        
        if '発注番号' in df1.columns:
            df1_with_order = df1[df1['発注番号'] != '']
            df1_without_order = df1[df1['発注番号'] == '']
            
            if not df1_with_order.empty:
                df1_with_order = df1_with_order.sort_values('発注番号')
            
            df1 = pd.concat([df1_with_order, df1_without_order], ignore_index=True)
        
        return df1
    except Exception as e:
        raise Exception(f"Error processing dataframes: {str(e)}")
    
def create_order_detail_with_parts(row, order, all_received_items, safe_str, safe_int):
    """OrderDetail作成"""
    import pandas as pd
    
    order_type = safe_str(row.get('手配区分', ''))
    has_internal = '社内加工' in order_type or '追加工' in order_type
    
    # 仕入先CDを取得
    supplier_cd = row.get('仕入先CD', '')
    if isinstance(supplier_cd, (int, float)) and not pd.isna(supplier_cd):
        supplier_cd = str(int(supplier_cd))
    else:
        supplier_cd = safe_str(supplier_cd)
    
    detail = OrderDetail(
        order_id=order.id,
        delivery_date=safe_str(row.get('納期', '')),
        supplier=safe_str(row.get('仕入先略称', '')),
        supplier_cd=supplier_cd,  # 追加
        order_number=DataUtils.normalize_order_number(row.get('発注番号', '')),
        quantity=safe_int(row.get('手配数', 0)),
        unit_measure=safe_str(row.get('単位', '')),
        item_name=safe_str(row.get('品名', '')),
        spec1=safe_str(row.get('仕様１', '')),
        spec2=safe_str(row.get('仕様２', '')),
        item_code=safe_str(row.get('品目CD', '')),
        order_type_code=DataUtils.normalize_order_number(row.get('手配区分CD', '')),
        order_type=order_type,
        maker=safe_str(row.get('メーカー', '')),
        remarks=safe_str(row.get('備考', '')),
        member_count=safe_int(row.get('員数', 0)),
        required_count=safe_int(row.get('必要数', 0)),
        seiban=safe_str(row.get('製番', '')),
        material=safe_str(row.get('材質', '')).replace('-', ''),
        has_internal_processing=has_internal,
        part_number=safe_str(row.get('部品No', '')),
        page_number=safe_str(row.get('ページNo', '')),
        row_number=safe_str(row.get('行No', '')),
        hierarchy=safe_int(row.get('階層', 0)),
        reply_delivery_date=safe_str(row.get('回答納期', ''))
    )
    
    _restore_received_status(detail, all_received_items)
    return detail

def _restore_received_status(detail, all_received_items):
    """受入状態復元（既存データ優先、なければReceivedHistoryから復元）"""
    restored = False

    # 1. まず既存データ（同じ製番内）から復元を試みる
    if detail.order_number and detail.order_number in all_received_items:
        for received in all_received_items[detail.order_number]:
            if (received['item_name'] == detail.item_name and
                received['spec1'] == detail.spec1 and
                received['quantity'] == detail.quantity):
                detail.is_received = True
                detail.received_at = received['received_at']
                restored = True
                break

    # 2. 既存データで復元できなかった場合、ReceivedHistoryから復元
    if not restored and detail.order_number:
        history = ReceivedHistory.get_received_info(
            order_number=detail.order_number,
            item_name=detail.item_name,
            spec1=detail.spec1,
            quantity=detail.quantity
        )
        if history:
            detail.is_received = True
            detail.received_at = history.received_at
            print(f"✅ 受入履歴から復元: 発注番号={detail.order_number}, 品名={detail.item_name}")
            
def update_order_status(order):
    """注文ステータス更新"""
    if not order.details:
        return
    
    all_received = all(d.is_received for d in order.details)
    any_received = any(d.is_received for d in order.details)
    
    if all_received:
        order.status = Constants.STATUS_COMPLETED
    elif any_received:
        order.status = Constants.STATUS_IN_PROGRESS
    else:
        order.status = Constants.STATUS_BEFORE
    
    order.updated_at = datetime.now(timezone.utc)

def save_order_to_excel(order, filepath, data_filepath=None):
    """注文をExcelファイルに保存（dataフォルダに元データ保存→メインファイルにコピー）"""
    import shutil

    try:
        unit_display = order.unit if order.unit else 'ユニット名無し'
        sheet_name = f"{order.seiban}_{unit_display}"
        sheet_name = re.sub(r'[\\\/\?\*\[\]:]', '', sheet_name)[:31]

        # dataフォルダのパスを取得（指定がなければ自動生成）
        if data_filepath is None:
            data_filepath = get_order_excel_data_path(order.seiban, order.product_name, order.customer_abbr)

        # === Step 1: dataフォルダに元データを保存 ===
        if Path(data_filepath).exists():
            try:
                wb = load_workbook(data_filepath)
            except PermissionError:
                # dataフォルダのファイルも使用中の場合は新規作成
                wb = Workbook()
                wb.remove(wb.active)

            # 既存シートを削除
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

            # ガントチャートシートを削除して再作成
            if "納期ガントチャート" in wb.sheetnames:
                del wb["納期ガントチャート"]
        else:
            wb = Workbook()
            wb.remove(wb.active)

        # 全ユニットを取得してガントチャートを再作成
        orders = Order.query.filter_by(seiban=order.seiban, is_archived=False).all()
        create_gantt_chart_sheet(wb, order.seiban, orders)

        # 新しいシートを作成
        ws = wb.create_sheet(sheet_name)
        create_order_sheet(ws, order, sheet_name)

        # dataフォルダに保存
        wb.save(data_filepath)
        wb.close()
        print(f"✅ 元データ保存完了: {data_filepath}")

        # === Step 2: メインファイルにコピー（閲覧用） ===
        try:
            shutil.copy2(data_filepath, filepath)
            print(f"✅ メインファイル更新: {filepath}")
            return True, None
        except PermissionError:
            # メインファイルが使用中でもdataフォルダには保存済み
            print(f"⚠️ メインファイル使用中（元データは保存済み）: {filepath}")
            return True, "メインファイルは使用中ですが、元データは保存されました"

    except Exception as e:
        return False, str(e)
    
def get_order_excel_filename(seiban, product_name=None, customer_abbr=None):
    """製番に対応するExcelファイル名を取得（品名・客先名付き）"""
    # 品名と客先名をファイル名に含める（Windowsファイル名禁止文字を除去）
    if product_name:
        safe_product_name = re.sub(r'[\\/:*?"<>|]', '', product_name)
        if customer_abbr:
            safe_customer_abbr = re.sub(r'[\\/:*?"<>|]', '', customer_abbr)
            filename = f"{seiban}_{safe_product_name}_{safe_customer_abbr}_手配発注リスト.xlsx"
        else:
            filename = f"{seiban}_{safe_product_name}_手配発注リスト.xlsx"
    else:
        filename = f"{seiban}_手配発注リスト.xlsx"
    return filename

def get_order_excel_path(seiban, product_name=None, customer_abbr=None):
    """製番に対応するExcelファイルパスを取得（閲覧用メインファイル）"""
    export_dir = Path(app.config['EXPORT_EXCEL_PATH'])
    export_dir.mkdir(parents=True, exist_ok=True)
    filename = get_order_excel_filename(seiban, product_name, customer_abbr)
    return str(export_dir / filename)

def get_order_excel_data_path(seiban, product_name=None, customer_abbr=None):
    """製番に対応するExcelファイルパスを取得（元データ用dataフォルダ）"""
    export_dir = Path(app.config['EXPORT_EXCEL_PATH']) / 'data'
    export_dir.mkdir(parents=True, exist_ok=True)
    filename = get_order_excel_filename(seiban, product_name, customer_abbr)
    return str(export_dir / filename)
    
def update_order_excel(order_id):
    """注文IDに対応するExcelファイルを更新（同じ製番の全ユニットを再生成）"""
    import shutil
    try:
        order = db.session.get(Order, order_id)
        if not order:
            return False, "注文が見つかりません"

        # 品名と客先名を渡す
        filepath = get_order_excel_path(order.seiban, order.product_name, order.customer_abbr)
        data_filepath = get_order_excel_data_path(order.seiban, order.product_name, order.customer_abbr)

        # 同じ製番の全ユニットを取得
        all_orders = Order.query.filter_by(seiban=order.seiban, is_archived=False).all()

        # 新規ワークブック作成（全シート再生成）
        wb = Workbook()
        wb.remove(wb.active)

        # ガントチャートシート作成
        create_gantt_chart_sheet(wb, order.seiban, all_orders)

        # 全ユニットのシートを作成
        for unit_order in all_orders:
            unit_display = unit_order.unit if unit_order.unit else 'ユニット名無し'
            sheet_name = f"{unit_order.seiban}_{unit_display}"
            sheet_name = re.sub(r'[\\\/\?\*\[\]:]', '', sheet_name)[:31]

            ws = wb.create_sheet(sheet_name)
            create_order_sheet(ws, unit_order, sheet_name)

        # dataフォルダに保存
        Path(data_filepath).parent.mkdir(parents=True, exist_ok=True)
        wb.save(data_filepath)
        wb.close()
        print(f"✅ 全ユニットExcel保存完了: {data_filepath} ({len(all_orders)}シート)")

        # メインファイルにコピー
        try:
            shutil.copy2(data_filepath, filepath)
            print(f"✅ メインファイル更新: {filepath}")
            return True, None
        except PermissionError:
            print(f"⚠️ メインファイル使用中（元データは保存済み）: {filepath}")
            return True, "メインファイルは使用中ですが、元データは保存されました"

    except Exception as e:
        import traceback
        traceback.print_exc()
        return False, str(e)

def update_unit_excel_only(order_id):
    """受入処理用の軽量Excel更新 - 対象ユニットのシートのみ差し替え（ガントチャート・他ユニットはスキップ）"""
    import shutil
    try:
        order = db.session.get(Order, order_id)
        if not order:
            return False, "注文が見つかりません"

        data_filepath = get_order_excel_data_path(order.seiban, order.product_name, order.customer_abbr)
        filepath = get_order_excel_path(order.seiban, order.product_name, order.customer_abbr)

        unit_display = order.unit if order.unit else 'ユニット名無し'
        sheet_name = f"{order.seiban}_{unit_display}"
        sheet_name = re.sub(r'[\\\/\?\*\[\]:]', '', sheet_name)[:31]

        # 既存ワークブックを開く（なければフル再生成にフォールバック）
        if Path(data_filepath).exists():
            try:
                wb = load_workbook(data_filepath)
            except Exception:
                return update_order_excel(order_id)
        else:
            return update_order_excel(order_id)

        # 対象シートだけ削除して再作成
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws = wb.create_sheet(sheet_name)
        create_order_sheet(ws, order, sheet_name)

        # 保存
        wb.save(data_filepath)
        wb.close()
        print(f"✅ ユニットシート更新: {sheet_name}")

        # メインファイルにコピー
        try:
            shutil.copy2(data_filepath, filepath)
        except PermissionError:
            pass  # dataフォルダには保存済み

        return True, None

    except Exception as e:
        import traceback
        traceback.print_exc()
        return False, str(e)
    
def save_to_database(df, seiban_prefix):
    """Save processed data to database"""
    try:
        seiban_info = load_seiban_info()
        info = seiban_info.get(seiban_prefix, {})
        product_name = info.get('product_name', '')  # 品名を取得
        customer_abbr = info.get('customer_abbr', '')  # 客先名を取得
        
        df['材質'] = df['材質'].replace('', '-')
        materials = df['材質'].unique()
        
        cols_to_keep = ['部品No', 'ページNo', '行No', '階層']
        for col in cols_to_keep:
            if col not in df.columns:
                df[col] = ''
        
        all_received_items = {}
        existing_orders = Order.query.filter_by(seiban=seiban_prefix).all()
        for existing_order in existing_orders:
            for detail in existing_order.details:
                if detail.is_received and detail.order_number:
                    key = str(detail.order_number)
                    if key not in all_received_items:
                        all_received_items[key] = []
                    all_received_items[key].append({
                        'item_name': detail.item_name,
                        'spec1': detail.spec1,
                        'quantity': detail.quantity,
                        'is_received': True,
                        'received_at': detail.received_at
                    })
        
        def safe_str(value):
            if pd.isna(value) or value is None:
                return ''
            if isinstance(value, float) and value == value:
                try:
                    return str(int(value))
                except:
                    return str(value)
            return str(value)
        
        def safe_int(value, default=0):
            if pd.isna(value) or value is None:
                return default
            try:
                return int(float(value))
            except (ValueError, TypeError):
                return default
        
        for material in materials:
            material_df = df[df['材質'] == material]
            
            # 🔥 ユニット名を正規化（検索と作成で統一）
            unit_name = material if material and material != '-' else ''
            
            # 🔥 正規化されたユニット名で検索
            order = Order.query.filter_by(seiban=seiban_prefix, unit=unit_name).first()
            if not order:
                # 🔥 新規作成
                order = Order(
                    seiban=seiban_prefix, 
                    unit=unit_name,  # 正規化された値を使用
                    product_name=product_name,
                    customer_abbr=info.get('customer_abbr', ''),
                    memo2=info.get('memo2', '')
                )
                db.session.add(order)
                db.session.flush()
                print(f"✅ 新規ユニット作成: {seiban_prefix} - {unit_name or 'ユニット名無し'}")
            else:
                # 🔥 既存レコードを更新（remarks, image_path, location, pallet, statusは保持）
                order.product_name = product_name
                order.customer_abbr = info.get('customer_abbr', '')
                order.memo2 = info.get('memo2', '')
                # order.remarks / order.image_path / order.location / order.pallet_number / order.status は変更しない
                print(f"🔄 既存ユニット更新: {seiban_prefix} - {unit_name or 'ユニット名無し'} (ID: {order.id}, 備考保持: {'有' if order.remarks else '無'})")
            
            # 既存の詳細を削除して再作成
            OrderDetail.query.filter_by(order_id=order.id).delete()
            
            part_groups = {}
            
            for _, row in material_df.iterrows():
                part_no = safe_str(row.get('部品No', ''))
                page_no = safe_str(row.get('ページNo', ''))
                material_key = safe_str(row.get('材質', ''))
                
                group_key = (part_no, page_no, material_key)
                
                if group_key not in part_groups:
                    part_groups[group_key] = []
                part_groups[group_key].append(row)
            
            for group_key, rows in part_groups.items():
                part_no, page_no, material_key = group_key
                
                blanks = []
                processed = []
                others = []
                
                for row in rows:
                    order_type_code = safe_str(row.get('手配区分CD', ''))

                    # 手配区分CDが空欄のものは除外
                    if not order_type_code or order_type_code.strip() == '':
                        item_name = safe_str(row.get('品名', ''))
                        print(f"除外: {item_name} - 手配区分CDが空欄")
                        continue

                    if order_type_code == '13':
                        blanks.append(row)
                    elif order_type_code == '11':
                        processed.append(row)
                    else:
                        others.append(row)
                
                blanks.sort(key=lambda r: safe_int(r.get('行No', 0)))
                processed.sort(key=lambda r: safe_int(r.get('行No', 0)))
                
                if blanks or processed:
                    print(f"\nグループ: 部品No={part_no}, ページNo={page_no}")
                    print(f"追加工（親）候補: {len(processed)}個, ブランク（子）候補: {len(blanks)}個")
                
                used_blanks = set()

                # 🔥 追加工(11,階層1)が親 → ブランク(13,階層2)が子
                # ルール: 追加工の行No < ブランクの行No <= 追加工の行No+300, 階層差=+1
                for proc_row in processed:
                    proc_row_no = safe_int(proc_row.get('行No', 0))
                    proc_hierarchy = safe_int(proc_row.get('階層', 0))

                    matching_blank = None

                    for i, blank_row in enumerate(blanks):
                        if i in used_blanks:
                            continue

                        blank_row_no = safe_int(blank_row.get('行No', 0))
                        blank_hierarchy = safe_int(blank_row.get('階層', 0))

                        # ブランクの行Noが追加工より大きく、300以内、階層が+1
                        if (blank_row_no > proc_row_no and
                            blank_row_no <= proc_row_no + 300 and
                            blank_hierarchy == proc_hierarchy + 1):
                            matching_blank = (i, blank_row)
                            break

                    proc_name = safe_str(proc_row.get('品名', ''))

                    if matching_blank is not None:
                        blank_idx, blank_row = matching_blank
                        used_blanks.add(blank_idx)

                        # 追加工(11)が親
                        parent_detail = create_order_detail_with_parts(
                            proc_row, order, all_received_items, safe_str, safe_int
                        )
                        db.session.add(parent_detail)
                        db.session.flush()

                        # ブランク(13)が子
                        child_detail = create_order_detail_with_parts(
                            blank_row, order, all_received_items, safe_str, safe_int
                        )
                        child_detail.parent_id = parent_detail.id
                        db.session.add(child_detail)

                        blank_name = safe_str(blank_row.get('品名', ''))
                        blank_row_no = safe_int(blank_row.get('行No', 0))
                        blank_hierarchy = safe_int(blank_row.get('階層', 0))

                        print(f"親子設定: 親・追加工({proc_name[:15]}, 行No={proc_row_no}, 階層={proc_hierarchy}) "
                              f"→ 子・ブランク({blank_name[:15]}, 行No={blank_row_no}, 階層={blank_hierarchy})")
                    else:
                        # 対応するブランクがない追加工は単独で保存
                        proc_detail = create_order_detail_with_parts(
                            proc_row, order, all_received_items, safe_str, safe_int
                        )
                        db.session.add(proc_detail)
                        print(f"追加工のみ: {proc_name[:15]} (行No={proc_row_no}) - 対応するブランクなし")

                # 未マッチのブランクを単独保存
                for i, blank_row in enumerate(blanks):
                    if i not in used_blanks:
                        blank_detail = create_order_detail_with_parts(
                            blank_row, order, all_received_items, safe_str, safe_int
                        )
                        db.session.add(blank_detail)
                        blank_name = safe_str(blank_row.get('品名', ''))
                        blank_row_no = safe_int(blank_row.get('行No', 0))
                        print(f"ブランクのみ: {blank_name[:15]} (行No={blank_row_no}) - 対応する追加工なし")
                
                for row in others:
                    order_type_code = safe_str(row.get('手配区分CD', ''))
                    spec1 = safe_str(row.get('仕様１', ''))
                    
                    if order_type_code == '15' and re.match(r'^M\d', spec1):
                        item_name = safe_str(row.get('品名', ''))
                        print(f"除外: {item_name} ({spec1}) - 在庫部品のM+数値")
                        continue
                    
                    detail = create_order_detail_with_parts(
                        row, order, all_received_items, safe_str, safe_int
                    )
                    db.session.add(detail)
        
        for order in Order.query.filter_by(seiban=seiban_prefix).all():
            update_order_status(order)

        db.session.commit()
                
        try:
            filepath = get_order_excel_path(seiban_prefix, product_name, customer_abbr)
            
            if filepath:
                orders = Order.query.filter_by(seiban=seiban_prefix, is_archived=False).all()
                
                if orders:
                    wb = Workbook()
                    wb.remove(wb.active)  # デフォルトシートを削除
                    
                    # 🔥 1シート目: ガントチャート
                    create_gantt_chart_sheet(wb, seiban_prefix, orders)
                    
                    # 2シート目以降: 各ユニットの手配リスト
                    for order in orders:
                        unit = order.unit if order.unit else 'ユニット名無し'
                        sheet_name = f"{seiban_prefix}_{unit}"[:31]
                        ws = wb.create_sheet(title=sheet_name)
                        create_order_sheet(ws, order, sheet_name)
                    
                    wb.save(filepath)
                    wb.close()
                    print(f"✅ Excel自動出力成功（ガントチャート付き）: {filepath}")
            else:
                print("⚠️  Excel出力パスの取得失敗")
        except Exception as excel_error:
            print(f"⚠️  Excel出力エラー（処理は継続）: {excel_error}")
            import traceback
            traceback.print_exc()
        
        return True
        
    except Exception as e:
        db.session.rollback()
        raise Exception(f"Database error: {str(e)}")

def get_server_url():
    """サーバーのURLを取得（IP + ポート）"""
    try:
        import socket
        # ホスト名からIPアドレスを取得
        hostname = socket.gethostname()
        ip_address = socket.gethostbyname(hostname)
        
        # 設定からHTTPSの使用状況を確認
        config_obj = get_config() if hasattr(app, 'config') else None
        use_https = False
        port = 8080
        
        if config_obj:
            use_https = getattr(config_obj, 'USE_HTTPS', False)
            port = getattr(config_obj, 'PORT', 8080)
        
        protocol = 'https' if use_https else 'http'
        return f"{protocol}://{ip_address}:{port}"
    except Exception as e:
        print(f"サーバーURL取得エラー: {e}")
        # フォールバック
        return "http://localhost:8080"
    
def create_order_sheet(ws, order, sheet_name=None):
    """ワークシート作成（縦向き印刷、QRコードH列配置）"""
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.drawing.image import Image
    from io import BytesIO
    import qrcode
    
    if sheet_name:
        ws.title = sheet_name
    
    # ヘッダー情報
    unit_display = order.unit if order.unit else 'ユニット名無し'
    customer = order.customer_abbr if order.customer_abbr else ''
    memo = order.memo2 if order.memo2 else ''
    product_name = order.product_name if order.product_name else ''
    
    # 🔥 QRコード生成（受入専用ページURL）- 製番/ユニットでURL固定
    try:
        from urllib.parse import quote
        server_url = get_server_url()
        unit_encoded = quote(order.unit, safe='') if order.unit else ''
        receive_url = f"{server_url}/receive/{order.seiban}/{unit_encoded}"
        
        # QRコード画像を生成
        qr = qrcode.QRCode(
            version=1,
            box_size=8,
            border=3
        )
        qr.add_data(receive_url)
        qr.make(fit=True)
        
        qr_img = qr.make_image(fill_color="black", back_color="white")
        
        # BytesIOに保存
        qr_buffer = BytesIO()
        qr_img.save(qr_buffer, format='PNG')
        qr_buffer.seek(0)
        
        # Excelに画像を挿入
        img = Image(qr_buffer)
        img.width = 100
        img.height = 100
        
        # 🔥 QRコードをH1セルに配置
        ws.add_image(img, 'I1')

        # 🔥 URLテキストとラベルをM列に配置（QRコードの右側）
        ws['M1'] = '💻️ 受入確認専用ページ(社内LANよりアクセス)'
        ws['M1'].font = Font(size=9, bold=True)
        ws['M1'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        ws['M2'] = receive_url
        ws['M2'].font = Font(size=8, color='0000FF', underline='single')
        ws['M2'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
    except Exception as e:
        print(f"⚠️ QRコード生成エラー: {e}")
    
    # 🔥 1行目: 4列に情報を配置
    # A1: 製番 + 品名 + 得意先 + メモ
    a1_text_parts = [order.seiban]
    if product_name:
        a1_text_parts.append(product_name)
    if customer:
        a1_text_parts.append(customer)
    if memo:
        a1_text_parts.append(memo)
    
    ws['A1'] = ' '.join(a1_text_parts)
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    
    # A2: ユニット名
    ws['A2'] = unit_display
    ws['A2'].font = Font(size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    # A3: 注意書き（赤字）
    ws['A3'] = '※ピンク塗は受入済 製番外の持ち出しは必ず記録を残すこと データは保存先にて随時更新'
    ws['A3'].font = Font(size=9, bold=True, color=Constants.COLOR_RED)
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')

    # A4: ネットワークパス（赤字）
    ws['A4'] = r'\\SERVER3\Share-data\Document\仕入れ\002_手配リスト\手配発注リスト'
    ws['A4'].font = Font(size=9, bold=True, color=Constants.COLOR_RED)
    ws['A4'].alignment = Alignment(horizontal='left', vertical='center')

    # 🔥 J2: 備考
    remarks_text = order.remarks if order.remarks else ''
    ws['J2'] = f'備考：{remarks_text}'
    ws['J2'].font = Font(size=9)
    ws['J2'].alignment = Alignment(horizontal='left', vertical='center')

    # 🔥 J3-K4: 保管場所情報
    ws['J3'] = '保管場所：'
    ws['J3'].font = Font(size=10, bold=True)
    ws['J3'].alignment = Alignment(horizontal='right', vertical='center')

    ws['K3'] = order.floor if order.floor else ''
    ws['K3'].font = Font(size=10)
    ws['K3'].alignment = Alignment(horizontal='left', vertical='center')

    ws['J4'] = '場所番号：'
    ws['J4'].font = Font(size=10, bold=True)
    ws['J4'].alignment = Alignment(horizontal='right', vertical='center')

    ws['K4'] = order.pallet_number if order.pallet_number else ''
    ws['K4'].font = Font(size=10)
    ws['K4'].alignment = Alignment(horizontal='left', vertical='center')

    # 🔥 行の高さ調整
    ws.row_dimensions[1].height = 35
    
    # 🔥 ヘッダー行（6行目）
    headers = Constants.EXCEL_COLUMNS
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color=Constants.COLOR_HEADER, 
                               end_color=Constants.COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 🔥 列幅設定（縦向き印刷用に最適化）
    column_widths = {
        'A': 9,   # 納入日（新規）
        'B': 6,   # 納入数（新規）
        'C': 9,   # 納期
        'D': 11,  # 仕入先略称
        'E': 9,   # 発注番号
        'F': 5,   # 手配数
        'G': 4,   # 単位
        'H': 18,  # 品名
        'I': 15,  # 仕様１
        'J': 12,  # 仕様２
        'K': 10,  # 手配区分
        'L': 8,   # メーカー
        'M': 12   # 備考
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 🔥 検収データを読み込み
    delivery_dict = DeliveryUtils.load_delivery_data()

    # 🔥 データ行を書き込む（7行目から開始）
    row_idx = 7
    parent_details = [d for d in order.details if d.parent_id is None]

    for detail in parent_details:
        row_idx = _write_detail_row(ws, detail, row_idx, is_parent=True, delivery_dict=delivery_dict)

        # 子アイテム
        children = [d for d in order.details if d.parent_id == detail.id]
        for child in children:
            row_idx = _write_detail_row(ws, child, row_idx, is_parent=False, delivery_dict=delivery_dict)
    
    # 🔥 ページ設定（縦向き印刷）
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # 🔥 余白を最小化
    ws.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.3,
        bottom=0.5,
        header=0.15,
        footer=0.2
    )
    
    # 🔥 印刷タイトル行（ヘッダーを毎ページ印刷）
    ws.print_title_rows = '1:6'
    ws.print_area = f'A1:M{row_idx - 1}'
    
    # 🔥 フッター設定（フォントサイズ10に縮小）
    footer_parts = []
    
    # 製番は必須
    footer_parts.append(order.seiban)
    
    # ユニット名
    if unit_display and unit_display != 'ユニット名無し':
        footer_parts.append(unit_display)
    
    # 品名（長すぎる場合は省略）
    if product_name:
        display_product_name = product_name if len(product_name) <= 20 else product_name[:20] + '...'
        footer_parts.append(display_product_name)
    
    # 得意先略称
    if customer:
        footer_parts.append(customer)
    
    # メモ２
    if memo:
        footer_parts.append(memo)
    
    footer_text = '_'.join(footer_parts)
    
    # 🔥 フッター設定（フォントサイズを10に変更）
    for footer in [ws.oddFooter, ws.evenFooter, ws.firstFooter]:
        footer.left.text = f"&10&B{footer_text}"  # &10でフォントサイズ10
        footer.center.text = "&P / &N"
        footer.right.text = f"&10&B{footer_text}"
    
    # 🔥 改ページプレビュー表示
    ws.sheet_view.view = 'pageBreakPreview'
    
    return ws

def _setup_page_settings(ws):
    """ページ設定"""
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.2, bottom=1.2, header=0.2, footer=0.1)

def _create_header_rows(ws, order, unit_display, customer, memo):
    """ヘッダー行作成"""
    header_parts = [f"{order.seiban}({order.product_name})"]
    if unit_display:
        header_parts.append(unit_display)
    if customer:
        header_parts.append(customer)
    if memo:
        header_parts.append(memo)
    
    ws['A1'] = '_'.join(header_parts) + "_"
    ws['A1'].font = Font(size=26, bold=True)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['A2'] = '※ピンク塗は受入済　製番外の持ち出しは必ず記録を残すこと　データは保存先にて随時更新: \\\\SERVER3\\Share-data\\Document\\仕入れ\\002_手配リスト\\手配発注リスト'
    ws['A2'].font = Font(size=11, bold=True, color=Constants.COLOR_RED)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws.merge_cells('A1:K1')
    ws.merge_cells('A2:K2')
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 20


def _create_column_headers(ws):
    """列ヘッダー作成"""
    for col, header in enumerate(Constants.EXCEL_COLUMNS, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color=Constants.COLOR_WHITE)
        cell.fill = PatternFill(start_color=Constants.COLOR_HEADER, 
                               end_color=Constants.COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _create_data_rows(ws, order):
    """データ行作成"""
    row_idx = 4
    parent_details = [d for d in order.details if d.parent_id is None]

    # 検収データを読み込み
    delivery_dict = DeliveryUtils.load_delivery_data()

    for detail in parent_details:
        row_idx = _write_detail_row(ws, detail, row_idx, is_parent=True, delivery_dict=delivery_dict)

        # 子アイテム
        children = [d for d in order.details if d.parent_id == detail.id]
        for child in children:
            row_idx = _write_detail_row(ws, child, row_idx, is_parent=False, delivery_dict=delivery_dict)

    return row_idx


def _get_cad_hyperlink(spec1):
    """仕様1からCADファイルのハイパーリンクパスを取得"""
    if not spec1 or not str(spec1).startswith('N'):
        return None
    spec1 = str(spec1)
    parts = spec1.split('-')
    if len(parts) < 2 or len(parts[0]) < 2:
        return None
    folder_letter = parts[0][1].upper()

    import glob

    # SERVER3のCADフォルダ（PDF優先 → mx2 → フォルダ）
    cad_folder = f"\\\\SERVER3\\Share-data\\CadData\\Parts\\{folder_letter}"
    try:
        pdf_files = glob.glob(os.path.join(cad_folder, f"{spec1}*.pdf"))
        if pdf_files:
            return pdf_files[0]
        mx2_files = glob.glob(os.path.join(cad_folder, f"{spec1}*.mx2"))
        if mx2_files:
            return mx2_files[0]
    except Exception:
        pass
    # ファイルが見つからない場合はフォルダへのリンク
    return cad_folder


def _write_detail_row(ws, detail, row_idx, is_parent=True, delivery_dict=None):
    """詳細行を出力"""
    is_blank = '加工用ブランク' in str(detail.order_type)
    supplier_cd = getattr(detail, 'supplier_cd', None)
    spec1_value = detail.spec1 or ''
    spec2_value = detail.spec2 or ''
    is_mekki = MekkiUtils.is_mekki_target(supplier_cd, spec2_value, spec1_value)

    remarks = MekkiUtils.add_mekki_alert(detail.remarks) if is_mekki else (detail.remarks or '')

    # 検収データから納入日・納入数を取得
    delivery_info = DeliveryUtils.get_delivery_info(detail.order_number, delivery_dict)
    delivery_date = delivery_info.get('納入日', '')
    delivery_qty = delivery_info.get('納入数', 0)
    # 納入数が0の場合は空欄表示
    delivery_qty_display = delivery_qty if delivery_qty > 0 else ''

    # 仕様1のCADリンクを事前に取得
    cad_link = _get_cad_hyperlink(spec1_value)

    data = [
        delivery_date,  # 検収日
        delivery_qty_display,  # 検収数
        detail.delivery_date, detail.supplier, detail.order_number,
        detail.quantity, detail.unit_measure, detail.item_name,
        detail.spec1, spec2_value, detail.order_type, detail.maker, remarks
    ]

    row_fill = ExcelStyler.get_fill(detail.is_received, row_idx % 2 == 0, not is_parent)
    cell_font = ExcelStyler.get_font(is_blank, False)

    for col, value in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.fill = row_fill
        cell.alignment = Alignment(vertical='center')

        if col == 10 and is_mekki:  # 仕様２のカラムがJ(10)に変更
            cell.font = ExcelStyler.get_font(False, True)
        elif cell_font:
            cell.font = cell_font

        if col == 8 and not is_parent:  # 品名のカラムがH(8)に変更
            cell.value = f"  └ {value}"

        # 仕様１(col=9)にCADハイパーリンクを設定
        if col == 9 and cad_link:
            cell.hyperlink = cad_link
            cell.font = Font(color="0000FF", underline="single", size=cell_font.size if cell_font and cell_font.size else 10)

    ws.row_dimensions[row_idx].height = 27
    return row_idx + 1

def _setup_print_settings(ws, row_idx, order, unit_display, customer, memo):
    """印刷設定"""
    ws.print_title_rows = '1:3'
    ws.print_area = f'A1:M{row_idx - 1}'
    
    footer_parts = [order.seiban]
    if unit_display:
        footer_parts.append(unit_display)
    if customer:
        footer_parts.append(customer)
    if memo:
        footer_parts.append(memo)
    
    footer_text = f"&24&B{'_'.join(footer_parts)}"
    
    for footer in [ws.oddFooter, ws.evenFooter, ws.firstFooter]:
        footer.center.text = "&P / &N"
        footer.left.text = footer_text
        footer.right.text = footer_text
    
    ws.sheet_view.view = 'pageBreakPreview'
    
# Excelファイル更新用の関数を追加
# 🔥 更新対象のExcelファイル一覧（DB直接クエリに移行中）
EXCEL_FILES_TO_REFRESH = [
    {
        'path': r"\\SERVER3\share-data\Document\Acrossデータ\製番一覧表.xlsx",
        'name': '製番一覧表',
        'sheet': '製番'
    }
]

def refresh_single_excel(excel_path, file_name):
    """単一のExcelファイルを更新"""
    excel = None
    wb = None
    try:
        if not os.path.exists(excel_path):
            return False, f"ファイルが見つかりません: {excel_path}"

        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.EnableEvents = False

        # ファイルを開く（手配発注_ALLは外部リンク更新が必要）
        update_links = 3 if '手配発注_ALL' in file_name else 0
        wb = excel.Workbooks.Open(
            Filename=excel_path,
            UpdateLinks=update_links,
            ReadOnly=False,
            Notify=False
        )

        # 背景クエリを無効化
        if hasattr(wb, 'Connections'):
            for i in range(1, wb.Connections.Count + 1):
                try:
                    conn = wb.Connections(i)
                    if hasattr(conn, 'ODBCConnection'):
                        conn.ODBCConnection.BackgroundQuery = False
                    elif hasattr(conn, 'OLEDBConnection'):
                        conn.OLEDBConnection.BackgroundQuery = False
                except:
                    pass

        # 全接続を更新
        wb.RefreshAll()
        excel.CalculateUntilAsyncQueriesDone()
        excel.CalculateFull()
        time.sleep(2)

        # 保存して閉じる
        wb.Save()
        wb.Close(SaveChanges=False)
        excel.Quit()

        return True, f"{file_name}を更新しました"

    except Exception as e:
        return False, f"{file_name}の更新エラー: {str(e)}"

    finally:
        try:
            if wb:
                wb.Close(SaveChanges=False)
        except:
            pass
        try:
            if excel:
                excel.Quit()
        except:
            pass

def refresh_excel_file():
    """複数のExcelファイルを順番に更新"""
    results = []
    all_success = True

    try:
        # COMを初期化（重要）
        pythoncom.CoInitialize()

        for file_info in EXCEL_FILES_TO_REFRESH:
            excel_path = file_info['path']
            file_name = file_info['name']

            print(f"📊 {file_name} を更新中...")
            success, message = refresh_single_excel(excel_path, file_name)
            results.append({'name': file_name, 'success': success, 'message': message})

            if not success:
                all_success = False
                print(f"  ❌ {message}")
            else:
                print(f"  ✅ {message}")

            # ファイル間で少し待機
            time.sleep(3)

        # 結果メッセージを作成
        success_count = sum(1 for r in results if r['success'])
        total_count = len(results)

        if all_success:
            message = f"全{total_count}ファイルを更新しました"
        else:
            message = f"{success_count}/{total_count}ファイルを更新（一部エラー）"

        return all_success, message, results

    except Exception as e:
        return False, f"更新エラー: {str(e)}", results

    finally:
        try:
            pythoncom.CoUninitialize()
        except:
            pass
        
def detect_seibans_from_excel(file_path, sheet_name, min_seiban='MHT0600'):
    """Excelから製番を自動検出"""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)
        
        if '製番' not in df.columns:
            return []
        
        # 製番列から一意の値を取得
        seibans = df['製番'].dropna().unique()
        
        # MHTで始まり、指定番号以降のものをフィルター
        filtered_seibans = []
        for seiban in seibans:
            seiban_str = str(seiban).strip()
            if seiban_str.startswith('MHT'):
                # MHT0600 -> 600として比較
                try:
                    seiban_num = int(seiban_str.replace('MHT', ''))
                    min_num = int(min_seiban.replace('MHT', ''))
                    if seiban_num >= min_num:
                        filtered_seibans.append(seiban_str)
                except ValueError:
                    continue
        
        return sorted(filtered_seibans)
    except Exception as e:
        print(f"製番検出エラー: {e}")
        return []
    
# グローバル変数に追加
previous_seiban_counts = {}

def get_seiban_counts(file_path, sheet_name='手配リスト_ALL'):
    """製番ごとの件数を取得"""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)
        if '製番' not in df.columns:
            return {}
        
        counts = df['製番'].value_counts().to_dict()
        return {str(k): int(v) for k, v in counts.items()}
    except Exception as e:
        print(f"件数取得エラー: {e}")
        return {}
    
@app.route('/api/order/<int:order_id>/send-completion-email', methods=['POST'])
def send_completion_email(order_id):
    """納品完了メールを作成してメーラーを起動"""
    try:
        order = Order.query.get_or_404(order_id)
        
        # 🔥 実際のExcelファイルパスを取得（get_order_excel_path()を使用）
        excel_path = get_order_excel_path(
            seiban=order.seiban,
            product_name=order.product_name,
            customer_abbr=order.customer_abbr
        )
        
        # メール送信
        success = EmailSender.send_completion_notification(
            seiban=order.seiban,
            product_name=order.product_name or '',
            customer_abbr=order.customer_abbr or '',
            unit=order.unit or '',
            memo2=order.memo2 or '',
            floor=order.floor or '',
            pallet_number=order.pallet_number or '',
            excel_path=excel_path,  # 🔥 実際のファイルパスを渡す
            sender_name='丸山'
        )
        
        if success:
            return jsonify({
                'success': True,
                'message': 'メーラーを起動しました'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'メーラーの起動に失敗しました'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/check-network-file-with-diff')
def check_network_file_with_diff():
    """ネットワークファイルの存在確認と差分検出（廃止済み）"""
    return jsonify({
        'accessible': False,
        'error': 'この機能は廃止されました。DBから直接取得してください。'
    })

@app.route('/api/seiban-list', methods=['GET'])
def get_seiban_list():
    """製番一覧を取得（V_D受注から直接取得）"""
    try:
        min_seiban = request.args.get('min_seiban')
        source = request.args.get('source', 'db')  # デフォルトはDB

        if source == 'db':
            # V_D受注から直接取得
            result = across_db.get_seiban_list_from_db(min_seiban)
            if result['success']:
                # UIとの互換性のためcustomer_nameをcustomer_abbrにマッピング
                items = []
                for item in result['items']:
                    items.append({
                        'seiban': item['seiban'],
                        'product_name': item['product_name'],
                        'customer_abbr': item['customer_name'],  # まとめ区分２
                        'memo2': item.get('memo2', '')
                    })
                return jsonify({
                    'success': True,
                    'items': items,
                    'count': len(items),
                    'source': 'V_D受注'
                })
            else:
                return jsonify({'success': False, 'error': result.get('error', 'DB取得エラー'), 'items': []})
        else:
            # 従来のExcel読み込み（フォールバック）
            seiban_info = load_seiban_info()
            if not seiban_info:
                return jsonify({'success': False, 'error': '製番一覧表を読み込めません', 'items': []})

            items = []
            for seiban, info in seiban_info.items():
                items.append({
                    'seiban': seiban,
                    'product_name': info.get('product_name', ''),
                    'customer_abbr': info.get('customer_abbr', '')
                })

            # 製番の降順でソート（新しいものが上）
            items.sort(key=lambda x: x['seiban'], reverse=True)

            return jsonify({'success': True, 'items': items, 'count': len(items), 'source': 'Excel'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'items': []})


@app.route('/api/detect-seibans', methods=['POST'])
def detect_seibans():
    """製番を自動検出"""
    try:
        data = request.json
        filepath = data.get('filepath')
        sheet_name = data.get('sheet_name')
        min_seiban = data.get('min_seiban', 'MHT0600')
        
        if not filepath or not sheet_name:
            return jsonify({'error': 'ファイルパスとシート名が必要です'}), 400
        
        seibans = detect_seibans_from_excel(filepath, sheet_name, min_seiban)
        
        return jsonify({
            'success': True,
            'seibans': seibans,
            'count': len(seibans)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/process', methods=['POST'])
def process_file_endpoint():
    """Excelファイル処理のメインエンドポイント"""
    try:
        data = request.json
        filepath = data['filepath']
        sheet1 = data['sheet1']
        sheet2 = data['sheet2']
        seiban = data['seiban']
        order_date_from = data.get('order_date_from')
        order_date_to = data.get('order_date_to')
        
        df_merged = process_excel_file(
            filepath, 
            sheet1, 
            sheet2, 
            seiban, 
            order_date_from, 
            order_date_to
        )
        
        if df_merged is None or len(df_merged) == 0:
            return jsonify({
                'success': False, 
                'error': f'製番 {seiban} のデータが見つかりません'
            })
        
        save_to_database(df_merged, seiban)
        
        return jsonify({
            'success': True,
            'message': f'{seiban} の処理が完了しました（{len(df_merged)}件）'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

# Routes
@app.route('/api/refresh-excel', methods=['POST'])
def refresh_excel_endpoint():
    """複数のExcelファイルを更新するエンドポイント（廃止済み）"""
    return jsonify({
        'success': False,
        'error': 'この機能は廃止されました。DBから直接取得してください。'
    }), 400

# Pythonスクリプトを実行するエンドポイントを修正
@app.route('/api/run-refresh-script', methods=['POST'])
def run_refresh_script():
    """refresh_order_list.pyを実行"""
    try:
        # スクリプトパスのリスト（優先順位順）
        script_paths = [
            r"C:\Users\t.maruyama\order_merge\refresh_order_list.py",  # アプリと同じフォルダ
            r"C:\Users\t.maruyama\refresh_order_list.py",  # 元のパス
            os.path.join(os.path.dirname(__file__), "refresh_order_list.py")  # 相対パス
        ]
        
        # 存在するスクリプトを探す
        script_path = None
        for path in script_paths:
            if os.path.exists(path):
                script_path = path
                break
        
        if not script_path:
            return jsonify({
                'success': False,
                'error': f'スクリプトが見つかりません。以下のパスを確認してください:\n' + '\n'.join(script_paths)
            }), 404
        
        # Pythonスクリプトを実行（タイムアウトを延長）
        result = subprocess.run(
            [sys.executable, script_path],
            capture_output=True,
            text=True,
            timeout=120,  # 120秒に延長
            encoding='utf-8',  # 文字エンコーディングを指定
            errors='replace'  # エンコーディングエラーを回避
        )
        
        if result.returncode == 0:
            return jsonify({
                'success': True,
                'message': 'スクリプト実行成功',
                'output': result.stdout
            })
        else:
            return jsonify({
                'success': False,
                'error': result.stderr or 'スクリプト実行失敗',
                'output': result.stdout  # エラー時も出力を表示
            }), 500
            
    except subprocess.TimeoutExpired:
        return jsonify({
            'success': False,
            'error': 'タイムアウト（120秒）- スクリプトが長時間実行されています'
        }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# 🔥 製番単位でデータを更新（マージ）するAPI
@app.route('/api/generate-labels', methods=['POST'])
def generate_labels_endpoint():
    """製番のラベルをExcelで生成してダウンロード"""
    try:
        data = request.json
        seiban = data.get('seiban')

        if not seiban:
            return jsonify({'success': False, 'error': '製番が指定されていません'}), 400

        from label_maker import create_labels_for_seiban

        # labelsフォルダに出力
        labels_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'labels')
        os.makedirs(labels_dir, exist_ok=True)
        safe_seiban = seiban.replace('/', '_').replace('\\', '_')
        output_path = os.path.join(labels_dir, f'{safe_seiban}_ラベル.xlsx')

        result = create_labels_for_seiban(seiban, output_path)
        if result is None:
            return jsonify({'success': False, 'error': f'製番 {seiban} のデータが見つかりません'}), 404

        return send_file(
            output_path,
            as_attachment=True,
            download_name=f'{safe_seiban}_ラベル.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"❌ ラベル生成エラー: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/refresh-seiban', methods=['POST'])
def refresh_seiban_endpoint():
    """製番単位でデータを最新に更新（廃止済み）"""
    return jsonify({
        'success': False,
        'error': 'この機能は廃止されました。DBから直接取得してください。'
    }), 400

@app.route('/')
def index():
    """Main page"""
    return render_template('index.html')

@app.route('/api/debug-paths')
def debug_paths():
    """パスの接続状態をデバッグ"""
    import os

    debug_info = {
        'configured_paths': {
            'history': app.config.get('HISTORY_EXCEL_PATH', 'Not configured'),
            'seiban': app.config.get('SEIBAN_LIST_PATH', 'Not configured')
        },
        'path_checks': {}
    }
    
    # 各パスの存在確認
    for name, path in debug_info['configured_paths'].items():
        if path != 'Not configured':
            # os.path.exists での確認
            exists_os = os.path.exists(path)
            
            # Path オブジェクトでの確認
            path_obj = Path(path)
            exists_path = path_obj.exists()
            
            debug_info['path_checks'][name] = {
                'path': path,
                'os_exists': exists_os,
                'path_exists': exists_path,
                'is_file': path_obj.is_file() if exists_path else False,
                'parent_exists': path_obj.parent.exists()
            }
            
            # ファイルサイズの取得
            if exists_os:
                try:
                    stats = os.stat(path)
                    debug_info['path_checks'][name]['size_mb'] = round(stats.st_size / (1024 * 1024), 2)
                    debug_info['path_checks'][name]['readable'] = os.access(path, os.R_OK)
                except Exception as e:
                    debug_info['path_checks'][name]['error'] = str(e)
    
    # サーバーへの接続確認
    server_path = r'\\server3'
    debug_info['server_connection'] = {
        'path': server_path,
        'connected': os.path.exists(server_path)
    }
    
    if debug_info['server_connection']['connected']:
        try:
            shares = os.listdir(server_path)
            debug_info['server_connection']['shares'] = shares[:10]  # 最初の10個
        except Exception as e:
            debug_info['server_connection']['error'] = str(e)
    
    # Python環境情報
    debug_info['environment'] = {
        'python_version': sys.version,
        'platform': sys.platform,
        'cwd': os.getcwd(),
        'flask_env': app.config.get('ENV', 'not set')
    }
    
    return jsonify(debug_info)

@app.route('/api/check-network-file')
def check_network_file():
    """Check if network file is accessible（廃止済み）"""
    return jsonify({
        'accessible': False,
        'error': 'この機能は廃止されました。DBから直接取得してください。'
    })

@app.route('/api/load-network-file', methods=['POST'])
def load_network_file():
    """Load file from network location（廃止済み）"""
    return jsonify({
        'success': False,
        'error': 'この機能は廃止されました。DBから直接取得してください。'
    }), 400

@app.route('/api/load-from-odbc', methods=['POST'])
def load_from_odbc_endpoint():
    """Load data directly from ODBC"""
    try:
        data = request.json
        seiban = data.get('seiban', '')
        
        if not seiban:
            return jsonify({'error': '製番を入力してください'}), 400
        
        # Load from ODBC
        odbc_data, error = load_from_odbc()
        
        if error:
            return jsonify({
                'success': False,
                'error': error
            }), 500
        
        # Process the data
        df1 = odbc_data['手配リスト']
        df2 = odbc_data['発注リスト']
        
        # Filter by seiban
        df1 = df1[df1['製番'].astype(str).str.startswith(seiban)]
        df2 = df2[df2['製番'].astype(str).str.startswith(seiban)]
        
        # Merge and save to database
        df_merged = process_excel_file_from_dataframes(df1, df2, seiban)
        save_to_database(df_merged, seiban)
        
        return jsonify({
            'success': True,
            'message': f'ODBC経由で{len(df_merged)}件のデータを処理しました'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ========== Across DB 直接クエリ API ==========
import across_db

@app.route('/api/across-db/test')
def across_db_test():
    """Across DB 接続テスト"""
    try:
        result = across_db.test_connection()
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/across-db/check-updates')
def across_db_check_updates():
    """DB更新チェック（手配・発注リストの変更検知）"""
    try:
        result = across_db.check_db_updates()
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/across-db/status')
def across_db_status():
    """DB現在状態取得"""
    try:
        result = across_db.get_db_status()
        if result.get('success'):
            # setはJSON化できないので変換
            return jsonify({
                'success': True,
                'tehai': {
                    'count': result['tehai']['count'],
                    'seiban_count': result['tehai']['seiban_count']
                },
                'hacchu': {
                    'count': result['hacchu']['count']
                },
                'timestamp': result['timestamp'].isoformat()
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/across-db/seiban-status/<seiban>')
def across_db_seiban_status(seiban):
    """製番別の手配・発注状況取得"""
    try:
        result = across_db.get_seiban_updates(seiban)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/across-db/delivery-schedule')
def across_db_delivery_schedule():
    """発注DBから納品予定を取得"""
    try:
        start_date = request.args.get('start_date', None)
        days = int(request.args.get('days', 7))
        seibans_str = request.args.get('seibans', '')
        seibans = [s.strip() for s in seibans_str.split(',') if s.strip()] if seibans_str else None

        result = across_db.get_delivery_schedule_from_db(start_date, days, seibans)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/across-db/columns')
def across_db_columns():
    """ビューのカラム一覧取得"""
    try:
        view_name = request.args.get('view', 'V_D発注')
        columns = across_db.get_view_columns(view_name)
        return jsonify({'columns': columns})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/across-db/query')
def across_db_query():
    """ビューへの自由検索"""
    try:
        view_name = request.args.get('view', 'V_D発注')
        search_type = request.args.get('search_type', '')
        search_value = request.args.get('search_value', '').strip()
        limit = min(int(request.args.get('limit', 100)), 500)

        where_clause = None
        params = None

        if search_type and search_value:
            if search_type == '発注番号':
                # ゼロパディング対応
                search_value = search_value.zfill(8)
            where_clause = f'{search_type} = ?'
            params = [search_value]

        result = across_db.query_view(view_name, where_clause, params, limit)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/across-db/order-detail')
def across_db_order_detail():
    """発注番号の詳細情報（発注 + 発注残 + 仕入を統合）"""
    try:
        order_number = request.args.get('order_number', '').strip()
        if not order_number:
            return jsonify({'error': '発注番号を入力してください'}), 400

        order = across_db.search_order(order_number)
        remaining = across_db.search_order_remaining(order_number)
        receipts = across_db.search_receipts(order_number)

        return jsonify({
            'order': order,
            'remaining': remaining,
            'receipts': receipts
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/across-db/process', methods=['POST'])
def across_db_process():
    """DB直接クエリでマージ→DB保存（Excel不要）- 変更内容をフィードバック"""
    try:
        data = request.json
        seiban = data.get('seiban', '').strip()
        order_date_from = data.get('order_date_from')
        order_date_to = data.get('order_date_to')
        include_mihatchu = data.get('include_mihatchu', True)  # デフォルトでON

        if not seiban:
            return jsonify({'success': False, 'error': '製番を入力してください'}), 400

        # 更新前の状態を取得
        before_orders = Order.query.filter_by(seiban=seiban, is_archived=False).all()
        before_units = {}
        for order in before_orders:
            unit_name = order.unit or ''
            detail_count = OrderDetail.query.filter_by(order_id=order.id).count()
            before_units[unit_name] = {
                'order_id': order.id,
                'detail_count': detail_count,
                'status': order.status
            }

        # DB直接クエリでマージ済みDataFrameを取得
        if include_mihatchu:
            df_merged = across_db.merge_from_db_with_mihatchu(seiban, order_date_from, order_date_to)
        else:
            df_merged = across_db.merge_from_db(seiban, order_date_from, order_date_to)

        if df_merged is None or len(df_merged) == 0:
            return jsonify({
                'success': False,
                'error': f'製番 {seiban} のデータが見つかりません（Across DB）'
            })

        # 既存のsave_to_database()でDB保存
        save_to_database(df_merged, seiban)

        # 更新後の状態を取得
        after_orders = Order.query.filter_by(seiban=seiban, is_archived=False).all()
        after_units = {}
        for order in after_orders:
            unit_name = order.unit or ''
            detail_count = OrderDetail.query.filter_by(order_id=order.id).count()
            after_units[unit_name] = {
                'order_id': order.id,
                'detail_count': detail_count,
                'status': order.status
            }

        # 変更内容を分析
        changes = {
            'added_units': [],
            'updated_units': [],
            'unchanged_units': [],
            'total_before': len(before_units),
            'total_after': len(after_units)
        }

        for unit_name, after_info in after_units.items():
            if unit_name not in before_units:
                # 新規追加されたユニット
                changes['added_units'].append({
                    'unit': unit_name or '(名称なし)',
                    'detail_count': after_info['detail_count']
                })
            else:
                before_info = before_units[unit_name]
                if after_info['detail_count'] != before_info['detail_count']:
                    # 内容が更新されたユニット
                    diff = after_info['detail_count'] - before_info['detail_count']
                    changes['updated_units'].append({
                        'unit': unit_name or '(名称なし)',
                        'before_count': before_info['detail_count'],
                        'after_count': after_info['detail_count'],
                        'diff': diff
                    })
                else:
                    changes['unchanged_units'].append(unit_name or '(名称なし)')

        # メッセージ生成
        msg_parts = [f'{seiban} の更新完了']
        if changes['added_units']:
            msg_parts.append(f"新規ユニット: {len(changes['added_units'])}件")
        if changes['updated_units']:
            msg_parts.append(f"更新ユニット: {len(changes['updated_units'])}件")
        msg_parts.append(f"合計: {len(df_merged)}件")

        return jsonify({
            'success': True,
            'message': ' / '.join(msg_parts),
            'changes': changes,
            'total_items': len(df_merged)
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/across-db/merge-test')
def across_db_merge_test():
    """製番でマージテスト（V_D手配リスト + V_D発注）"""
    try:
        seiban = request.args.get('seiban', '').strip()
        if not seiban:
            return jsonify({'error': '製番を入力してください'}), 400

        result = across_db.merge_test_by_seiban(seiban)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/across-db/mihatchu')
def across_db_mihatchu():
    """V_D未発注から検索（社内加工品）"""
    try:
        seiban = request.args.get('seiban', '').strip()
        supplier_cd = request.args.get('supplier_cd', '').strip() or None
        order_type_cd = request.args.get('order_type_cd', '').strip() or None
        if not seiban:
            return jsonify({'error': '製番を入力してください'}), 400

        result = across_db.search_mihatchu(seiban, supplier_cd, order_type_cd)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/across-db/zaiko-buhin', methods=['POST'])
def across_db_zaiko_buhin():
    """在庫部品（手配区分CD=15）を検索"""
    try:
        data = request.get_json() or {}
        seibans = data.get('seibans', None)
        result = across_db.search_zaiko_buhin(seibans)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/across-db/0zaiko')
def across_db_0zaiko():
    """0ZAIKO（在庫品発注用製番）の手配リストを検索"""
    try:
        result = across_db.search_0zaiko_tehai()
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/get-system-status')
def get_system_status():
    """Get system status"""
    try:
        status = {
            'last_refresh': last_refresh_time.isoformat() if last_refresh_time else None,
            'cached_file': cached_file_info,
            'odbc_enabled': app.config.get('USE_ODBC', False)
        }
        return jsonify(status)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Handle file upload and processing"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Save uploaded file
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Get sheet names
        wb = load_workbook(filepath, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        return jsonify({
            'success': True,
            'filepath': filepath,
            'sheet_names': sheet_names
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/orders')
def get_orders():
    """Get all active orders"""
    try:
        from sqlalchemy import func, case
        
        # ユニット名無しを最後にソート
        results = db.session.query(
            Order,
            func.count(OrderDetail.id).label('detail_count'),
            func.sum(OrderDetail.is_received.cast(db.Integer)).label('received_count'),
            func.max(OrderDetail.has_internal_processing.cast(db.Integer)).label('has_internal_processing')
        ).outerjoin(OrderDetail).filter(
            Order.is_archived == False
        ).group_by(Order.id).order_by(
            Order.seiban.desc(),
            case(
                (Order.unit == '', 1),  # ユニット名無しを最後に
                (Order.unit == None, 1),
                else_=0
            ),
            Order.unit
        ).all()
        
        orders = []
        for order, detail_count, received_count, has_internal in results:
            orders.append({
                'id': order.id,
                'seiban': order.seiban,
                'unit': order.unit or '',
                'product_name': order.product_name or '',
                'customer_abbr': order.customer_abbr or '',
                'memo2': order.memo2 or '',
                'location': order.location,
                'status': order.status,
                'remarks': order.remarks or '',
                'created_at': to_jst(order.created_at).strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': to_jst(order.updated_at).strftime('%Y-%m-%d %H:%M:%S'),
                'detail_count': detail_count or 0,
                'received_count': received_count or 0,
                'has_internal_processing': bool(has_internal)
            })
        
        return jsonify(orders)

    except Exception as e:
        import traceback
        print(f"Error getting orders: {str(e)}")
        print(traceback.print_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/orders/update-info')
def get_orders_update_info():
    """備考・画像が設定されているユニットの一覧を取得"""
    try:
        from sqlalchemy import or_

        orders = Order.query.filter(
            Order.is_archived == False,
            or_(
                Order.remarks != None,
                Order.remarks != '',
                Order.image_path != None,
                Order.image_path != ''
            )
        ).order_by(Order.updated_at.desc()).all()

        # 備考または画像が実際に入っているものだけ抽出
        result = []
        for order in orders:
            has_remarks = bool(order.remarks and order.remarks.strip())
            has_image = bool(order.image_path and order.image_path.strip())
            if not has_remarks and not has_image:
                continue

            result.append({
                'order_id': order.id,
                'seiban': order.seiban,
                'unit': order.unit or '',
                'status': order.status or '',
                'location': order.floor or '未設定',
                'pallet_number': order.pallet_number or '未設定',
                'remarks': order.remarks or '',
                'has_image': has_image,
                'image_url': f'/api/order/{order.id}/image' if has_image else None,
                'updated_at': order.updated_at.strftime('%m/%d %H:%M') if order.updated_at else ''
            })

        return jsonify({'success': True, 'orders': result, 'total': len(result)})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/delivery-schedule')
def get_delivery_schedule():
    """今日の納品リストと1週間の予定を取得"""
    try:
        from datetime import date, timedelta

        # カスタム開始日対応
        start_date_str = request.args.get('start_date', '')
        if start_date_str:
            try:
                parts = start_date_str.split('-')
                today = date(int(parts[0]), int(parts[1]), int(parts[2]))
            except (ValueError, IndexError):
                today = date.today()
        else:
            today = date.today()
        week_end = today + timedelta(days=7)

        # 全アクティブ注文の詳細を取得
        orders = Order.query.filter_by(is_archived=False).all()

        schedule = {}  # {date_str: [items]}

        for order in orders:
            for detail in order.details:
                if not detail.delivery_date or detail.delivery_date.strip() == '' or detail.delivery_date == '-':
                    continue

                parsed = _parse_delivery_date_to_date(detail.delivery_date)
                if not parsed:
                    continue

                if parsed < today or parsed > week_end:
                    continue

                date_key = parsed.isoformat()
                if date_key not in schedule:
                    schedule[date_key] = []

                # 加工用ブランクの場合、親（追加工）の処理先を取得
                # DB構造: 追加工(11)=parent → ブランク(13)=child (parent_id)
                next_steps = []
                is_blank = (str(detail.order_type_code or '').strip() == '13' or
                           '加工用ブランク' in str(detail.order_type or ''))
                if is_blank:
                    from utils.mekki_utils import MekkiUtils
                    parent = None

                    # 方法1: parent_idで親（追加工）を取得
                    if detail.parent_id:
                        parent = OrderDetail.query.get(detail.parent_id)

                    # 方法2: parent_idがない場合、同じ注文内で行No・階層ルールでマッチング
                    if not parent:
                        blank_row_no = int(detail.row_number or 0) if detail.row_number else 0
                        blank_hierarchy = detail.hierarchy or 0
                        for d in order.details:
                            if d.id == detail.id:
                                continue
                            d_code = str(d.order_type_code or '').strip()
                            if d_code != '11' and '追加工' not in str(d.order_type or ''):
                                continue
                            d_row_no = int(d.row_number or 0) if d.row_number else 0
                            d_hierarchy = d.hierarchy or 0
                            # 追加工の行No < ブランクの行No <= 追加工の行No+300, 階層差=+1
                            if (d_row_no < blank_row_no and
                                blank_row_no <= d_row_no + 300 and
                                blank_hierarchy == d_hierarchy + 1):
                                parent = d
                                break

                    if parent:
                        step = {
                            'supplier': parent.supplier or '',
                            'item_name': parent.item_name or '',
                            'order_type': parent.order_type or '',
                            'is_mekki': False
                        }
                        if MekkiUtils.is_mekki_target(parent.supplier_cd, parent.spec2, parent.spec1):
                            step['is_mekki'] = True
                        next_steps.append(step)

                schedule[date_key].append({
                    'detail_id': detail.id,
                    'order_id': order.id,
                    'seiban': order.seiban,
                    'unit': order.unit or '',
                    'item_name': detail.item_name or '',
                    'spec1': detail.spec1 or '',
                    'spec2': detail.spec2 or '',
                    'supplier': detail.supplier or '',
                    'order_number': detail.order_number or '',
                    'quantity': detail.quantity or 0,
                    'unit_measure': detail.unit_measure or '',
                    'is_received': detail.is_received,
                    'delivery_date': detail.delivery_date,
                    'reply_delivery_date': detail.reply_delivery_date or '',
                    'order_type': detail.order_type or '',
                    'order_type_code': detail.order_type_code or '',
                    'product_name': order.product_name or '',
                    'customer_abbr': order.customer_abbr or '',
                    'cad_link': _get_cad_hyperlink(detail.spec1 or '') or '',
                    'is_blank': is_blank,
                    'next_steps': next_steps
                })

        # 日付順にソート
        result = []
        for date_key in sorted(schedule.keys()):
            items = schedule[date_key]
            parsed_date = date.fromisoformat(date_key)
            weekday_names = ['月', '火', '水', '木', '金', '土', '日']
            weekday = weekday_names[parsed_date.weekday()]

            result.append({
                'date': date_key,
                'display_date': f"{parsed_date.month}/{parsed_date.day}({weekday})",
                'is_today': parsed_date == date.today(),
                'is_weekend': parsed_date.weekday() >= 5,
                'total': len(items),
                'received': sum(1 for i in items if i['is_received']),
                'items': items
            })

        # 集計情報
        all_items = [item for d in result for item in d['items']]
        unique_seibans = sorted(set(item['seiban'] for item in all_items))
        unique_units = sorted(set(item['unit'] for item in all_items if item['unit']))
        unique_suppliers = sorted(set(item['supplier'] for item in all_items if item['supplier']))

        return jsonify({
            'success': True,
            'start_date': today.isoformat(),
            'today': date.today().isoformat(),
            'week_end': week_end.isoformat(),
            'days': result,
            'total_items': len(all_items),
            'summary': {
                'seibans': unique_seibans,
                'seiban_count': len(unique_seibans),
                'units': unique_units,
                'unit_count': len(unique_units),
                'suppliers': unique_suppliers,
                'supplier_count': len(unique_suppliers)
            }
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def _parse_delivery_date_to_date(date_str):
    """納期文字列をdateオブジェクトに変換"""
    from datetime import date
    import re
    if not date_str:
        return None
    # YY/MM/DD
    m = re.match(r'^(\d{2})/(\d{1,2})/(\d{1,2})$', date_str)
    if m:
        return date(2000 + int(m.group(1)), int(m.group(2)), int(m.group(3)))
    # YYYY/MM/DD
    m = re.match(r'^(\d{4})/(\d{1,2})/(\d{1,2})$', date_str)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    # YYYY-MM-DD
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', date_str)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


@app.route('/api/orders/gantt-data')
def get_gantt_data():
    """ガントチャート用に最適化された納期データを一括取得"""
    try:
        from sqlalchemy import func

        # 全アクティブ注文と詳細を一度に取得
        orders = Order.query.filter_by(is_archived=False).all()

        gantt_data = []
        for order in orders:
            # 各注文の納期を取得
            delivery_dates = [
                d.delivery_date for d in order.details
                if d.delivery_date and d.delivery_date.strip() and d.delivery_date != '-'
            ]

            if delivery_dates:
                # 進捗計算
                total_details = len(order.details)
                received_details = sum(1 for d in order.details if d.is_received)
                progress = (received_details / total_details * 100) if total_details > 0 else 0

                gantt_data.append({
                    'id': order.id,
                    'seiban': order.seiban,
                    'unit': order.unit or 'ユニット名無し',
                    'status': order.status,
                    'progress': progress,
                    'delivery_dates': delivery_dates  # 納期のリスト
                })

        return jsonify(gantt_data)

    except Exception as e:
        import traceback
        print(f"Error getting gantt data: {str(e)}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/archived-orders')
def get_archived_orders():
    """Get archived orders"""
    try:
        orders = Order.query.filter_by(is_archived=True).order_by(Order.archived_at.desc()).all()
        result = []
        for order in orders:
            # unitが空の場合は'-'として表示
            unit_display = order.unit if order.unit else '-'
            
            result.append({
                'id': order.id,
                'seiban': order.seiban,
                'unit': unit_display,
                'product_name': order.product_name or '',
                'status': order.status,
                'location': order.location,
                'remarks': order.remarks,
                'archived_at': order.archived_at.isoformat() if order.archived_at else None,
                'detail_count': len(order.details),
                'received_count': sum(1 for d in order.details if d.is_received)
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/order/<int:order_id>/archive', methods=['POST'])
def archive_order(order_id):
    """Archive an order"""
    try:
        order = Order.query.get_or_404(order_id)
        
        # 納品完了チェック
        if order.status != '納品完了':
            return jsonify({
                'success': False,
                'error': '納品完了の注文のみアーカイブできます'
            }), 400
        
        order.is_archived = True
        order.archived_at = datetime.now(timezone.utc)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'製番 {order.seiban} をアーカイブしました'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/order/<int:order_id>/unarchive', methods=['POST'])
def unarchive_order(order_id):
    """Unarchive an order"""
    try:
        order = Order.query.get_or_404(order_id)
        
        order.is_archived = False
        order.archived_at = None
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'製番 {order.seiban} を復元しました'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
    
@app.route('/receive/<seiban>/<path:unit>')
@app.route('/receive/<seiban>/')  # ユニット名が空の場合
def receive_page(seiban, unit=''):
    """受入専用ページ（スマートフォン用）- 製番/ユニットでURL固定"""
    try:
        from urllib.parse import unquote
        unit = unquote(unit)  # URLデコード
        order = Order.query.filter_by(seiban=seiban, unit=unit, is_archived=False).first_or_404()

        # 🔥 検収データを読み込み
        delivery_dict = DeliveryUtils.load_delivery_data()

        # 詳細リストを取得
        details = []
        for detail in order.details:
            # 検収データから納入日・納入数を取得
            delivery_info = DeliveryUtils.get_delivery_info(detail.order_number, delivery_dict)

            details.append({
                'id': detail.id,
                'delivery_date': detail.delivery_date,
                'reply_delivery_date': detail.reply_delivery_date or '',
                'supplier': detail.supplier,
                'order_number': detail.order_number,
                'quantity': detail.quantity,
                'unit_measure': detail.unit_measure,
                'item_name': detail.item_name,
                'spec1': detail.spec1,
                'spec2': detail.spec2,
                'order_type': detail.order_type,
                'remarks': detail.remarks,
                'is_received': detail.is_received,
                'received_quantity': detail.received_quantity,  # 実際の受入数量
                'parent_id': detail.parent_id,
                'has_internal_processing': detail.has_internal_processing,
                # 🔥 検収データを追加
                'received_delivery_date': delivery_info.get('納入日', ''),
                'received_delivery_qty': delivery_info.get('納入数', 0)
            })
        
        # スマートフォン用のシンプルなHTMLを返す
        html = f"""
<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>受入 - {order.seiban}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            background: #f5f5f5;
            padding: 10px;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 15px;
            text-align: center;
        }}
        .header h1 {{
            font-size: 1.5em;
            margin-bottom: 5px;
        }}
        .info-box {{
            background: white;
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 15px;
            box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        }}
        .info-row {{
            display: flex;
            justify-content: space-between;
            padding: 8px 0;
            border-bottom: 1px solid #eee;
        }}
        .info-row:last-child {{
            border-bottom: none;
        }}
        .label {{
            font-weight: bold;
            color: #666;
        }}
        .detail-item {{
            background: white;
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 10px;
            box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        }}
        .detail-item.received {{
            background: #d4edda;
            border-left: 4px solid #28a745;
        }}
        .detail-item.child {{
            margin-left: 20px;
            background: #f8f9fa;
        }}
        .detail-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
        }}
        .item-name {{
            font-weight: bold;
            font-size: 1.1em;
            color: #333;
        }}
        .btn {{
            padding: 10px 20px;
            border: none;
            border-radius: 5px;
            font-size: 1em;
            cursor: pointer;
            width: 100%;
            margin-top: 10px;
        }}
        .btn-primary {{
            background: #667eea;
            color: white;
        }}
        .btn-success {{
            background: #28a745;
            color: white;
        }}
        .btn-warning {{
            background: #ffc107;
            color: #212529;
        }}
        .status-badge {{
            padding: 5px 10px;
            border-radius: 15px;
            font-size: 0.85em;
            font-weight: bold;
        }}
        .badge-success {{
            background: #28a745;
            color: white;
        }}
        .badge-warning {{
            background: #ffc107;
            color: #212529;
        }}
        .detail-row {{
            padding: 5px 0;
            font-size: 0.9em;
        }}
        .toast {{
            position: fixed;
            top: 20px;
            right: 20px;
            background: #28a745;
            color: white;
            padding: 15px 20px;
            border-radius: 5px;
            box-shadow: 0 5px 15px rgba(0,0,0,0.3);
            z-index: 1000;
            display: none;
        }}
        .toast.show {{
            display: block;
            animation: slideIn 0.3s;
        }}
        @keyframes slideIn {{
            from {{ transform: translateX(100%); }}
            to {{ transform: translateX(0); }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📦 {order.seiban}</h1>
        <div>{order.unit or ''}</div>
    </div>
    
    <div class="info-box">
        <div class="info-row">
            <span class="label">ステータス:</span>
            <span>{order.status}</span>
        </div>
        <div class="info-row">
            <span class="label">品名:</span>
            <span>{order.product_name or ''}</span>
        </div>
        <div class="info-row">
            <span class="label">得意先:</span>
            <span>{order.customer_abbr or ''}</span>
        </div>
    </div>

    <!-- 🔥 場所・パレット番号編集セクション -->
    <div class="info-box" style="background: #e7f3ff; border-left: 4px solid #007bff;">
        <div style="margin-bottom: 10px; font-weight: bold; color: #004085;">📍 保管場所</div>
        <div style="margin-bottom: 10px;">
            <label style="display: block; font-size: 0.9em; color: #666; margin-bottom: 5px;">場所</label>
            <select id="floorInput" style="width: 100%; padding: 10px; border: 1px solid #007bff; border-radius: 5px; font-size: 0.95em;">
                <option value="">未設定</option>
                <option value="1F" {'selected' if order.floor == '1F' else ''}>1F</option>
                <option value="2F" {'selected' if order.floor == '2F' else ''}>2F</option>
            </select>
        </div>
        <div style="margin-bottom: 10px;">
            <label style="display: block; font-size: 0.9em; color: #666; margin-bottom: 5px;">パレット（棚）番号</label>
            <select id="palletInput" style="width: 100%; padding: 10px; border: 1px solid #007bff; border-radius: 5px; font-size: 0.95em;">
                <option value="">未設定</option>
                <!-- パレット -->
                <option value="P001" {'selected' if order.pallet_number == 'P001' else ''}>P001(パレット)</option>
                <option value="P002" {'selected' if order.pallet_number == 'P002' else ''}>P002(パレット)</option>
                <option value="P003" {'selected' if order.pallet_number == 'P003' else ''}>P003(パレット)</option>
                <option value="P004" {'selected' if order.pallet_number == 'P004' else ''}>P004(パレット)</option>
                <option value="P005" {'selected' if order.pallet_number == 'P005' else ''}>P005(パレット)</option>
                <option value="P006" {'selected' if order.pallet_number == 'P006' else ''}>P006(パレット)</option>
                <option value="P007" {'selected' if order.pallet_number == 'P007' else ''}>P007(パレット)</option>
                <option value="P008" {'selected' if order.pallet_number == 'P008' else ''}>P008(パレット)</option>
                <option value="P009" {'selected' if order.pallet_number == 'P009' else ''}>P009(パレット)</option>
                <option value="P010" {'selected' if order.pallet_number == 'P010' else ''}>P010(パレット)</option>
                <!-- 台車 -->
                <option value="D001" {'selected' if order.pallet_number == 'D001' else ''}>D001（台車）</option>
                <option value="D002" {'selected' if order.pallet_number == 'D002' else ''}>D002（台車）</option>
                <option value="D003" {'selected' if order.pallet_number == 'D003' else ''}>D003（台車）</option>
                <!-- 棚 -->
                <option value="T001" {'selected' if order.pallet_number == 'T001' else ''}>T001(棚)</option>
                <option value="T002" {'selected' if order.pallet_number == 'T002' else ''}>T002(棚)</option>
                <option value="T003" {'selected' if order.pallet_number == 'T003' else ''}>T003(棚)</option>
                <option value="T004" {'selected' if order.pallet_number == 'T004' else ''}>T004(棚)</option>
                <option value="T005" {'selected' if order.pallet_number == 'T005' else ''}>T005(棚)</option>
            </select>
        </div>
    </div>

    <!-- 🔥 備考セクション追加 -->
    <div class="info-box" style="background: #fff3cd;">
        <div style="margin-bottom: 10px; font-weight: bold; color: #856404;">📝 備考</div>
        <textarea id="remarksInput" style="width: 100%; min-height: 80px; padding: 10px; border: 1px solid #ffc107; border-radius: 5px; font-size: 0.95em; resize: vertical;">{order.remarks or ''}</textarea>
    </div>

    <!-- 📷 画像セクション -->
    <div class="info-box" style="background: #e8f5e9; border-left: 4px solid #4caf50;">
        <div style="margin-bottom: 10px; font-weight: bold; color: #2e7d32;">📷 画像</div>
        <div id="imagePreviewArea" style="margin: 10px 0; text-align: center;">
            <img id="orderImage" src="/api/order/{order.id}/image"
                 style="max-width: 100%; max-height: 250px; border-radius: 8px; display: none; cursor: pointer;"
                 onclick="openImageFullscreen(this.src)"
                 onerror="this.style.display='none'; document.getElementById('noImageText').style.display='block';"
                 onload="this.style.display='block'; document.getElementById('noImageText').style.display='none';">
            <p id="noImageText" style="color: #888; font-style: italic;">画像なし</p>
        </div>
        <div style="display: flex; gap: 8px; flex-wrap: wrap;">
            <label style="flex: 1; min-width: 120px;">
                <div class="btn btn-primary" style="text-align: center; margin: 0;">📤 ファイル選択</div>
                <input type="file" id="imageUploadFile" accept="image/*"
                       style="display: none;" onchange="uploadOrderImage({order.id})">
            </label>
            <label style="flex: 1; min-width: 120px;">
                <div class="btn btn-success" style="text-align: center; margin: 0;">📸 カメラ撮影</div>
                <input type="file" id="imageUploadCamera" accept="image/*" capture="environment"
                       style="display: none;" onchange="uploadOrderImageFromCamera({order.id})">
            </label>
        </div>
        <button class="btn" style="background: #dc3545; color: white; margin-top: 8px;" onclick="deleteOrderImage({order.id})">🗑️ 画像を削除</button>
        <p style="font-size: 0.75em; color: #666; margin-top: 8px; text-align: center;">※FullHD (1920x1080) に自動圧縮されます</p>
    </div>

    <!-- 🔥 自動保存インジケーター -->
    <div id="autoSaveIndicator" style="text-align: center; padding: 10px; color: #28a745; font-size: 0.9em; display: none;">
        ✅ 自動保存済み
    </div>

    <!-- 🔥 バーコードスキャン入力欄 -->
    <div class="info-box" style="background: #f0f7ff; border-left: 4px solid #0066cc;">
        <div style="margin-bottom: 10px; font-weight: bold; color: #004085;">📷 バーコードスキャン</div>
        <div style="display: flex; gap: 8px;">
            <input type="text" id="barcodeInput"
                   placeholder="バーコードをスキャン (例: 00088333P)"
                   style="flex: 1; padding: 12px; border: 2px solid #0066cc; border-radius: 5px; font-size: 1em;"
                   onkeypress="if(event.key==='Enter') processBarcode()">
            <button class="btn btn-primary" onclick="processBarcode()" style="white-space: nowrap;">
                🔍 検索
            </button>
        </div>
        <div id="barcodeResult" style="margin-top: 10px; padding: 10px; border-radius: 5px; display: none;"></div>
        <p style="font-size: 0.75em; color: #666; margin-top: 8px;">※ 8桁数字+チェック文字(例: 00088333P → 88333)</p>
    </div>

    <h3 style="margin: 20px 0 10px 5px;">詳細リスト</h3>
    <div id="detailsList">
        {''.join([create_detail_html(d, details) for d in details if not d['parent_id']])}
    </div>

    <div id="toast" class="toast"></div>

    <!-- 🔥 検索中ローディングオーバーレイ -->
    <div id="loadingOverlay" style="display: none; position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.7); z-index: 9999; justify-content: center; align-items: center;">
        <div style="background: white; padding: 30px 50px; border-radius: 15px; text-align: center; box-shadow: 0 10px 30px rgba(0,0,0,0.3);">
            <div style="font-size: 3em; margin-bottom: 15px;">🔍</div>
            <div style="font-size: 1.2em; font-weight: bold; color: #333;" id="loadingText">検索中...</div>
            <div style="margin-top: 15px;">
                <div style="width: 50px; height: 50px; border: 5px solid #f3f3f3; border-top: 5px solid #007bff; border-radius: 50%; animation: spin 1s linear infinite; margin: 0 auto;"></div>
            </div>
            <div style="font-size: 0.85em; color: #666; margin-top: 15px;">初回は読み込みに時間がかかります</div>
        </div>
    </div>
    <style>
        @keyframes spin {{
            0% {{ transform: rotate(0deg); }}
            100% {{ transform: rotate(360deg); }}
        }}
    </style>

    <script>
        // 🔥 自動保存用変数
        let remarksTimeout = null;

        // 🔥 ページ読み込み後にイベントリスナーを設定
        document.addEventListener('DOMContentLoaded', function() {{
            // CADリンクにイベントリスナーを追加
            document.querySelectorAll('.cad-link').forEach(function(link) {{
                link.addEventListener('click', function(e) {{
                    e.preventDefault();
                    const detailId = this.getAttribute('data-detail-id');
                    openCadFile(detailId);
                }});
            }});

            // 🔥 自動保存イベントリスナーを追加
            // 場所の変更時に自動保存
            document.getElementById('floorInput').addEventListener('change', function() {{
                autoSave();
            }});

            // パレット番号の変更時に自動保存
            document.getElementById('palletInput').addEventListener('change', function() {{
                autoSave();
            }});

            // 備考の変更時に自動保存（debounce）
            document.getElementById('remarksInput').addEventListener('input', function() {{
                clearTimeout(remarksTimeout);
                remarksTimeout = setTimeout(function() {{
                    autoSave();
                }}, 1000);  // 1秒後に保存
            }});
        }});

        // 🔥 自動保存関数
        async function autoSave() {{
            const floor = document.getElementById('floorInput').value;
            const palletNumber = document.getElementById('palletInput').value;
            const remarks = document.getElementById('remarksInput').value;

            try {{
                const response = await fetch('/api/order/{order.id}/update', {{
                    method: 'POST',
                    headers: {{ 'Content-Type': 'application/json' }},
                    body: JSON.stringify({{
                        floor: floor,
                        pallet_number: palletNumber,
                        remarks: remarks
                    }})
                }});

                const data = await response.json();

                if (data.success) {{
                    const indicator = document.getElementById('autoSaveIndicator');
                    indicator.style.display = 'block';
                    setTimeout(function() {{
                        indicator.style.display = 'none';
                    }}, 2000);
                }} else {{
                    showToast('❌ 自動保存エラー: ' + data.error, 'error');
                }}
            }} catch (error) {{
                showToast('❌ 自動保存エラー: ' + error, 'error');
            }}
        }}
        
        // CADファイルを開く関数
        async function openCadFile(detailId) {{
            try {{
                const response = await fetch('/api/open-cad/' + detailId);
                
                // JSONレスポンスの場合（ローカルで直接起動した場合）
                if (response.headers.get('content-type')?.includes('application/json')) {{
                    const data = await response.json();
                    
                    if (data.success) {{
                        if (data.opened_locally) {{
                            showToast('🔧 ' + data.message + ': ' + data.file_name, 'success');
                        }} else {{
                            showToast('📄 ' + data.message, 'success');
                        }}
                    }} else {{
                        showToast('❌ ' + data.error, 'error');
                    }}
                }} else {{
                    // ファイルダウンロード/表示の場合
                    const contentType = response.headers.get('content-type');
                    
                    if (contentType.includes('application/pdf')) {{
                        // PDFは新しいタブで開く
                        const url = `/api/open-cad/${{detailId}}`;
                        window.open(url, '_blank');
                        showToast('📄 PDF図面を開きました', 'success');
                    }} else {{
                        // MX2ファイルをダウンロード
                        const blob = await response.blob();
                        const url = window.URL.createObjectURL(blob);
                        const a = document.createElement('a');
                        a.href = url;
                        a.download = response.headers.get('content-disposition')?.split('filename=')[1] || 'file.mx2';
                        document.body.appendChild(a);
                        a.click();
                        a.remove();
                        window.URL.revokeObjectURL(url);
                        showToast('💾 MX2ファイルをダウンロードしました', 'info');
                    }}
                }}
                
            }} catch (error) {{
                showToast('❌ ファイルを開けませんでした: ' + error, 'error');
            }}
        }}
        
        // 🔥 統合保存関数（保管場所と備考を一度に保存）
        async function saveAll() {{
            const floor = document.getElementById('floorInput').value;
            const palletNumber = document.getElementById('palletInput').value;
            const remarks = document.getElementById('remarksInput').value;

            try {{
                const response = await fetch('/api/order/{order.id}/update', {{
                    method: 'POST',
                    headers: {{ 'Content-Type': 'application/json' }},
                    body: JSON.stringify({{
                        floor: floor,
                        pallet_number: palletNumber,
                        remarks: remarks
                    }})
                }});

                const data = await response.json();

                if (data.success) {{
                    showToast('✅ 保存しました', 'success');
                }} else {{
                    showToast('❌ エラー: ' + data.error, 'error');
                }}
            }} catch (error) {{
                showToast('❌ 保存エラー: ' + error, 'error');
            }}
        }}

        // 🔥 バーコード検証・処理関数
        function validateBarcode(barcode) {{
            // 前後の空白を除去し、大文字に変換
            barcode = barcode.trim().toUpperCase();

            // 長さチェック: 9文字（8桁数字 + 1文字アルファベット）
            if (barcode.length !== 9) {{
                return {{ valid: false, error: '長さが不正です（9文字必要）', orderNumber: null }};
            }}

            const digits = barcode.substring(0, 8);
            const checkChar = barcode.charAt(8);

            // 8桁が全て数字かチェック
            if (!/^\\d{{8}}$/.test(digits)) {{
                return {{ valid: false, error: '数字部分に不正な文字が含まれています', orderNumber: null }};
            }}

            // チェック文字がアルファベットかチェック
            if (!/^[A-Z]$/.test(checkChar)) {{
                return {{ valid: false, error: 'チェック文字がアルファベットではありません', orderNumber: null }};
            }}

            // チェックディジット計算: (各桁の合計 + 16) mod 26 = アルファベット位置
            // これは (合計 - 10 + 26) mod 26 と等価で、合計が10未満でも正しく計算できる
            let digitSum = 0;
            for (let i = 0; i < 8; i++) {{
                digitSum += parseInt(digits.charAt(i), 10);
            }}
            const expectedCharCode = 65 + ((digitSum + 16) % 26);  // A=65, mod 26で循環
            const expectedChar = String.fromCharCode(expectedCharCode);

            if (checkChar !== expectedChar) {{
                return {{
                    valid: false,
                    error: 'チェックディジット不一致（期待: ' + expectedChar + ', 実際: ' + checkChar + '）',
                    orderNumber: null
                }};
            }}

            // 先頭の0を除いた発注番号を返す
            const orderNumber = digits.replace(/^0+/, '');
            return {{ valid: true, error: null, orderNumber: orderNumber }};
        }}

        // 🔥 ローディング表示/非表示関数
        function showLoading(message = '検索中...') {{
            const overlay = document.getElementById('loadingOverlay');
            const text = document.getElementById('loadingText');
            text.textContent = message;
            overlay.style.display = 'flex';
        }}

        function hideLoading() {{
            document.getElementById('loadingOverlay').style.display = 'none';
        }}

        async function processBarcode() {{
            const input = document.getElementById('barcodeInput');
            const resultDiv = document.getElementById('barcodeResult');
            const barcode = input.value;

            if (!barcode) {{
                resultDiv.style.display = 'none';
                return;
            }}

            const result = validateBarcode(barcode);

            if (!result.valid) {{
                // エラー表示
                resultDiv.style.display = 'block';
                resultDiv.style.background = '#f8d7da';
                resultDiv.style.color = '#721c24';
                resultDiv.style.border = '1px solid #f5c6cb';
                resultDiv.innerHTML = '❌ <strong>無効なバーコード</strong><br>' + result.error + '<br>再スキャンしてください';
                input.value = '';
                input.focus();

                // バイブレーション（エラー）
                if (navigator.vibrate) navigator.vibrate([100, 50, 100]);
                return;
            }}

            // 成功: 発注番号で検索
            const orderNumber = result.orderNumber;

            // まずページ内検索を試行
            const foundInPage = highlightAndScrollToItem(orderNumber);

            if (foundInPage) {{
                // ページ内で見つかった場合
                resultDiv.style.display = 'block';
                resultDiv.style.background = '#d4edda';
                resultDiv.style.color = '#155724';
                resultDiv.style.border = '1px solid #c3e6cb';
                resultDiv.innerHTML = '✅ <strong>発注番号: ' + orderNumber + '</strong>';
                if (navigator.vibrate) navigator.vibrate(100);
            }} else {{
                // ページ内で見つからない場合、API検索
                showLoading('発注番号 ' + orderNumber + ' を検索中...');

                try {{
                    const response = await fetch('/api/search-by-purchase-order/' + orderNumber);
                    const data = await response.json();
                    hideLoading();

                    if (data.results && data.results.length > 0) {{
                        // API検索で見つかった
                        const item = data.results[0];
                        resultDiv.style.display = 'block';
                        resultDiv.style.background = '#cce5ff';
                        resultDiv.style.color = '#004085';
                        resultDiv.style.border = '1px solid #b8daff';
                        resultDiv.innerHTML = '📋 <strong>発注番号: ' + orderNumber + '</strong><br>' +
                            '製番: ' + (item['製番'] || '-') + '<br>' +
                            '品名: ' + (item['品名'] || '-') + '<br>' +
                            '仕入先: ' + (item['仕入先略称'] || '-') + '<br>' +
                            '<span style="color:#856404;">⚠️ このユニットには含まれていません</span>';
                        if (navigator.vibrate) navigator.vibrate([50, 30, 50]);
                    }} else {{
                        // どこにも見つからない
                        resultDiv.style.display = 'block';
                        resultDiv.style.background = '#fff3cd';
                        resultDiv.style.color = '#856404';
                        resultDiv.style.border = '1px solid #ffeeba';
                        resultDiv.innerHTML = '⚠️ <strong>発注番号: ' + orderNumber + '</strong><br>データが見つかりません';
                        if (navigator.vibrate) navigator.vibrate([100, 50, 100]);
                    }}
                }} catch (error) {{
                    hideLoading();
                    resultDiv.style.display = 'block';
                    resultDiv.style.background = '#f8d7da';
                    resultDiv.style.color = '#721c24';
                    resultDiv.style.border = '1px solid #f5c6cb';
                    resultDiv.innerHTML = '❌ 検索エラー: ' + error.message;
                }}
            }}

            // 入力をクリアして次のスキャンに備える
            input.value = '';
            input.focus();
        }}

        function highlightAndScrollToItem(orderNumber) {{
            // 全てのハイライトを解除
            document.querySelectorAll('.detail-item').forEach(item => {{
                item.style.boxShadow = '';
                item.style.border = '';
            }});

            // 発注番号が一致するアイテムを探す
            let found = false;
            document.querySelectorAll('.detail-item').forEach(item => {{
                const text = item.textContent;
                // 発注番号: XXXXX の形式で検索
                if (text.includes('発注番号: ' + orderNumber) || text.includes('発注番号:' + orderNumber)) {{
                    found = true;
                    // ハイライト
                    item.style.boxShadow = '0 0 15px 5px rgba(0, 123, 255, 0.5)';
                    item.style.border = '3px solid #007bff';
                    // スクロール
                    item.scrollIntoView({{ behavior: 'smooth', block: 'center' }});

                    // 5秒後にハイライト解除
                    setTimeout(() => {{
                        item.style.boxShadow = '';
                        item.style.border = '';
                    }}, 5000);
                }}
            }});

            // 見つかったかどうかを返す
            return found;
        }}

        // 受入切替関数（数量入力対応）
        async function toggleReceive(detailId, setReceived, orderNumber, itemName, spec1, quantity) {{
            const action = setReceived ? '受入' : '受入取消';

            // 手配数量を数値として抽出
            const expectedQty = parseInt((quantity || '0').toString().replace(/[^0-9]/g, ''), 10) || 0;

            if (setReceived) {{
                // 受入時：数量入力モーダルを表示
                const infoText = '発注番号: ' + (orderNumber || '未設定') + '\\n' +
                    '品名: ' + (itemName || '未設定') + '\\n' +
                    '仕様１: ' + (spec1 || '未設定') + '\\n' +
                    '手配数: ' + expectedQty;

                const inputQty = prompt(
                    '受入数量を入力してください。\\n' +
                    '（全数受入の場合は空欄またはそのままOK）\\n\\n' +
                    infoText,
                    expectedQty.toString()
                );

                if (inputQty === null) {{
                    return; // キャンセル
                }}

                // 数量のパース（空欄の場合は全数受入）
                let receivedQty = null;
                if (inputQty.trim() !== '' && inputQty.trim() !== expectedQty.toString()) {{
                    receivedQty = parseInt(inputQty.trim(), 10);
                    if (isNaN(receivedQty) || receivedQty < 0) {{
                        showToast('❌ 数量は0以上の数値を入力してください', 'error');
                        return;
                    }}

                    // 不足・超過の確認
                    if (receivedQty !== expectedQty) {{
                        const diff = expectedQty - receivedQty;
                        let confirmMsg;
                        if (diff > 0) {{
                            confirmMsg = '手配数より ' + diff + '個 不足しています。\\n不足分は備考に自動記録されます。\\nこのまま受入しますか？';
                        }} else {{
                            confirmMsg = '手配数より ' + (-diff) + '個 超過しています。\\n超過分は備考に自動記録されます。\\nこのまま受入しますか？';
                        }}
                        if (!confirm(confirmMsg)) {{
                            return;
                        }}
                    }}
                }}

                // 受入API呼び出し
                try {{
                    const body = {{ is_received: true }};
                    if (receivedQty !== null) {{
                        body.received_quantity = receivedQty;
                    }}

                    const response = await fetch('/api/detail/' + detailId + '/receive-with-quantity', {{
                        method: 'POST',
                        headers: {{ 'Content-Type': 'application/json' }},
                        body: JSON.stringify(body)
                    }});

                    const data = await response.json();

                    if (data.success) {{
                        showToast(data.message || '✅ 受入しました');
                        setTimeout(function() {{ location.reload(); }}, 1000);
                    }} else {{
                        showToast('❌ エラー: ' + (data.error || '不明なエラー'), 'error');
                    }}
                }} catch (error) {{
                    showToast('❌ ネットワークエラー: ' + error, 'error');
                    console.error('Error:', error);
                }}
            }} else {{
                // 受入取消時
                const confirmMessage = 'このアイテムの受入を取り消しますか？\\n\\n' +
                    '発注番号: ' + (orderNumber || '未設定') + '\\n' +
                    '品名: ' + (itemName || '未設定') + '\\n' +
                    '仕様１: ' + (spec1 || '未設定');

                if (!confirm(confirmMessage)) {{
                    return;
                }}

                try {{
                    const response = await fetch('/api/detail/' + detailId + '/receive-with-quantity', {{
                        method: 'POST',
                        headers: {{ 'Content-Type': 'application/json' }},
                        body: JSON.stringify({{ is_received: false }})
                    }});

                    const data = await response.json();

                    if (data.success) {{
                        showToast('⚠️ 受入を取り消しました');
                        setTimeout(function() {{ location.reload(); }}, 1000);
                    }} else {{
                        showToast('❌ エラー: ' + (data.error || '不明なエラー'), 'error');
                    }}
                }} catch (error) {{
                    showToast('❌ ネットワークエラー: ' + error, 'error');
                    console.error('Error:', error);
                }}
            }}
        }}
        
        // トースト表示関数
        function showToast(message, type) {{
            type = type || 'success';
            const toast = document.getElementById('toast');
            toast.textContent = message;

            if (type === 'error') {{
                toast.style.background = '#dc3545';
            }} else if (type === 'info') {{
                toast.style.background = '#17a2b8';
            }} else {{
                toast.style.background = '#28a745';
            }}

            toast.classList.add('show');
            setTimeout(function() {{
                toast.classList.remove('show');
            }}, 3000);
        }}

        // 📷 画像アップロード（ファイル選択）
        async function uploadOrderImage(orderId) {{
            const fileInput = document.getElementById('imageUploadFile');
            await processImageUpload(orderId, fileInput);
        }}

        // 📸 画像アップロード（カメラ撮影）
        async function uploadOrderImageFromCamera(orderId) {{
            const fileInput = document.getElementById('imageUploadCamera');
            await processImageUpload(orderId, fileInput);
        }}

        // 画像アップロード処理
        async function processImageUpload(orderId, fileInput) {{
            const file = fileInput.files[0];

            if (!file) {{
                return;
            }}

            showToast('📤 アップロード中...', 'info');

            const formData = new FormData();
            formData.append('image', file);

            try {{
                const response = await fetch('/api/order/' + orderId + '/upload-image', {{
                    method: 'POST',
                    body: formData
                }});

                const data = await response.json();

                if (data.success) {{
                    showToast('✅ 画像をアップロードしました', 'success');
                    // 画像を再読み込み
                    const img = document.getElementById('orderImage');
                    img.src = '/api/order/' + orderId + '/image?t=' + Date.now();
                }} else {{
                    showToast('❌ エラー: ' + data.error, 'error');
                }}
            }} catch (error) {{
                showToast('❌ アップロードエラー: ' + error, 'error');
            }}

            // ファイル選択をリセット
            fileInput.value = '';
        }}

        // 🗑️ 画像削除
        async function deleteOrderImage(orderId) {{
            if (!confirm('画像を削除しますか？')) {{
                return;
            }}

            try {{
                const response = await fetch('/api/order/' + orderId + '/delete-image', {{
                    method: 'DELETE'
                }});

                const data = await response.json();

                if (data.success) {{
                    showToast('✅ 画像を削除しました', 'success');
                    document.getElementById('orderImage').style.display = 'none';
                    document.getElementById('noImageText').style.display = 'block';
                }} else {{
                    showToast('❌ エラー: ' + data.error, 'error');
                }}
            }} catch (error) {{
                showToast('❌ 削除エラー: ' + error, 'error');
            }}
        }}

        // 🔍 画像をフルスクリーンで表示
        function openImageFullscreen(src) {{
            const overlay = document.createElement('div');
            overlay.style.cssText = 'position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.95); display: flex; justify-content: center; align-items: center; z-index: 10000; cursor: pointer;';

            const img = document.createElement('img');
            img.src = src;
            img.style.cssText = 'max-width: 95%; max-height: 95%; object-fit: contain;';

            const closeBtn = document.createElement('div');
            closeBtn.innerHTML = '✕';
            closeBtn.style.cssText = 'position: absolute; top: 15px; right: 20px; color: white; font-size: 2em; cursor: pointer;';
            closeBtn.onclick = function() {{ overlay.remove(); }};

            overlay.appendChild(img);
            overlay.appendChild(closeBtn);
            overlay.onclick = function(e) {{ if (e.target === overlay) overlay.remove(); }};
            document.body.appendChild(overlay);
        }}
    </script>
</body>
</html>
"""
        return html
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"<html><body><h1>エラー</h1><p>{str(e)}</p></body></html>", 500

@app.route('/api/order/<int:order_id>/update-remarks', methods=['POST'])
def update_order_remarks(order_id):
    """備考のみを更新するAPI"""
    try:
        data = request.json
        order = Order.query.get_or_404(order_id)
        
        order.remarks = data.get('remarks', '')
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '備考を更新しました'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

def create_detail_html(detail, all_details):
    """詳細アイテムのHTML生成"""
    is_received = detail['is_received']
    has_children = any(d['parent_id'] == detail['id'] for d in all_details)
    
    def escape_js(text):
        if not text:
            return ''
        return str(text).replace("'", "\\'").replace('"', '\\"').replace('\n', '\\n')
    
    order_number = escape_js(detail.get('order_number', ''))
    item_name = escape_js(detail.get('item_name', ''))
    spec1 = escape_js(detail.get('spec1', ''))
    quantity_str = f"{detail.get('quantity', '')} {detail.get('unit_measure', '')}".strip()
    
    # CAD図面情報を取得
    cad_info = get_cad_file_info(detail.get('spec1', ''))
    spec1_display = detail.get('spec1', '-')
    
    # 🔥 仕様1の表示（data属性を使用）
    if cad_info:
        if cad_info['has_pdf']:
            file_info = f"📄 PDF有 ({len(cad_info['pdf_files'])}件)"
            spec1_html = f'''
            <div>
                <strong>仕様１:</strong> 
                <a href="#" class="cad-link" data-detail-id="{detail['id']}" 
                   style="color: #007bff; text-decoration: underline; cursor: pointer;">
                    {spec1_display}
                </a>
                <span style="font-size: 0.8em; color: #28a745; margin-left: 5px;">{file_info}</span>
            </div>
            '''
        elif cad_info['has_mx2']:
            file_info = f"🔧 mx2のみ ({len(cad_info['mx2_files'])}件)"
            spec1_html = f'''
            <div>
                <strong>仕様１:</strong> 
                <a href="#" class="cad-link" data-detail-id="{detail['id']}" 
                   style="color: #007bff; text-decoration: underline; cursor: pointer;">
                    {spec1_display}
                </a>
                <div style="font-size: 0.75em; color: #856404; margin-top: 3px;">
                    {file_info}<br>
                    ⚠️ iCAD MX導入PCのみ(ダウンロードファイル)
                </div>
            </div>
            '''
        else:
            spec1_html = f'<div><strong>仕様１:</strong> {spec1_display}</div>'
    else:
        spec1_html = f'<div><strong>仕様１:</strong> {spec1_display}</div>'
    
    # 受入数量の表示テキストを生成
    received_qty_html = ''
    if is_received:
        expected_qty = detail.get('quantity') or 0
        received_qty = detail.get('received_quantity')
        if received_qty is not None and received_qty != expected_qty:
            diff = expected_qty - received_qty
            if diff > 0:
                received_qty_html = f'<div style="background: #f8d7da; padding: 8px; border-radius: 5px; margin: 10px 0; font-size: 0.9em; border-left: 3px solid #dc3545;"><strong>受入数量:</strong> {received_qty}個 <span style="color: #dc3545; font-weight: bold;">（不足 {diff}個）</span></div>'
            else:
                received_qty_html = f'<div style="background: #fff3cd; padding: 8px; border-radius: 5px; margin: 10px 0; font-size: 0.9em; border-left: 3px solid #ffc107;"><strong>受入数量:</strong> {received_qty}個 <span style="color: #856404; font-weight: bold;">（超過 {-diff}個）</span></div>'
        elif received_qty is not None:
            received_qty_html = f'<div style="background: #d4edda; padding: 8px; border-radius: 5px; margin: 10px 0; font-size: 0.9em; border-left: 3px solid #28a745;"><strong>受入数量:</strong> {received_qty}個 <span style="color: #155724;">（全数）</span></div>'

    # 親アイテムのHTML
    html = f"""
    <div class="detail-item {'received' if is_received else ''}">
        <div class="detail-header">
            <div class="item-name">{detail['item_name'] or '-'}</div>
            {f'<span class="status-badge badge-success">✅ 受入済</span>' if is_received else ''}
        </div>

        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 5px 15px; font-size: 0.9em; margin: 10px 0;">
            <div><strong>発注番号:</strong> {detail['order_number'] or '-'}</div>
            <div><strong>納期:</strong> {detail['delivery_date'] or '-'}</div>
            {spec1_html}
            <div><strong>手配数:</strong> {detail['quantity'] or ''} {detail['unit_measure'] or ''}</div>
            <div><strong>仕入先:</strong> {detail['supplier'] or '-'}</div>
            <div><strong>手配区分:</strong> {detail['order_type'] or '-'}</div>
        </div>

        {received_qty_html}

        {f'<div style="background: #e3f2fd; padding: 8px; border-radius: 5px; margin: 10px 0; font-size: 0.85em; border-left: 3px solid #2196f3;"><strong>📝納品書入力日:</strong> {detail.get("received_delivery_date", "-")} / {int(detail.get("received_delivery_qty", 0)) if detail.get("received_delivery_qty") else "-"}個</div>' if detail.get('received_delivery_qty') else ''}

        {f'<div style="background: #fff3cd; padding: 8px; border-radius: 5px; margin: 10px 0; font-size: 0.9em;"><strong>備考:</strong> {detail["remarks"]}</div>' if detail.get('remarks') else ''}
        {f'<span class="status-badge badge-warning">追加工有</span>' if has_children else ''}

        <button class="btn {'btn-warning' if is_received else 'btn-primary'}"
                onclick="toggleReceive({detail['id']}, {str(not is_received).lower()}, '{order_number}', '{item_name}', '{spec1}', '{quantity_str}')">
            {('受入取消' if is_received else '受入')}
        </button>
    </div>
    """
    
    # 子アイテムも同様に処理
    children = [d for d in all_details if d['parent_id'] == detail['id']]
    for child in children:
        child_received = child['is_received']
        
        child_order_number = escape_js(child.get('order_number', ''))
        child_item_name = escape_js(child.get('item_name', ''))
        child_spec1 = escape_js(child.get('spec1', ''))
        child_quantity_str = f"{child.get('quantity', '')} {child.get('unit_measure', '')}".strip()
        
        child_cad_info = get_cad_file_info(child.get('spec1', ''))
        child_spec1_display = child.get('spec1', '-')
        
        if child_cad_info:
            if child_cad_info['has_pdf']:
                child_file_info = f"📄 PDF有"
                child_spec1_html = f'''
                <div>
                    <strong>仕様１:</strong> 
                    <a href="#" class="cad-link" data-detail-id="{child['id']}" 
                       style="color: #007bff; text-decoration: underline; cursor: pointer;">
                        {child_spec1_display}
                    </a>
                    <span style="font-size: 0.8em; color: #28a745; margin-left: 5px;">{child_file_info}</span>
                </div>
                '''
            elif child_cad_info['has_mx2']:
                child_file_info = f"🔧 mx2のみ"
                child_spec1_html = f'''
                <div>
                    <strong>仕様１:</strong> 
                    <a href="#" class="cad-link" data-detail-id="{child['id']}" 
                       style="color: #007bff; text-decoration: underline; cursor: pointer;">
                        {child_spec1_display}
                    </a>
                    <div style="font-size: 0.75em; color: #856404; margin-top: 3px;">
                        {child_file_info}<br>
                        ⚠️ iCAD MX導入PCのみ
                    </div>
                </div>
                '''
            else:
                child_spec1_html = f'<div><strong>仕様１:</strong> {child_spec1_display}</div>'
        else:
            child_spec1_html = f'<div><strong>仕様１:</strong> {child_spec1_display}</div>'
        
        # 子アイテムの受入数量表示
        child_received_qty_html = ''
        if child_received:
            child_expected_qty = child.get('quantity') or 0
            child_recv_qty = child.get('received_quantity')
            if child_recv_qty is not None and child_recv_qty != child_expected_qty:
                child_diff = child_expected_qty - child_recv_qty
                if child_diff > 0:
                    child_received_qty_html = f'<div style="background: #f8d7da; padding: 6px; border-radius: 5px; margin: 8px 0; font-size: 0.85em; border-left: 3px solid #dc3545;"><strong>受入数量:</strong> {child_recv_qty}個 <span style="color: #dc3545; font-weight: bold;">（不足 {child_diff}個）</span></div>'
                else:
                    child_received_qty_html = f'<div style="background: #fff3cd; padding: 6px; border-radius: 5px; margin: 8px 0; font-size: 0.85em; border-left: 3px solid #ffc107;"><strong>受入数量:</strong> {child_recv_qty}個 <span style="color: #856404; font-weight: bold;">（超過 {-child_diff}個）</span></div>'
            elif child_recv_qty is not None:
                child_received_qty_html = f'<div style="background: #d4edda; padding: 6px; border-radius: 5px; margin: 8px 0; font-size: 0.85em; border-left: 3px solid #28a745;"><strong>受入数量:</strong> {child_recv_qty}個 <span style="color: #155724;">（全数）</span></div>'

        html += f"""
    <div class="detail-item child {'received' if child_received else ''}">
        <div class="detail-header">
            <div class="item-name">└─ {child['item_name'] or '-'}</div>
            {f'<span class="status-badge badge-success">✅ 受入済</span>' if child_received else ''}
        </div>

        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 5px 15px; font-size: 0.9em; margin: 10px 0;">
            <div><strong>発注番号:</strong> {child['order_number'] or '-'}</div>
            <div><strong>納期:</strong> {child['delivery_date'] or '-'}</div>
            {child_spec1_html}
            <div><strong>手配数:</strong> {child['quantity'] or ''} {child['unit_measure'] or ''}</div>
            <div><strong>仕入先:</strong> {child['supplier'] or '-'}</div>
            <div><strong>手配区分:</strong> {child['order_type'] or '-'}</div>
        </div>

        {child_received_qty_html}

        <button class="btn {'btn-warning' if child_received else 'btn-primary'}"
                onclick="toggleReceive({child['id']}, {str(not child_received).lower()}, '{child_order_number}', '{child_item_name}', '{child_spec1}', '{child_quantity_str}')">
            {'受入取消' if child_received else '受入'}
        </button>
    </div>
        """
    
    return html

@app.route('/api/order/<int:order_id>')
def get_order_details(order_id):
    """Get order details"""
    try:
        from urllib.parse import quote
        order = Order.query.get_or_404(order_id)

        # 🔥 検収データを読み込み
        delivery_dict = DeliveryUtils.load_delivery_data()

        details = []
        for detail in order.details:
            # 🔥 検収データを取得
            delivery_info = DeliveryUtils.get_delivery_info(detail.order_number, delivery_dict)

            detail_dict = {
                'id': detail.id,
                'delivery_date': detail.delivery_date,
                'reply_delivery_date': detail.reply_delivery_date or '',
                'supplier': detail.supplier,
                'order_number': detail.order_number,
                'quantity': detail.quantity,
                'unit_measure': detail.unit_measure,
                'item_name': detail.item_name,
                'spec1': detail.spec1,
                'spec2': detail.spec2,
                'order_type': detail.order_type,
                'remarks': detail.remarks,
                'is_received': detail.is_received,
                'received_quantity': detail.received_quantity,  # 実際の受入数量
                'received_at': detail.received_at.isoformat() if detail.received_at else None,
                'has_internal_processing': detail.has_internal_processing,
                'parent_id': detail.parent_id,  # 🔥 親子関係を追加
                # 🔥 検収データを追加
                'received_delivery_date': delivery_info.get('納入日', ''),
                'received_delivery_qty': delivery_info.get('納入数', 0),
                # CAD情報は遅延ロード（/api/detail/{id}/cad-info で取得）
                'cad_info': None
            }

            details.append(detail_dict)
        
        return jsonify({
            'order': {
                'id': order.id,
                'seiban': order.seiban,
                'created_at': to_jst(order.created_at).strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': to_jst(order.updated_at).strftime('%Y-%m-%d %H:%M:%S'),
                'unit': order.unit or '',  # 空の場合は空文字列を返す
                'product_name': order.product_name or '',
                'customer_abbr': order.customer_abbr or '',
                'pallet_number': order.pallet_number,
                'floor': order.floor,
                'memo2': order.memo2 or '',
                'status': order.status,
                'location': order.location,
                'remarks': order.remarks
            },
            'details': details,
            'qr_code': generate_qr_code(f"{get_server_url()}/receive/{order.seiban}/{quote(order.unit, safe='') if order.unit else ''}")
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/order/<int:order_id>/update', methods=['POST'])
def update_order(order_id):
    """Update order status, location, pallet and floor"""
    try:
        order = Order.query.get_or_404(order_id)
        data = request.json
        
        # 🔥 ステータス変更前の値を保存
        old_status = order.status
        new_status = data.get('status', order.status)
        
        # 更新処理
        if 'status' in data:
            order.status = data['status']
        if 'location' in data:
            order.location = data['location']
        if 'remarks' in data:
            order.remarks = data['remarks']
        if 'pallet_number' in data:
            order.pallet_number = data['pallet_number']
        if 'floor' in data:
            order.floor = data['floor']
        
        order.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        # 🔥 Excelファイルも更新（対象ユニットのみ・非同期）
        _oid = order_id
        def _bg_update():
            try:
                with app.app_context():
                    update_unit_excel_only(_oid)
            except Exception as excel_error:
                print(f"⚠️ Excel更新エラー（DB保存は成功）: {excel_error}")
        Thread(target=_bg_update, daemon=True).start()

        # 🔥 納品完了になった場合の処理
        response_data = {
            'success': True,
            'message': 'Order updated successfully'
        }
        
        if old_status != '納品完了' and new_status == '納品完了':
            # メール送信を促す
            response_data['show_email_prompt'] = True
            response_data['order_info'] = {
                'seiban': order.seiban,
                'product_name': order.product_name or '',
                'customer_abbr': order.customer_abbr or '',
                'floor': order.floor or '',
                'pallet_number': order.pallet_number or '',
                'order_id': order.id
            }
        
        return jsonify(response_data)
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/detail/<int:detail_id>/toggle-receive', methods=['POST'])
def toggle_receive_detail(detail_id):
    """Toggle receive status for a detail item"""
    try:
        detail = OrderDetail.query.get_or_404(detail_id)

        # 現在の状態を取得
        was_received = detail.is_received
        action = 'unreceive' if was_received else 'receive'

        # ステータスをトグル
        detail.is_received = not was_received
        detail.received_at = None if not detail.is_received else datetime.now(timezone.utc)

        # クライアントIPを取得
        client_ip = request.headers.get('X-Forwarded-For', request.remote_addr)
        if ',' in client_ip:
            client_ip = client_ip.split(',')[0].strip()

        # 🔥 受入履歴を記録（発注番号がある場合のみ）
        if detail.order_number:
            if detail.is_received:
                ReceivedHistory.record_receive(
                    order_number=detail.order_number,
                    item_name=detail.item_name,
                    spec1=detail.spec1,
                    quantity=detail.quantity,
                    client_ip=client_ip
                )
            else:
                ReceivedHistory.record_cancel(
                    order_number=detail.order_number,
                    item_name=detail.item_name,
                    spec1=detail.spec1,
                    quantity=detail.quantity,
                    client_ip=client_ip
                )

        # 編集ログを記録
        log = EditLog(
            detail_id=detail_id,
            action=action,
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string if request.user_agent else 'Unknown'
        )
        db.session.add(log)
        
        # 注文全体のステータスを更新（1回のループで計算）
        order = detail.order
        details_list = order.details
        total_count = len(details_list)
        received_count = sum(1 for d in details_list if d.is_received)

        if received_count == total_count:
            order.status = '納品完了'
        elif received_count > 0:
            order.status = '納品中'
        else:
            order.status = '受入準備前'
        
        order.updated_at = datetime.now(timezone.utc)
        
        db.session.commit()

        # Excelファイルを非同期で軽量更新（対象ユニットのシートのみ）
        _order_id = order.id
        def _bg_excel_update():
            try:
                with app.app_context():
                    update_unit_excel_only(_order_id)
            except Exception as excel_error:
                print(f"⚠️ Excel更新エラー（DB保存は成功）: {excel_error}")
        Thread(target=_bg_excel_update, daemon=True).start()

        # メッセージ作成（詳細情報を含む）
        if detail.is_received:
            message = f'✅ 受入完了\n'
        else:
            message = f'❌ 受入取消\n'
    
        # 社内加工の警告
        has_internal = False
        if detail.has_internal_processing:
            message += '\n\n⚠️ 注意: 社内加工/追加工品です'
            has_internal = True
        
        return jsonify({
            'success': True,
            'message': message,
            'is_received': detail.is_received,
            'order_status': order.status,
            'has_internal_processing': has_internal
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/detail/<int:detail_id>/receive', methods=['POST'])
def receive_detail(detail_id):
    """Mark detail as received (deprecated - use toggle-receive instead)"""
    return toggle_receive_detail(detail_id)


@app.route('/api/detail/<int:detail_id>/receive-with-quantity', methods=['POST'])
def receive_detail_with_quantity(detail_id):
    """数量指定での受入処理
    リクエストボディ:
    {
        "received_quantity": 10,  // 実際に受け入れた数量（Nullまたは省略で全数受入）
        "is_received": true       // true=受入、false=取消
    }
    """
    try:
        detail = OrderDetail.query.get_or_404(detail_id)
        data = request.get_json() or {}

        is_received = data.get('is_received', True)
        received_quantity = data.get('received_quantity')

        # 数量のバリデーション
        if received_quantity is not None:
            try:
                received_quantity = int(received_quantity)
                if received_quantity < 0:
                    return jsonify({'success': False, 'error': '数量は0以上で入力してください'}), 400
            except (ValueError, TypeError):
                return jsonify({'success': False, 'error': '数量は数値で入力してください'}), 400

        # 現在の状態を取得
        was_received = detail.is_received
        action = 'receive' if is_received else 'unreceive'

        # ステータスを更新
        detail.is_received = is_received
        detail.received_at = datetime.now(timezone.utc) if is_received else None

        # 受入数量を設定
        if is_received:
            detail.received_quantity = received_quantity  # Noneの場合は全数受入扱い
        else:
            detail.received_quantity = None

        # 不足時の備考追加処理
        shortage_note = ''
        if is_received and received_quantity is not None and detail.quantity:
            shortage = detail.quantity - received_quantity
            if shortage > 0:
                # 不足がある場合、備考に追加
                shortage_note = f"【不足：{shortage}個】"
                existing_remarks = detail.remarks or ''

                # 既存の不足備考を削除（重複防止）
                existing_remarks = re.sub(r'【不足：\d+個】', '', existing_remarks).strip()

                # 新しい備考を設定
                if existing_remarks:
                    detail.remarks = f"{shortage_note} {existing_remarks}"
                else:
                    detail.remarks = shortage_note
            elif shortage < 0:
                # 超過の場合
                overage = -shortage
                shortage_note = f"【超過：{overage}個】"
                existing_remarks = detail.remarks or ''
                existing_remarks = re.sub(r'【(不足|超過)：\d+個】', '', existing_remarks).strip()
                if existing_remarks:
                    detail.remarks = f"{shortage_note} {existing_remarks}"
                else:
                    detail.remarks = shortage_note
            else:
                # 過不足なしの場合、不足/超過備考を削除
                if detail.remarks:
                    detail.remarks = re.sub(r'【(不足|超過)：\d+個】\s*', '', detail.remarks).strip()

        # クライアントIPを取得
        client_ip = request.headers.get('X-Forwarded-For', request.remote_addr)
        if ',' in client_ip:
            client_ip = client_ip.split(',')[0].strip()

        # 受入履歴を記録
        if detail.order_number:
            if is_received:
                ReceivedHistory.record_receive(
                    order_number=detail.order_number,
                    item_name=detail.item_name,
                    spec1=detail.spec1,
                    quantity=detail.quantity,
                    client_ip=client_ip,
                    received_quantity=received_quantity
                )
            else:
                ReceivedHistory.record_cancel(
                    order_number=detail.order_number,
                    item_name=detail.item_name,
                    spec1=detail.spec1,
                    quantity=detail.quantity,
                    client_ip=client_ip
                )

        # 編集ログを記録
        log = EditLog(
            detail_id=detail_id,
            action=action,
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string if request.user_agent else 'Unknown'
        )
        db.session.add(log)

        # 注文全体のステータスを更新
        order = detail.order
        details_list = order.details
        total_count = len(details_list)
        received_count = sum(1 for d in details_list if d.is_received)

        if received_count == total_count:
            order.status = '納品完了'
        elif received_count > 0:
            order.status = '納品中'
        else:
            order.status = '受入準備前'

        order.updated_at = datetime.now(timezone.utc)

        db.session.commit()

        # Excelファイルを非同期で更新
        _order_id = order.id
        def _bg_excel_update():
            try:
                with app.app_context():
                    update_unit_excel_only(_order_id)
            except Exception as excel_error:
                print(f"⚠️ Excel更新エラー（DB保存は成功）: {excel_error}")
        Thread(target=_bg_excel_update, daemon=True).start()

        # メッセージ作成
        if is_received:
            qty_msg = f"{received_quantity}" if received_quantity is not None else f"{detail.quantity}(全数)"
            message = f'✅ 受入完了 ({qty_msg}個)\n'
            if shortage_note:
                message += f'\n{shortage_note}'
        else:
            message = f'❌ 受入取消\n'

        # 社内加工の警告
        has_internal = False
        if detail.has_internal_processing:
            message += '\n\n⚠️ 注意: 社内加工/追加工品です'
            has_internal = True

        return jsonify({
            'success': True,
            'message': message,
            'is_received': detail.is_received,
            'received_quantity': detail.received_quantity,
            'expected_quantity': detail.quantity,
            'order_status': order.status,
            'has_internal_processing': has_internal,
            'remarks': detail.remarks
        })

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/detail/<int:detail_id>/logs')
def get_detail_logs(detail_id):
    """Get edit logs for a specific detail"""
    try:
        logs = EditLog.query.filter_by(detail_id=detail_id).order_by(EditLog.timestamp.desc()).all()
        
        log_data = []
        for log in logs:
            log_data.append({
                'id': log.id,
                'action': '受入' if log.action == 'receive' else '取消',
                'ip_address': log.ip_address,
                'timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'user_agent': log.user_agent[:50] if log.user_agent else 'Unknown'
            })
        
        return jsonify({
            'success': True,
            'logs': log_data,
            'total': len(log_data)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/detail/<int:detail_id>/cad-info')
def get_detail_cad_info(detail_id):
    """詳細アイテムのCADファイル情報を取得（遅延ロード用）"""
    try:
        detail = OrderDetail.query.get_or_404(detail_id)
        cad_info = get_cad_file_info(detail.spec1)

        if cad_info:
            return jsonify({
                'success': True,
                'cad_info': {
                    'has_pdf': cad_info['has_pdf'],
                    'has_mx2': cad_info['has_mx2'],
                    'pdf_count': len(cad_info['pdf_files']),
                    'mx2_count': len(cad_info['mx2_files'])
                }
            })
        else:
            return jsonify({
                'success': True,
                'cad_info': None
            })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/check-update')
def check_update():
    """ファイルの更新をチェック（廃止済み - DBから直接取得に移行）"""
    return jsonify({
        'has_update': False,
        'message': 'この機能は廃止されました。DBから直接取得してください。'
    })

@app.route('/api/export/<int:order_id>')
def export_order(order_id):
    """注文データをExcelファイルとしてエクスポート"""
    try:
        order = Order.query.get(order_id)
        if not order:
            return jsonify({'success': False, 'error': '注文が見つかりません'}), 404
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"{order.seiban}_{order.unit}"
        
        # ヘッダー
        headers = ['製番', 'ユニット', '品名', '仕様１', '仕様２', '数量', '単位', 
                   '納期', '手配区分', '発注番号', '仕入先', '仕入先CD', '備考', '検収日', '検収数']
        ws.append(headers)
        
        # データ
        for detail in order.details:
            row = [
                order.seiban,
                order.unit,
                detail.item_name,
                detail.spec1,
                detail.spec2,
                detail.quantity,
                detail.unit_measure,
                detail.delivery_date,
                detail.order_type,
                detail.order_number,
                detail.supplier,
                detail.supplier_cd,
                detail.remarks,
                detail.received_at.strftime('%Y-%m-%d %H:%M:%S') if detail.received_at else '',
                '受入済' if detail.is_received else '未受入'
            ]
            ws.append(row)
        
        # メモリ上に保存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()
        
        filename = f"{order.seiban}_{order.unit}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/order/<int:order_id>/delete', methods=['DELETE'])
def delete_order(order_id):
    """Delete an order and its details"""
    try:
        order = Order.query.get_or_404(order_id)
        
        # Delete all details first
        OrderDetail.query.filter_by(order_id=order_id).delete()
        
        # Delete the order
        db.session.delete(order)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Order deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/search-by-spec1/<spec1>')
def search_by_spec1(spec1):
    """仕様１で検索"""
    try:
        # マージ済みデータから検索
        details = OrderDetail.query.filter(
            OrderDetail.spec1.contains(spec1)
        ).all()

        result_list = []

        for detail in details:
            result_list.append({
                'id': detail.id,
                'order_id': detail.order_id,
                'seiban': detail.seiban,
                'unit': detail.material,
                'item_name': detail.item_name,
                'spec1': detail.spec1,
                'order_number': detail.order_number,
                'quantity': detail.quantity,
                'unit_measure': detail.unit_measure,
                'is_received': detail.is_received,
                'delivery_date': detail.delivery_date,
                'reply_delivery_date': detail.reply_delivery_date or '',
                'supplier': detail.supplier,
                'staff': '',
                'source': 'merged'
            })

        if not result_list:
            return jsonify({
                'found': False,
                'message': f'仕様１ "{spec1}" が見つかりません'
            }), 404

        return jsonify({
            'found': True,
            'count': len(result_list),
            'details': result_list
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/search-by-purchase-order/<purchase_order_number>')
def search_by_purchase_order(purchase_order_number):
    """発注番号で検索"""
    try:
        # 浮動小数点数として入力された場合の対策
        search_number = purchase_order_number
        if '.' in search_number and search_number.endswith('.0'):
            search_number = search_number.replace('.0', '')

        # マージ済みデータから検索
        details = OrderDetail.query.filter(
            db.or_(
                OrderDetail.order_number == search_number,
                OrderDetail.order_number == purchase_order_number
            )
        ).all()

        result_list = []

        for detail in details:
            result_list.append({
                'id': detail.id,
                'order_id': detail.order_id,
                'seiban': detail.seiban,
                'unit': detail.material,
                'item_name': detail.item_name,
                'spec1': detail.spec1,
                'quantity': detail.quantity,
                'unit_measure': detail.unit_measure,
                'is_received': detail.is_received,
                'delivery_date': detail.delivery_date,
                'reply_delivery_date': detail.reply_delivery_date or '',
                'supplier': detail.supplier,
                'source': 'merged',
                'staff': '-'
            })

        if not result_list:
            return jsonify({
                'found': False,
                'message': f'発注番号 {purchase_order_number} が見つかりません'
            }), 404

        return jsonify({
            'found': True,
            'count': len(result_list),
            'details': result_list
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/receive-by-purchase-order', methods=['POST'])
def receive_by_purchase_order():
    """発注番号で一括受入"""
    try:
        data = request.json
        purchase_order_number = data.get('purchase_order_number')
        
        if not purchase_order_number:
            return jsonify({'error': '発注番号を入力してください'}), 400
        
        # 発注番号に紐づく全アイテムを取得
        details = OrderDetail.query.filter_by(
            order_number=purchase_order_number
        ).all()
        
        if not details:
            return jsonify({
                'error': f'発注番号 {purchase_order_number} が見つかりません'
            }), 404
        
        # 社内加工チェック
        has_internal_processing = False
        received_count = 0
        
        for detail in details:
            if not detail.is_received:
                detail.is_received = True
                detail.received_at = datetime.now(timezone.utc)
                received_count += 1
                
                if detail.has_internal_processing:
                    has_internal_processing = True
                
                # ログを記録
                log = EditLog(
                    detail_id=detail.id,
                    action='receive',
                    ip_address=request.remote_addr,
                    user_agent=request.user_agent.string if request.user_agent else 'Unknown'
                )
                db.session.add(log)
        
        # 注文のステータス更新
        orders_to_update = set()
        for detail in details:
            orders_to_update.add(detail.order)
        
        for order in orders_to_update:
            # 1回のループで計算（二重ループ解消）
            details_list = order.details
            total_count = len(details_list)
            order_received_count = sum(1 for d in details_list if d.is_received)

            if order_received_count == total_count:
                order.status = '納品完了'
            elif order_received_count > 0:
                order.status = '納品中'
        
        db.session.commit()
        
        message = f'発注番号 {purchase_order_number} の {received_count} 件を受入しました'
        if has_internal_processing:
            message += '\n⚠️ 注意: 社内加工/追加工品が含まれています'
        
        return jsonify({
            'success': True,
            'message': message,
            'received_count': received_count,
            'has_internal_processing': has_internal_processing
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/purchase-order-stats')
def purchase_order_stats():
    """発注番号の統計情報を取得"""
    try:
        from sqlalchemy import func
        
        stats = db.session.query(
            OrderDetail.order_number,
            func.count(OrderDetail.id).label('total_items'),
            func.sum(OrderDetail.quantity).label('total_quantity'),
            func.sum(OrderDetail.is_received.cast(db.Integer)).label('received_items'),
            func.min(OrderDetail.spec1).label('spec1'),
            func.min(OrderDetail.item_name).label('item_name')  # 品名を追加
        ).filter(
            OrderDetail.order_number != '',
            OrderDetail.order_number != None
        ).group_by(
            OrderDetail.order_number
        ).all()
        
        result = []
        for stat in stats:
            order_number = stat.order_number
            if order_number and '.0' in order_number:
                order_number = order_number.replace('.0', '')
            
            completion_rate = (stat.received_items / stat.total_items * 100) if stat.total_items > 0 else 0
            result.append({
                'purchase_order_number': order_number,
                'total_items': stat.total_items,
                'total_quantity': stat.total_quantity or 0,
                'received_items': stat.received_items or 0,
                'completion_rate': round(completion_rate, 1),
                'spec1': stat.spec1 or '-',
                'item_name': stat.item_name or '-' 
            })
        
        result.sort(key=lambda x: x['completion_rate'])
        
        return jsonify({
            'total_orders': len(result),
            'stats': result
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/export-seiban/<seiban>')
def export_seiban(seiban):
    """製番全体をExcelファイルとしてエクスポート"""
    try:
        orders = Order.query.filter_by(seiban=seiban).all()
        
        if not orders:
            return jsonify({'success': False, 'error': '製番が見つかりません'}), 404
        
        wb = Workbook()
        ws = wb.active
        ws.title = seiban
        
        # ヘッダー
        headers = ['製番', 'ユニット', '品名', '仕様１', '仕様２', '数量', '単位', 
                   '納期', '手配区分', '発注番号', '仕入先', '仕入先CD', '備考', '検収日', '検収数']
        ws.append(headers)
        
        # 全ユニットのデータを出力
        for order in orders:
            for detail in order.details:
                row = [
                    order.seiban,
                    order.unit,
                    detail.item_name,
                    detail.spec1,
                    detail.spec2,
                    detail.quantity,
                    detail.unit_measure,
                    detail.delivery_date,
                    detail.order_type,
                    detail.order_number,
                    detail.supplier,
                    detail.supplier_cd,
                    detail.remarks,
                    detail.received_at.strftime('%Y-%m-%d %H:%M:%S') if detail.received_at else '',
                    '受入済' if detail.is_received else '未受入'
                ]
                ws.append(row)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()
        
        filename = f"{seiban}_all_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/export-seiban-family/<seiban>')
def export_seiban_family(seiban):
    """枝番ファミリー全体を1つのExcelファイルにまとめてエクスポート
    MHT0620を指定 → MHT0620, MHT0620-001, MHT0620-002... を1つのファイルに
    MHT0620-001を指定 → 同上（親製番を自動判定）
    """
    try:
        # 枝番ファミリーを取得
        family_seibans = get_seiban_family(seiban)

        if not family_seibans:
            return jsonify({'success': False, 'error': '製番が見つかりません'}), 404

        # 親製番を取得（ファイル名用）
        parent = get_parent_seiban(seiban)
        base_seiban = parent if parent else seiban

        # 全注文を取得
        orders = Order.query.filter(
            Order.seiban.in_(family_seibans),
            Order.is_archived == False
        ).all()

        if not orders:
            return jsonify({'success': False, 'error': '注文が見つかりません'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = f"{base_seiban}_枝番統合"

        # ヘッダー
        headers = ['製番', 'ユニット', '品名', '仕様１', '仕様２', '数量', '単位',
                   '納期', '手配区分', '発注番号', '仕入先', '仕入先CD', '備考',
                   '受入数量', '検収日', '受入状態']
        ws.append(headers)

        # ヘッダーのスタイル設定
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 製番順にソートしてデータを出力
        sorted_orders = sorted(orders, key=lambda o: (
            0 if o.seiban == base_seiban else 1,
            o.seiban,
            o.unit or ''
        ))

        for order in sorted_orders:
            for detail in order.details:
                # 受入数量の表示（received_quantityがNoneの場合は手配数と同じ）
                received_qty = ''
                if detail.is_received:
                    received_qty = detail.received_quantity if detail.received_quantity is not None else detail.quantity

                row = [
                    order.seiban,
                    order.unit,
                    detail.item_name,
                    detail.spec1,
                    detail.spec2,
                    detail.quantity,
                    detail.unit_measure,
                    detail.delivery_date,
                    detail.order_type,
                    detail.order_number,
                    detail.supplier,
                    detail.supplier_cd,
                    detail.remarks,
                    received_qty,
                    detail.received_at.strftime('%Y-%m-%d %H:%M:%S') if detail.received_at else '',
                    '受入済' if detail.is_received else '未受入'
                ]
                ws.append(row)

        # 列幅の調整
        column_widths = {
            'A': 15, 'B': 20, 'C': 25, 'D': 20, 'E': 20, 'F': 8,
            'G': 6, 'H': 12, 'I': 12, 'J': 12, 'K': 15, 'L': 10,
            'M': 20, 'N': 10, 'O': 18, 'P': 10
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # ===== シート2: 備考・仕様1キー集計シート（ピックアップ用） =====
        ws_pickup = wb.create_sheet(title="ピックアップ集計")

        # 備考と仕様1をキーに数量を集計（在庫部品のみ）
        pickup_data = {}  # キー: (備考, 仕様1) -> 集計データ

        for order in sorted_orders:
            for detail in order.details:
                # 在庫部品のみ対象
                if detail.order_type != '在庫部品':
                    continue

                key = (detail.remarks or '', detail.spec1 or '')
                if key not in pickup_data:
                    pickup_data[key] = {
                        'remarks': detail.remarks or '',
                        'spec1': detail.spec1 or '',
                        'item_name': detail.item_name or '',
                        'spec2': detail.spec2 or '',
                        'unit_measure': detail.unit_measure or '',
                        'total_quantity': 0,
                        'items': []  # 詳細情報のリスト
                    }
                pickup_data[key]['total_quantity'] += detail.quantity or 0
                pickup_data[key]['items'].append({
                    'seiban': order.seiban,
                    'unit': order.unit,
                    'quantity': detail.quantity
                })

        # ヘッダー
        pickup_headers = ['備考', '仕様１', '品名', '仕様２', '合計数量', '単位', '内訳（ユニット）']
        ws_pickup.append(pickup_headers)

        # ヘッダーのスタイル設定
        for col_idx, header in enumerate(pickup_headers, start=1):
            cell = ws_pickup.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 備考でソートして出力
        sorted_pickup = sorted(pickup_data.items(), key=lambda x: (x[0][0], x[0][1]))

        for (remarks, spec1), data in sorted_pickup:
            # 内訳を作成
            breakdown = ', '.join([f"{item['unit']}({item['quantity']})" for item in data['items']])

            row = [
                data['remarks'],
                data['spec1'],
                data['item_name'],
                data['spec2'],
                data['total_quantity'],
                data['unit_measure'],
                breakdown
            ]
            ws_pickup.append(row)

        # ピックアップシートの列幅調整
        pickup_widths = {'A': 15, 'B': 25, 'C': 25, 'D': 20, 'E': 10, 'F': 6, 'G': 50}
        for col, width in pickup_widths.items():
            ws_pickup.column_dimensions[col].width = width

        # ===== シート3: ユニット別分類シート =====
        ws_unit = wb.create_sheet(title="ユニット別分類")

        # ヘッダー
        unit_headers = ['ユニット', '品名', '仕様１', '仕様２', '数量', '単位', '手配区分', '備考']
        ws_unit.append(unit_headers)

        # ヘッダーのスタイル設定
        for col_idx, header in enumerate(unit_headers, start=1):
            cell = ws_unit.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # ユニットでグループ化（在庫部品のみ）
        unit_groups = {}
        for order in sorted_orders:
            unit_name = order.unit or '（ユニットなし）'
            if unit_name not in unit_groups:
                unit_groups[unit_name] = []
            for detail in order.details:
                if detail.order_type != '在庫部品':
                    continue
                unit_groups[unit_name].append({
                    'item_name': detail.item_name,
                    'spec1': detail.spec1,
                    'spec2': detail.spec2,
                    'quantity': detail.quantity,
                    'unit_measure': detail.unit_measure,
                    'order_type': detail.order_type,
                    'remarks': detail.remarks
                })

        # ユニット順でソートして出力
        current_row = 2
        unit_colors = ['FFF2CC', 'E2EFDA', 'DEEBF7', 'FCE4D6', 'EDEDED', 'D9E1F2']
        color_idx = 0

        for unit_name in sorted(unit_groups.keys()):
            items = unit_groups[unit_name]
            if not items:
                continue

            unit_color = unit_colors[color_idx % len(unit_colors)]
            color_idx += 1

            for item in items:
                row = [
                    unit_name,
                    item['item_name'],
                    item['spec1'],
                    item['spec2'],
                    item['quantity'],
                    item['unit_measure'],
                    item['order_type'],
                    item['remarks']
                ]
                ws_unit.append(row)

                # 行に背景色を設定
                for col_idx in range(1, len(unit_headers) + 1):
                    cell = ws_unit.cell(row=current_row, column=col_idx)
                    cell.fill = PatternFill(start_color=unit_color, end_color=unit_color, fill_type="solid")
                current_row += 1

        # ユニット別シートの列幅調整
        unit_widths = {'A': 20, 'B': 25, 'C': 25, 'D': 20, 'E': 8, 'F': 6, 'G': 12, 'H': 20}
        for col, width in unit_widths.items():
            ws_unit.column_dimensions[col].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()

        filename = f"{base_seiban}_枝番統合_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/seiban-family/<seiban>')
def get_seiban_family_api(seiban):
    """製番の枝番ファミリー情報を取得するAPI"""
    try:
        family_seibans = get_seiban_family(seiban)

        # 各製番の統計情報を取得
        result = []
        for s in family_seibans:
            orders = Order.query.filter_by(seiban=s, is_archived=False).all()
            total_details = sum(len(o.details) for o in orders)
            received_details = sum(sum(1 for d in o.details if d.is_received) for o in orders)

            result.append({
                'seiban': s,
                'is_parent': get_parent_seiban(s) is None,
                'unit_count': len(orders),
                'total_details': total_details,
                'received_details': received_details
            })

        parent = get_parent_seiban(seiban)
        return jsonify({
            'success': True,
            'base_seiban': parent if parent else seiban,
            'family': result,
            'total_seibans': len(result)
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/orders/delete-multiple', methods=['POST'])
def delete_multiple_orders():
    """複数の注文を一括削除"""
    try:
        data = request.json
        order_ids = data.get('order_ids', [])
        
        if not order_ids:
            return jsonify({'error': '削除する注文を選択してください'}), 400
        
        # 削除対象を取得
        orders = Order.query.filter(Order.id.in_(order_ids)).all()
        
        if not orders:
            return jsonify({'error': '対象の注文が見つかりません'}), 404
        
        deleted_count = 0
        seibans = []
        
        for order in orders:
            # 詳細を削除
            OrderDetail.query.filter_by(order_id=order.id).delete()
            
            # 注文を削除
            seibans.append(f"{order.seiban}({order.unit or '-'})")
            db.session.delete(order)
            deleted_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{deleted_count}件の注文を削除しました',
            'deleted': seibans
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/pallets/list')
def get_pallets_list():
    """パレット一覧を取得（品名付き）"""
    try:
        from sqlalchemy import func
        
        # 製番一覧表から品名を読み込み
        seiban_info = load_seiban_info()
        
        # パレット番号でグループ化して取得
        pallets = db.session.query(
            Order.pallet_number,
            Order.floor,
            func.count(Order.id).label('order_count')
        ).filter(
            Order.pallet_number != None,
            Order.pallet_number != '',
            Order.is_archived == False
        ).group_by(
            Order.pallet_number,
            Order.floor
        ).all()
        
        result = []
        for pallet in pallets:
            # そのパレットに含まれる注文を取得
            orders = Order.query.filter_by(
                pallet_number=pallet.pallet_number,
                is_archived=False
            ).all()
            
            orders_data = []
            for order in orders:
                # 製番一覧表から品名を取得（既にDBに品名がある場合はそちらを優先）
                product_name = order.product_name
                if not product_name and order.seiban in seiban_info:
                    product_name = seiban_info[order.seiban].get('product_name', '')
                
                # 製番一覧表から得意先略称を取得（既にDBに得意先略称がある場合はそちらを優先）
                customer_abbr = order.customer_abbr
                if not customer_abbr and order.seiban in seiban_info:
                    customer_abbr = seiban_info[order.seiban].get('customer_abbr', '')
                
                orders_data.append({
                    'id': order.id,
                    'seiban': order.seiban,
                    'unit': order.unit,
                    'status': order.status,
                    'product_name': product_name,
                    'customer_abbr': customer_abbr  # ← 製番一覧表からも取得するように修正
                })
            
            result.append({
                'pallet_number': pallet.pallet_number,
                'floor': pallet.floor,
                'order_count': pallet.order_count,
                'orders': orders_data
            })
        
        # パレット番号でソート
        result.sort(key=lambda x: x['pallet_number'])
        
        return jsonify({
            'success': True,
            'pallets': result,
            'total_pallets': len(result)
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/pallets/search')
def search_pallet():
    """製番、品名、または得意先略称でパレットを検索"""
    try:
        search_query = request.args.get('query', '')  # queryパラメータに統一
        
        if not search_query:
            return jsonify({'error': '検索キーワードを入力してください'}), 400
        
        # 製番一覧表から品名と得意先略称を読み込み
        seiban_info = load_seiban_info()  # ← load_seiban_data() から変更
        
        # 1. 製番で検索（部分一致）
        orders_by_seiban = Order.query.filter(
            Order.seiban.like(f'%{search_query}%'),
            Order.is_archived == False
        ).all()
        
        # 2. DBの得意先略称で検索（部分一致）
        orders_by_customer = Order.query.filter(
            Order.customer_abbr.like(f'%{search_query}%'),
            Order.is_archived == False
        ).all()
        
        # 3. 製番一覧表から品名と得意先略称で検索
        matching_seibans = []
        for seiban, info in seiban_info.items():
            product_name = info.get('product_name', '')
            customer_abbr = info.get('customer_abbr', '')
            
            # 品名または得意先略称に検索キーワードが含まれる場合
            if (search_query.lower() in product_name.lower() or 
                search_query.lower() in customer_abbr.lower()):
                matching_seibans.append(seiban)
        
        # 品名または得意先略称で見つかった製番の注文を取得
        orders_by_info = []
        if matching_seibans:
            orders_by_info = Order.query.filter(
                Order.seiban.in_(matching_seibans),
                Order.is_archived == False
            ).all()
        
        # 重複を除いて結合
        all_orders = list(set(orders_by_seiban + orders_by_customer + orders_by_info))
        
        if not all_orders:
            return jsonify({
                'success': False,
                'error': f'「{search_query}」に一致する製番、品名、または得意先略称が見つかりません'
            }), 404
        
        results = []
        for order in all_orders:
            # 製番一覧表から品名を取得（既にDBに品名がある場合はそちらを優先）
            product_name = order.product_name
            if not product_name and order.seiban in seiban_info:
                product_name = seiban_info[order.seiban].get('product_name', '')
            
            # 製番一覧表から得意先略称を取得（既にDBに得意先略称がある場合はそちらを優先）
            customer_abbr = order.customer_abbr
            if not customer_abbr and order.seiban in seiban_info:
                customer_abbr = seiban_info[order.seiban].get('customer_abbr', '')
            
            results.append({
                'id': order.id,
                'seiban': order.seiban,
                'unit': order.unit,
                'status': order.status,
                'pallet_number': order.pallet_number or '未設定',
                'floor': order.floor or '未設定',
                'product_name': product_name,
                'customer_abbr': customer_abbr  # ← 製番一覧表からも取得した値を使用
            })
        
        return jsonify({
            'success': True,
            'results': results
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500



@app.route('/api/pallets/<pallet_number>/label')
def get_pallet_label(pallet_number):
    """パレットラベルを生成して返す"""
    try:
        # そのパレット番号の注文を取得
        orders = Order.query.filter_by(
            pallet_number=pallet_number,
            is_archived=False
        ).all()
        
        if not orders:
            return jsonify({'error': 'パレットが見つかりません'}), 404
        
        # QRコード生成
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(f'PALLET:{pallet_number}')
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        
        # 画像をBase64エンコード
        buffered = BytesIO()
        img.save(buffered, format="PNG")
        qr_base64 = base64.b64encode(buffered.getvalue()).decode()
        
        # 注文情報をまとめる
        orders_info = []
        for order in orders:
            orders_info.append({
                'seiban': order.seiban,
                'unit': order.unit,
                'status': order.status,
                'product_name': order.product_name
            })
        
        return jsonify({
            'success': True,
            'pallet_number': pallet_number,
            'floor': orders[0].floor if orders else '未設定',
            'qr_code': qr_base64,
            'orders': orders_info,
            'order_count': len(orders)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/pallets/stats')
def get_pallet_stats():
    """パレット統計情報を取得"""
    try:
        from sqlalchemy import func
        
        # パレット別の統計
        stats = db.session.query(
            Order.pallet_number,
            Order.floor,
            func.count(Order.id).label('total_orders'),
            func.sum(case((Order.status == '納品完了', 1), else_=0)).label('completed_orders'),
            func.sum(case((Order.status == '納品中', 1), else_=0)).label('in_progress_orders')
        ).filter(
            Order.pallet_number != None,
            Order.pallet_number != '',
            Order.is_archived == False
        ).group_by(
            Order.pallet_number,
            Order.floor
        ).all()
        
        result = []
        for stat in stats:
            completion_rate = (stat.completed_orders / stat.total_orders * 100) if stat.total_orders > 0 else 0
            result.append({
                'pallet_number': stat.pallet_number,
                'floor': stat.floor or '未設定',
                'total_orders': stat.total_orders,
                'completed_orders': stat.completed_orders,
                'in_progress_orders': stat.in_progress_orders,
                'completion_rate': round(completion_rate, 1)
            })
        
        result.sort(key=lambda x: x['pallet_number'])
        
        return jsonify({
            'success': True,
            'stats': result
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# 🔥 ユーザー設定API
@app.route('/api/user-settings', methods=['GET'])
def get_user_settings():
    """ユーザー設定を取得"""
    try:
        client_ip = request.remote_addr or '0.0.0.0'
        settings = UserSettings.get_settings(client_ip)
        return jsonify({
            'success': True,
            'settings': settings
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/user-settings', methods=['POST'])
def update_user_settings():
    """ユーザー設定を更新"""
    try:
        client_ip = request.remote_addr or '0.0.0.0'
        data = request.get_json()

        if not data:
            return jsonify({'error': 'データがありません'}), 400

        UserSettings.update_settings(client_ip, **data)
        settings = UserSettings.get_settings(client_ip)

        return jsonify({
            'success': True,
            'message': '設定を保存しました',
            'settings': settings
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# 🔥 箱（パレット）QRスキャン用API - 未受入部品リスト取得
@app.route('/api/box/<pallet_number>/unreceived-parts')
def get_box_unreceived_parts(pallet_number):
    """箱（パレット）に紐づく未受入部品リストを取得"""
    try:
        # パレット番号に紐づくOrderを検索
        orders = Order.query.filter(
            Order.pallet_number == pallet_number,
            Order.is_archived == False
        ).all()

        if not orders:
            return jsonify({
                'success': True,
                'found': False,
                'pallet_number': pallet_number,
                'message': f'箱 {pallet_number} に紐づく製番が見つかりません',
                'parts': [],
                'summary': {'total': 0, 'unreceived': 0, 'received': 0}
            })

        # 各Orderの詳細を取得
        parts = []
        total_count = 0
        unreceived_count = 0
        received_count = 0

        for order in orders:
            for detail in order.details:
                total_count += 1
                if detail.is_received:
                    received_count += 1
                else:
                    unreceived_count += 1
                    parts.append({
                        'id': detail.id,
                        'seiban': order.seiban,
                        'unit': order.unit or '',
                        'order_number': detail.order_number,
                        'item_name': detail.item_name,
                        'spec1': detail.spec1,
                        'spec2': detail.spec2,
                        'quantity': detail.quantity,
                        'unit_measure': detail.unit_measure,
                        'delivery_date': detail.delivery_date,
                        'supplier': detail.supplier,
                        'order_type': detail.order_type,
                        'is_received': detail.is_received
                    })

        # 納期順でソート
        parts.sort(key=lambda x: x.get('delivery_date') or '9999-99-99')

        return jsonify({
            'success': True,
            'found': True,
            'pallet_number': pallet_number,
            'orders': [{'seiban': o.seiban, 'unit': o.unit or '', 'product_name': o.product_name or ''} for o in orders],
            'parts': parts,
            'summary': {
                'total': total_count,
                'unreceived': unreceived_count,
                'received': received_count
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# 🔥 箱QRコードで箱を検索
@app.route('/api/search-by-box-qr/<qr_data>')
def search_by_box_qr(qr_data):
    """箱QRコードから箱情報を検索
    QRコード形式: PALLET:P001, BOX:P001, またはP001などの直接パレット番号
    """
    try:
        # QRコードから箱番号を抽出
        pallet_number = None

        if qr_data.upper().startswith('PALLET:'):
            pallet_number = qr_data[7:].strip()
        elif qr_data.upper().startswith('BOX:'):
            pallet_number = qr_data[4:].strip()
        elif re.match(r'^[PDT]\d{3}$', qr_data.upper()):
            # P001, D001, T001 形式
            pallet_number = qr_data.upper()
        else:
            # そのまま使用
            pallet_number = qr_data.strip()

        if not pallet_number:
            return jsonify({
                'success': False,
                'error': '箱番号を認識できませんでした'
            }), 400

        # パレット番号に紐づくOrderを検索
        orders = Order.query.filter(
            Order.pallet_number == pallet_number,
            Order.is_archived == False
        ).all()

        if not orders:
            return jsonify({
                'success': True,
                'found': False,
                'pallet_number': pallet_number,
                'message': f'箱 {pallet_number} に紐づく製番が見つかりません'
            })

        # 未受入部品数をカウント
        total_unreceived = 0
        for order in orders:
            for detail in order.details:
                if not detail.is_received:
                    total_unreceived += 1

        return jsonify({
            'success': True,
            'found': True,
            'pallet_number': pallet_number,
            'floor': orders[0].floor if orders else None,
            'order_count': len(orders),
            'unreceived_count': total_unreceived,
            'orders': [{
                'id': o.id,
                'seiban': o.seiban,
                'unit': o.unit or '',
                'product_name': o.product_name or '',
                'status': o.status
            } for o in orders]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


import subprocess
import os


# 🔥 部品分類記号API
@app.route('/api/part-category/<part_code>')
def get_part_category_api(part_code):
    """部品コードから分類情報を取得するAPI
    例: /api/part-category/NAA-00123-01-00
    """
    try:
        parsed = PartCategory.parse_part_number(part_code)
        if not parsed:
            return jsonify({
                'success': False,
                'error': '無効な部品コード形式です'
            }), 400

        # 分類情報がある場合
        if 'major_category' in parsed:
            return jsonify({
                'success': True,
                'part_code': part_code,
                'category_code': parsed['category_code'],
                'serial': parsed['serial'],
                'derivative': parsed['derivative'],
                'revision': parsed['revision'],
                'major_category': parsed.get('major_category', ''),
                'minor_category': parsed.get('minor_category', ''),
                'note': parsed.get('note', ''),
                'description': {
                    'serial': 'シリアル番号（00000～99999）',
                    'derivative': '派生番号（同一形状・同一用途で+1）',
                    'revision': 'リビジョン番号（出図後変更のたび+1）'
                }
            })
        else:
            return jsonify({
                'success': True,
                'part_code': part_code,
                'category_code': parsed['category_code'],
                'serial': parsed['serial'],
                'derivative': parsed['derivative'],
                'revision': parsed['revision'],
                'major_category': None,
                'minor_category': None,
                'note': None,
                'warning': f'分類コード {parsed["category_code"]} はマスタに登録されていません'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/part-categories')
def get_all_part_categories():
    """全ての分類記号を取得するAPI（管理画面用）"""
    try:
        categories = PartCategory.query.order_by(PartCategory.code).all()
        return jsonify({
            'success': True,
            'count': len(categories),
            'categories': [{
                'code': c.code,
                'major_category': c.major_category,
                'minor_category': c.minor_category,
                'note': c.note
            } for c in categories]
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/open-cad/<int:detail_id>')
def open_cad_file(detail_id):
    """CADファイルを開く（ローカルは直接起動、リモートはダウンロード）"""
    try:
        detail = OrderDetail.query.get_or_404(detail_id)
        cad_info = get_cad_file_info(detail.spec1)
        
        if not cad_info:
            return jsonify({
                'success': False,
                'error': 'CADファイルが見つかりません'
            }), 404
        
        # 優先順位: PDF → mx2
        if cad_info['has_pdf']:
            file_path = cad_info['pdf_files'][0]
            file_type = 'PDF'
            mimetype = 'application/pdf'
        elif cad_info['has_mx2']:
            file_path = cad_info['mx2_files'][0]
            file_type = 'MX2'
            mimetype = 'application/octet-stream'
        else:
            return jsonify({
                'success': False,
                'error': 'ファイルが見つかりません'
            }), 404
        
        # 🔥 リクエスト元のIPアドレスを取得
        client_ip = request.remote_addr
        
        # 🔥 ローカルホストまたはサーバー自身からのアクセスか判定
        is_local = client_ip in ['127.0.0.1', '::1', 'localhost'] or \
                   client_ip == request.host.split(':')[0]  # サーバーのIPアドレス
        
        print(f"🔍 CADファイルアクセス: IP={client_ip}, ローカル={is_local}, ファイル={file_type}")
        
        # 🔥 MX2ファイルかつローカルアクセスの場合のみ直接起動
        if file_type == 'MX2' and is_local:
            try:
                # サーバー側でiCAD MXを起動
                os.startfile(file_path)
                
                return jsonify({
                    'success': True,
                    'file_type': file_type,
                    'file_name': os.path.basename(file_path),
                    'message': 'iCAD MXで図面を開きました',
                    'opened_locally': True
                })
            except Exception as e:
                return jsonify({
                    'success': False,
                    'error': f'ファイルを開けませんでした: {str(e)}'
                }), 500
        
        # 🔥 それ以外（リモートアクセスまたはPDF）はダウンロード/表示
        try:
            return send_file(
                file_path,
                mimetype=mimetype,
                as_attachment=(file_type == 'MX2'),  # MX2はダウンロード、PDFは表示
                download_name=os.path.basename(file_path)
            )
            
        except Exception as e:
            return jsonify({
                'success': False,
                'error': f'ファイルを送信できませんでした: {str(e)}'
            }), 500
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/open-cad-by-spec/<path:spec1>')
def open_cad_file_by_spec(spec1):
    """仕様1から直接CADファイルを開く（DB納品予定用）"""
    try:
        cad_info = get_cad_file_info(spec1)

        if not cad_info:
            return jsonify({
                'success': False,
                'error': 'CADファイルが見つかりません'
            }), 404

        # 優先順位: PDF → mx2
        if cad_info['has_pdf']:
            file_path = cad_info['pdf_files'][0]
            file_type = 'PDF'
            mimetype = 'application/pdf'
        elif cad_info['has_mx2']:
            file_path = cad_info['mx2_files'][0]
            file_type = 'MX2'
            mimetype = 'application/octet-stream'
        else:
            return jsonify({
                'success': False,
                'error': 'ファイルが見つかりません'
            }), 404

        client_ip = request.remote_addr
        is_local = client_ip in ['127.0.0.1', '::1', 'localhost'] or \
                   client_ip == request.host.split(':')[0]

        # MX2ファイルかつローカルアクセスの場合のみ直接起動
        if file_type == 'MX2' and is_local:
            try:
                os.startfile(file_path)
                return jsonify({
                    'success': True,
                    'file_type': file_type,
                    'file_name': os.path.basename(file_path),
                    'message': 'iCAD MXで図面を開きました',
                    'opened_locally': True
                })
            except Exception as e:
                return jsonify({
                    'success': False,
                    'error': f'ファイルを開けませんでした: {str(e)}'
                }), 500

        # それ以外はダウンロード/表示
        try:
            return send_file(
                file_path,
                mimetype=mimetype,
                as_attachment=(file_type == 'MX2'),
                download_name=os.path.basename(file_path)
            )
        except Exception as e:
            return jsonify({
                'success': False,
                'error': f'ファイルを送信できませんでした: {str(e)}'
            }), 500

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==================== 画像アップロード/表示機能 ====================

# 画像保存ディレクトリ
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'images')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# FullHD解像度
FULLHD_WIDTH = 1920
FULLHD_HEIGHT = 1080


def compress_to_fullhd(image_data):
    """画像をFullHD（1920x1080）以下に圧縮"""
    img = Image.open(io.BytesIO(image_data))

    # EXIF情報に基づいて回転を修正
    try:
        from PIL import ExifTags
        for orientation in ExifTags.TAGS.keys():
            if ExifTags.TAGS[orientation] == 'Orientation':
                break
        exif = img._getexif()
        if exif is not None:
            orientation_value = exif.get(orientation)
            if orientation_value == 3:
                img = img.rotate(180, expand=True)
            elif orientation_value == 6:
                img = img.rotate(270, expand=True)
            elif orientation_value == 8:
                img = img.rotate(90, expand=True)
    except (AttributeError, KeyError, IndexError):
        pass

    # 元のサイズ
    original_width, original_height = img.size

    # リサイズが必要かチェック
    if original_width <= FULLHD_WIDTH and original_height <= FULLHD_HEIGHT:
        # リサイズ不要、でもJPEGに変換して圧縮
        output = io.BytesIO()
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        img.save(output, format='JPEG', quality=85, optimize=True)
        return output.getvalue()

    # アスペクト比を維持してリサイズ
    ratio = min(FULLHD_WIDTH / original_width, FULLHD_HEIGHT / original_height)
    new_width = int(original_width * ratio)
    new_height = int(original_height * ratio)

    # リサイズ
    img = img.resize((new_width, new_height), Image.LANCZOS)

    # JPEG形式で保存
    output = io.BytesIO()
    if img.mode in ('RGBA', 'P'):
        img = img.convert('RGB')
    img.save(output, format='JPEG', quality=85, optimize=True)

    return output.getvalue()


@app.route('/api/order/<int:order_id>/upload-image', methods=['POST'])
def upload_order_image(order_id):
    """注文に画像をアップロード（FullHD圧縮）"""
    try:
        order = Order.query.get(order_id)
        if not order:
            return jsonify({'success': False, 'error': '注文が見つかりません'}), 404

        if 'image' not in request.files:
            return jsonify({'success': False, 'error': '画像ファイルがありません'}), 400

        file = request.files['image']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'ファイルが選択されていません'}), 400

        # 拡張子チェック
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        if ext not in allowed_extensions:
            return jsonify({'success': False, 'error': '許可されていないファイル形式です'}), 400

        # 画像データを読み込み
        image_data = file.read()

        # FullHDに圧縮
        compressed_data = compress_to_fullhd(image_data)

        # ファイル名を生成（order_id + タイムスタンプ）
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"order_{order_id}_{timestamp}.jpg"
        filepath = os.path.join(UPLOAD_FOLDER, filename)

        # 古い画像があれば削除
        if order.image_path and os.path.exists(order.image_path):
            try:
                os.remove(order.image_path)
            except:
                pass

        # 保存
        with open(filepath, 'wb') as f:
            f.write(compressed_data)

        # DBに保存
        order.image_path = filepath
        db.session.commit()

        return jsonify({
            'success': True,
            'message': '画像をアップロードしました',
            'image_url': f'/api/order/{order_id}/image'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/order/<int:order_id>/image')
def get_order_image(order_id):
    """注文の画像を取得"""
    try:
        order = Order.query.get(order_id)
        if not order:
            return jsonify({'error': '注文が見つかりません'}), 404

        if not order.image_path or not os.path.exists(order.image_path):
            return jsonify({'error': '画像がありません'}), 404

        return send_file(
            order.image_path,
            mimetype='image/jpeg'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/order/<int:order_id>/delete-image', methods=['DELETE'])
def delete_order_image(order_id):
    """注文の画像を削除"""
    try:
        order = Order.query.get(order_id)
        if not order:
            return jsonify({'success': False, 'error': '注文が見つかりません'}), 404

        if order.image_path and os.path.exists(order.image_path):
            try:
                os.remove(order.image_path)
            except:
                pass

        order.image_path = None
        db.session.commit()

        return jsonify({'success': True, 'message': '画像を削除しました'})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


if __name__ == '__main__':

    # 設定を取得
    config_obj = get_config()

    # キャッシュプリロードは不要（DB直接クエリを使用）

    # SSL/TLSコンテキストを取得
    ssl_context = None
    if hasattr(config_obj, 'USE_HTTPS') and config_obj.USE_HTTPS:
        from config import get_ssl_context
        ssl_context = get_ssl_context(config_obj)
        
        if ssl_context:
            print("🔒 HTTPS有効")
            if ssl_context == 'adhoc':
                print("⚠️  自己署名証明書を使用（開発用）")
                print("📱 QRコードスキャナーが使用可能です")
        else:
            print("⚠️  HTTPS無効 - QRコードスキャナーは使用できません")
    
    # サーバー起動
    app.run(
        debug=config_obj.DEBUG,
        host='0.0.0.0',
        port=8080,
        ssl_context=ssl_context
    )
