"""
ラベル作成プログラム - label_maker.py
製番・客先・ユニット・受入QRコードのラベルをExcelで出力
A4縦サイズ、1ページに同一ユニットのラベルを2つ（箱の前後貼付用）
"""

import sys
import os
import socket
import sqlite3
from pathlib import Path
from urllib.parse import quote
from io import BytesIO

import qrcode
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# === 設定 ===
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance', 'order_management.db')
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'labels')
SERVER_PORT = 8080
USE_HTTPS = True


def get_server_url():
    """サーバーURLを取得"""
    try:
        hostname = socket.gethostname()
        ip_address = socket.gethostbyname(hostname)
        protocol = 'https' if USE_HTTPS else 'http'
        return f"{protocol}://{ip_address}:{SERVER_PORT}"
    except Exception:
        return f"http://localhost:{SERVER_PORT}"


def get_receive_url(seiban, unit):
    """受入ページのURLを生成"""
    server_url = get_server_url()
    unit_encoded = quote(unit, safe='') if unit else ''
    return f"{server_url}/receive/{seiban}/{unit_encoded}"


def generate_qr_image(url, box_size=10):
    """QRコード画像を生成してBytesIOで返す"""
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=box_size,
        border=2,
    )
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf


def get_orders_by_seiban(seiban):
    """指定された製番の全ユニット情報をDBから取得"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, seiban, unit, product_name, customer_abbr, status
        FROM "order"
        WHERE seiban = ? AND is_archived = 0
        ORDER BY unit
    """, (seiban,))
    rows = cursor.fetchall()
    conn.close()
    return [dict(row) for row in rows]


def get_all_active_seibans():
    """アクティブな全製番を取得"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT seiban
        FROM "order"
        WHERE is_archived = 0
        ORDER BY seiban
    """)
    rows = cursor.fetchall()
    conn.close()
    return [row[0] for row in rows]


def create_label_on_sheet(ws, start_row, order):
    """ワークシートの指定行からラベル1枚分を描画"""
    seiban = order['seiban']
    unit = order['unit'] or ''
    product_name = order['product_name'] or ''
    customer = order['customer_abbr'] or ''
    receive_url = get_receive_url(seiban, unit)

    # --- スタイル定義 ---
    title_font = Font(name='Meiryo UI', size=14, bold=True)
    large_font = Font(name='Meiryo UI', size=28, bold=True)
    medium_font = Font(name='Meiryo UI', size=18, bold=True)
    small_font = Font(name='Meiryo UI', size=9, color='666666')
    label_font = Font(name='Meiryo UI', size=11, color='333333')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    # --- ラベル枠（7行分使用） ---
    # Row 0: 製番（大きく）
    # Row 1: 品名
    # Row 2: 客先
    # Row 3: ユニット（大きく）
    # Row 4-6: QRコード + URL

    r = start_row

    # ラベル枠線と背景
    for row_offset in range(7):
        for col in range(1, 9):  # A-H
            cell = ws.cell(row=r + row_offset, column=col)
            cell.border = thin_border

    # Row 0: 「製番」ラベル + 値
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cell_label = ws.cell(row=r, column=1, value='製番')
    cell_label.font = title_font
    cell_label.alignment = center_align
    cell_label.fill = header_fill
    cell_label.border = thin_border

    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    cell_val = ws.cell(row=r, column=3, value=seiban)
    cell_val.font = large_font
    cell_val.alignment = center_align
    cell_val.border = thin_border

    # Row 1: 「品名」ラベル + 値
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cell_label = ws.cell(row=r, column=1, value='品名')
    cell_label.font = title_font
    cell_label.alignment = center_align
    cell_label.fill = header_fill
    cell_label.border = thin_border

    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    cell_val = ws.cell(row=r, column=3, value=product_name)
    cell_val.font = medium_font
    cell_val.alignment = center_align
    cell_val.border = thin_border

    # Row 2: 「客先」ラベル + 値
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cell_label = ws.cell(row=r, column=1, value='客先')
    cell_label.font = title_font
    cell_label.alignment = center_align
    cell_label.fill = header_fill
    cell_label.border = thin_border

    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    cell_val = ws.cell(row=r, column=3, value=customer)
    cell_val.font = medium_font
    cell_val.alignment = center_align
    cell_val.border = thin_border

    # Row 3: 「ユニット」ラベル + 値
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cell_label = ws.cell(row=r, column=1, value='ユニット')
    cell_label.font = title_font
    cell_label.alignment = center_align
    cell_label.fill = header_fill
    cell_label.border = thin_border

    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    cell_val = ws.cell(row=r, column=3, value=unit if unit else '（なし）')
    cell_val.font = large_font
    cell_val.alignment = center_align
    cell_val.border = thin_border

    # Row 4-6: QRコード（左側）+ URL説明（右側）
    r += 1
    # QRコードを3行分の高さで配置
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=3)
    ws.merge_cells(start_row=r, start_column=4, end_row=r + 2, end_column=8)

    # QRコード画像
    qr_buf = generate_qr_image(receive_url, box_size=6)
    qr_img = Image(qr_buf)
    qr_img.width = 150
    qr_img.height = 150
    anchor_cell = f"A{r}"
    ws.add_image(qr_img, anchor_cell)

    # URL説明テキスト
    cell_url = ws.cell(row=r, column=4, value='受入ページ QRコード\nスマホで読取り')
    cell_url.font = label_font
    cell_url.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_url.border = thin_border

    # URL表示（小さく）
    cell_url_text = ws.cell(row=r + 2, column=4, value=receive_url)
    cell_url_text.font = small_font
    cell_url_text.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

    return start_row + 7  # 次のラベル開始行


def create_labels_for_seiban(seiban, output_path=None):
    """製番の全ユニットのラベルをExcelファイルに出力"""
    orders = get_orders_by_seiban(seiban)
    if not orders:
        print(f"❌ 製番 {seiban} のデータが見つかりません")
        return None

    wb = Workbook()
    wb.remove(wb.active)

    for order in orders:
        unit_display = order['unit'] if order['unit'] else 'ユニット名無し'
        sheet_name = f"{unit_display}"
        sheet_name = sheet_name[:31]  # シート名31文字制限

        ws = wb.create_sheet(title=sheet_name)

        # --- ページ設定（A4縦） ---
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4

        # --- 列幅設定 ---
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12

        # --- 行高さ設定 ---
        # ラベル1: 行1-7, ラベル2: 行9-15 （行8は区切り）
        label_row_heights = [45, 40, 40, 50, 55, 55, 55]  # 7行分
        for i, h in enumerate(label_row_heights):
            ws.row_dimensions[1 + i].height = h

        ws.row_dimensions[8].height = 20  # 区切り行

        for i, h in enumerate(label_row_heights):
            ws.row_dimensions[9 + i].height = h

        # --- ラベル1（上半分）---
        create_label_on_sheet(ws, 1, order)

        # --- 区切り線（行8）---
        for col in range(1, 9):
            cell = ws.cell(row=8, column=col)
            cell.border = Border(
                bottom=Side(style='dashed', color='999999')
            )
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=8, column=4, value='✂ 切り取り線').font = Font(
            name='Meiryo UI', size=8, color='999999'
        )
        ws.cell(row=8, column=4).alignment = Alignment(horizontal='center', vertical='center')

        # --- ラベル2（下半分）---
        create_label_on_sheet(ws, 9, order)

    # --- 保存 ---
    if output_path is None:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        safe_seiban = seiban.replace('/', '_').replace('\\', '_')
        output_path = os.path.join(OUTPUT_DIR, f'{safe_seiban}_ラベル.xlsx')

    wb.save(output_path)
    wb.close()
    print(f"✅ ラベル出力完了: {output_path}")
    print(f"   製番: {seiban}")
    print(f"   ユニット数: {len(orders)}")
    for o in orders:
        unit = o['unit'] if o['unit'] else '（なし）'
        print(f"   - {unit}")
    return output_path


def create_labels_for_all():
    """全アクティブ製番のラベルを出力"""
    seibans = get_all_active_seibans()
    if not seibans:
        print("❌ アクティブな製番がありません")
        return

    print(f"全 {len(seibans)} 製番のラベルを作成します...")
    for seiban in seibans:
        create_labels_for_seiban(seiban)
    print(f"\n✅ 全ラベル出力完了（{OUTPUT_DIR}）")


def main():
    print("=" * 50)
    print("  ラベル作成プログラム")
    print("=" * 50)

    if len(sys.argv) > 1:
        # コマンドライン引数で製番指定
        seiban = sys.argv[1]
        output_path = sys.argv[2] if len(sys.argv) > 2 else None
        create_labels_for_seiban(seiban, output_path)
    else:
        # 対話モード
        seibans = get_all_active_seibans()
        if not seibans:
            print("❌ アクティブな製番がありません")
            return

        print("\n製番一覧:")
        for i, s in enumerate(seibans, 1):
            print(f"  {i}. {s}")
        print(f"  A. 全製番を出力")

        choice = input("\n番号を入力（またはAで全出力）: ").strip()

        if choice.upper() == 'A':
            create_labels_for_all()
        else:
            try:
                idx = int(choice) - 1
                if 0 <= idx < len(seibans):
                    create_labels_for_seiban(seibans[idx])
                else:
                    print("❌ 無効な番号です")
            except ValueError:
                # 直接製番を入力した場合
                create_labels_for_seiban(choice)


if __name__ == '__main__':
    main()
