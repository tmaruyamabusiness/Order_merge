"""
Excel出力・シート生成・更新 - excel_export.py
手配発注リストExcelの生成、ガントチャート、COM経由のリフレッシュ
"""
import os
import re
import time
import shutil
from pathlib import Path
from flask import current_app
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins

from app import db, Order, OrderDetail
from utils import Constants, DataUtils, MekkiUtils, ExcelStyler, generate_qr_code, create_gantt_chart_sheet, DeliveryUtils

# 更新対象のExcelファイル一覧（DB直接クエリに移行中）
EXCEL_FILES_TO_REFRESH = [
    {
        'path': r"\\SERVER3\share-data\Document\Acrossデータ\製番一覧表.xlsx",
        'name': '製番一覧表',
        'sheet': '製番'
    }
]


def get_order_excel_filename(seiban, product_name=None, customer_abbr=None):
    """製番に対応するExcelファイル名を取得"""
    if product_name:
        safe_product_name = re.sub(r'[\\/:*?"<>|]', '', product_name)
        if customer_abbr:
            safe_customer_abbr = re.sub(r'[\\/:*?"<>|]', '', customer_abbr)
            filename = f"{seiban}_{safe_product_name}_{safe_customer_abbr}_手配発注リスト.xlsx"
        else:
            filename = f"{seiban}_{safe_product_name}_手配発注リスト.xlsx"
    else:
        filename = f"{seiban}_手配発注リスト.xlsx"
    return filename


def get_order_excel_path(seiban, product_name=None, customer_abbr=None):
    """製番に対応するExcelファイルパスを取得（閲覧用メインファイル）"""
    export_dir = Path(current_app.config['EXPORT_EXCEL_PATH'])
    export_dir.mkdir(parents=True, exist_ok=True)
    filename = get_order_excel_filename(seiban, product_name, customer_abbr)
    return str(export_dir / filename)


def get_order_excel_data_path(seiban, product_name=None, customer_abbr=None):
    """製番に対応するExcelファイルパスを取得（元データ用dataフォルダ）"""
    export_dir = Path(current_app.config['EXPORT_EXCEL_PATH']) / 'data'
    export_dir.mkdir(parents=True, exist_ok=True)
    filename = get_order_excel_filename(seiban, product_name, customer_abbr)
    return str(export_dir / filename)


def get_server_url():
    """サーバーのURLを取得（IP + ポート）"""
    try:
        import socket
        from config import get_config
        hostname = socket.gethostname()
        ip_address = socket.gethostbyname(hostname)

        config_obj = get_config()
        use_https = getattr(config_obj, 'USE_HTTPS', False)
        port = getattr(config_obj, 'PORT', 8080)

        protocol = 'https' if use_https else 'http'
        return f"{protocol}://{ip_address}:{port}"
    except Exception as e:
        print(f"サーバーURL取得エラー: {e}")
        return "http://localhost:8080"


def save_order_to_excel(order, filepath, data_filepath=None):
    """注文をExcelファイルに保存（dataフォルダに元データ保存→メインファイルにコピー）"""
    try:
        unit_display = order.unit if order.unit else 'ユニット名無し'
        sheet_name = f"{order.seiban}_{unit_display}"
        sheet_name = re.sub(r'[\\\/\?\*\[\]:]', '', sheet_name)[:31]

        if data_filepath is None:
            data_filepath = get_order_excel_data_path(order.seiban, order.product_name, order.customer_abbr)

        # === Step 1: dataフォルダに元データを保存 ===
        if Path(data_filepath).exists():
            try:
                wb = load_workbook(data_filepath)
            except PermissionError:
                wb = Workbook()
                wb.remove(wb.active)

            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            if "納期ガントチャート" in wb.sheetnames:
                del wb["納期ガントチャート"]
        else:
            wb = Workbook()
            wb.remove(wb.active)

        orders = Order.query.filter_by(seiban=order.seiban, is_archived=False).all()
        create_gantt_chart_sheet(wb, order.seiban, orders)

        ws = wb.create_sheet(sheet_name)
        create_order_sheet(ws, order, sheet_name)

        wb.save(data_filepath)
        wb.close()
        print(f"✅ 元データ保存完了: {data_filepath}")

        # === Step 2: メインファイルにコピー（閲覧用） ===
        try:
            shutil.copy2(data_filepath, filepath)
            print(f"✅ メインファイル更新: {filepath}")
            return True, None
        except PermissionError:
            print(f"⚠️ メインファイル使用中（元データは保存済み）: {filepath}")
            return True, "メインファイルは使用中ですが、元データは保存されました"

    except Exception as e:
        return False, str(e)


def update_order_excel(order_id):
    """注文IDに対応するExcelファイルを更新"""
    try:
        order = db.session.get(Order, order_id)
        if not order:
            return False, "注文が見つかりません"

        filepath = get_order_excel_path(order.seiban, order.product_name, order.customer_abbr)
        success, error = save_order_to_excel(order, filepath)

        if success:
            print(f"✅ Excel更新成功: {filepath}")
        else:
            print(f"❌ Excel更新失敗: {error}")

        return success, error
    except Exception as e:
        return False, str(e)


def create_order_sheet(ws, order, sheet_name=None):
    """ワークシート作成（縦向き印刷、QRコードH列配置）"""
    from openpyxl.drawing.image import Image
    from io import BytesIO
    import qrcode

    if sheet_name:
        ws.title = sheet_name

    unit_display = order.unit if order.unit else 'ユニット名無し'
    customer = order.customer_abbr if order.customer_abbr else ''
    memo = order.memo2 if order.memo2 else ''
    product_name = order.product_name if order.product_name else ''

    # QRコード生成
    try:
        from urllib.parse import quote
        server_url = get_server_url()
        unit_encoded = quote(order.unit, safe='') if order.unit else ''
        receive_url = f"{server_url}/receive/{order.seiban}/{unit_encoded}"

        qr = qrcode.QRCode(version=1, box_size=8, border=3)
        qr.add_data(receive_url)
        qr.make(fit=True)

        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_buffer = BytesIO()
        qr_img.save(qr_buffer, format='PNG')
        qr_buffer.seek(0)

        img = Image(qr_buffer)
        img.width = 100
        img.height = 100
        ws.add_image(img, 'I1')

        ws['M1'] = '💻️ 受入確認専用ページ(社内LANよりアクセス)'
        ws['M1'].font = Font(size=9, bold=True)
        ws['M1'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        ws['M2'] = receive_url
        ws['M2'].font = Font(size=8, color='0000FF', underline='single')
        ws['M2'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    except Exception as e:
        print(f"⚠️ QRコード生成エラー: {e}")

    # 1行目: 製番 + 品名 + 得意先 + メモ
    a1_text_parts = [order.seiban]
    if product_name:
        a1_text_parts.append(product_name)
    if customer:
        a1_text_parts.append(customer)
    if memo:
        a1_text_parts.append(memo)

    ws['A1'] = ' '.join(a1_text_parts)
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    ws['A2'] = unit_display
    ws['A2'].font = Font(size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    ws['A3'] = '※ピンク塗は受入済 製番外の持ち出しは必ず記録を残すこと データは保存先にて随時更新'
    ws['A3'].font = Font(size=9, bold=True, color=Constants.COLOR_RED)
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')

    ws['A4'] = r'\\SERVER3\Share-data\Document\仕入れ\002_手配リスト\手配発注リスト'
    ws['A4'].font = Font(size=9, bold=True, color=Constants.COLOR_RED)
    ws['A4'].alignment = Alignment(horizontal='left', vertical='center')

    remarks_text = order.remarks if order.remarks else ''
    ws['J2'] = f'備考：{remarks_text}'
    ws['J2'].font = Font(size=9)
    ws['J2'].alignment = Alignment(horizontal='left', vertical='center')

    ws['J3'] = '保管場所：'
    ws['J3'].font = Font(size=10, bold=True)
    ws['J3'].alignment = Alignment(horizontal='right', vertical='center')

    ws['K3'] = order.floor if order.floor else ''
    ws['K3'].font = Font(size=10)
    ws['K3'].alignment = Alignment(horizontal='left', vertical='center')

    ws['J4'] = '場所番号：'
    ws['J4'].font = Font(size=10, bold=True)
    ws['J4'].alignment = Alignment(horizontal='right', vertical='center')

    ws['K4'] = order.pallet_number if order.pallet_number else ''
    ws['K4'].font = Font(size=10)
    ws['K4'].alignment = Alignment(horizontal='left', vertical='center')

    ws.row_dimensions[1].height = 35

    # ヘッダー行（6行目）
    headers = Constants.EXCEL_COLUMNS
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color=Constants.COLOR_HEADER,
                                end_color=Constants.COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    column_widths = {
        'A': 9, 'B': 6, 'C': 9, 'D': 11, 'E': 9, 'F': 5,
        'G': 4, 'H': 18, 'I': 15, 'J': 12, 'K': 10, 'L': 8, 'M': 12
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 検収データを読み込み
    delivery_dict = DeliveryUtils.load_delivery_data()

    # データ行（7行目から）
    row_idx = 7
    parent_details = [d for d in order.details if d.parent_id is None]

    for detail in parent_details:
        row_idx = _write_detail_row(ws, detail, row_idx, is_parent=True, delivery_dict=delivery_dict)
        children = [d for d in order.details if d.parent_id == detail.id]
        for child in children:
            row_idx = _write_detail_row(ws, child, row_idx, is_parent=False, delivery_dict=delivery_dict)

    # ページ設定
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.3, bottom=0.5, header=0.15, footer=0.2
    )

    ws.print_title_rows = '1:6'
    ws.print_area = f'A1:M{row_idx - 1}'

    footer_parts = [order.seiban]
    if unit_display and unit_display != 'ユニット名無し':
        footer_parts.append(unit_display)
    if product_name:
        display_product_name = product_name if len(product_name) <= 20 else product_name[:20] + '...'
        footer_parts.append(display_product_name)
    if customer:
        footer_parts.append(customer)
    if memo:
        footer_parts.append(memo)

    footer_text = '_'.join(footer_parts)

    for footer in [ws.oddFooter, ws.evenFooter, ws.firstFooter]:
        footer.left.text = f"&10&B{footer_text}"
        footer.center.text = "&P / &N"
        footer.right.text = f"&10&B{footer_text}"

    ws.sheet_view.view = 'pageBreakPreview'

    return ws


def _get_cad_hyperlink(spec1):
    """仕様1からCADファイルのハイパーリンクパスを取得"""
    if not spec1 or not str(spec1).startswith('N'):
        return None
    spec1 = str(spec1)
    parts = spec1.split('-')
    if len(parts) < 2 or len(parts[0]) < 2:
        return None
    folder_letter = parts[0][1].upper()

    import glob

    # SERVER3のCADフォルダ（PDF優先 → mx2 → フォルダ）
    cad_folder = f"\\\\SERVER3\\Share-data\\CadData\\Parts\\{folder_letter}"
    try:
        pdf_files = glob.glob(os.path.join(cad_folder, f"{spec1}*.pdf"))
        if pdf_files:
            return pdf_files[0]
        mx2_files = glob.glob(os.path.join(cad_folder, f"{spec1}*.mx2"))
        if mx2_files:
            return mx2_files[0]
    except Exception:
        pass
    return cad_folder


def _write_detail_row(ws, detail, row_idx, is_parent=True, delivery_dict=None):
    """詳細行を出力"""
    is_blank = '加工用ブランク' in str(detail.order_type)
    supplier_cd = getattr(detail, 'supplier_cd', None)
    spec1_value = detail.spec1 or ''
    spec2_value = detail.spec2 or ''
    is_mekki = MekkiUtils.is_mekki_target(supplier_cd, spec2_value, spec1_value)

    remarks = MekkiUtils.add_mekki_alert(detail.remarks) if is_mekki else (detail.remarks or '')

    delivery_info = DeliveryUtils.get_delivery_info(detail.order_number, delivery_dict)
    delivery_date = delivery_info.get('納入日', '')
    delivery_qty = delivery_info.get('納入数', 0)
    delivery_qty_display = delivery_qty if delivery_qty > 0 else ''

    # 仕様1のCADリンクを事前に取得
    cad_link = _get_cad_hyperlink(spec1_value)

    data = [
        delivery_date, delivery_qty_display,
        detail.delivery_date, detail.supplier, detail.order_number,
        detail.quantity, detail.unit_measure, detail.item_name,
        detail.spec1, spec2_value, detail.order_type, detail.maker, remarks
    ]

    row_fill = ExcelStyler.get_fill(detail.is_received, row_idx % 2 == 0, not is_parent)
    cell_font = ExcelStyler.get_font(is_blank, False)

    for col, value in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.fill = row_fill
        cell.alignment = Alignment(vertical='center')

        if col == 10 and is_mekki:
            cell.font = ExcelStyler.get_font(False, True)
        elif cell_font:
            cell.font = cell_font

        if col == 8 and not is_parent:
            cell.value = f"  └ {value}"

        # 仕様１(col=9)にCADハイパーリンクを設定
        if col == 9 and cad_link:
            cell.hyperlink = cad_link
            cell.font = Font(color="0000FF", underline="single", size=cell_font.size if cell_font and cell_font.size else 10)

    ws.row_dimensions[row_idx].height = 27
    return row_idx + 1


def refresh_single_excel(excel_path, file_name):
    """単一のExcelファイルをCOM経由で更新"""
    import win32com.client as win32
    excel = None
    wb = None
    try:
        if not os.path.exists(excel_path):
            return False, f"ファイルが見つかりません: {excel_path}"

        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.EnableEvents = False

        wb = excel.Workbooks.Open(
            Filename=excel_path, UpdateLinks=0, ReadOnly=False, Notify=False
        )

        if hasattr(wb, 'Connections'):
            for i in range(1, wb.Connections.Count + 1):
                try:
                    conn = wb.Connections(i)
                    if hasattr(conn, 'ODBCConnection'):
                        conn.ODBCConnection.BackgroundQuery = False
                    elif hasattr(conn, 'OLEDBConnection'):
                        conn.OLEDBConnection.BackgroundQuery = False
                except:
                    pass

        wb.RefreshAll()
        excel.CalculateUntilAsyncQueriesDone()
        excel.CalculateFull()
        time.sleep(2)

        wb.Save()
        wb.Close(SaveChanges=False)
        excel.Quit()

        return True, f"{file_name}を更新しました"

    except Exception as e:
        return False, f"{file_name}の更新エラー: {str(e)}"

    finally:
        try:
            if wb:
                wb.Close(SaveChanges=False)
        except:
            pass
        try:
            if excel:
                excel.Quit()
        except:
            pass


def refresh_excel_file():
    """複数のExcelファイルを順番に更新"""
    import pythoncom
    results = []
    all_success = True

    try:
        pythoncom.CoInitialize()

        for file_info in EXCEL_FILES_TO_REFRESH:
            excel_path = file_info['path']
            file_name = file_info['name']

            print(f"📊 {file_name} を更新中...")
            success, message = refresh_single_excel(excel_path, file_name)
            results.append({'name': file_name, 'success': success, 'message': message})

            if not success:
                all_success = False
                print(f"  ❌ {message}")
            else:
                print(f"  ✅ {message}")

            time.sleep(3)

        success_count = sum(1 for r in results if r['success'])
        total_count = len(results)

        if all_success:
            message = f"全{total_count}ファイルを更新しました"
        else:
            message = f"{success_count}/{total_count}ファイルを更新（一部エラー）"

        return all_success, message, results

    except Exception as e:
        return False, f"更新エラー: {str(e)}", results

    finally:
        try:
            pythoncom.CoUninitialize()
        except:
            pass
