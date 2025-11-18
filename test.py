"""
ネットワークアクセステストスクリプト
パスの形式による違いをテスト
"""

import os
from pathlib import Path

def test_path_formats():
    """異なるパス形式でのアクセステスト"""
    
    print("=" * 60)
    print("パス形式テスト")
    print("=" * 60)
    
    # テストするパスのバリエーション
    test_paths = [
        # Raw string
        (r'\\server3\Share-data\Document\仕入れ\002_手配リスト\手配発注_ALL.xlsx', "Raw string"),
        
        # 通常文字列（エスケープ）
        ('\\\\server3\\Share-data\\Document\\仕入れ\\002_手配リスト\\手配発注_ALL.xlsx', "Escaped string"),
        
        # Forward slash
        ('//server3/Share-data/Document/仕入れ/002_手配リスト/手配発注_ALL.xlsx', "Forward slash"),
        
        # Path.joinpath
        (str(Path('//server3/Share-data/Document/仕入れ/002_手配リスト/手配発注_ALL.xlsx')), "Path object"),
        
        # 大文字小文字のバリエーション
        (r'\\SERVER3\Share-data\Document\仕入れ\002_手配リスト\手配発注_ALL.xlsx', "SERVER3 (uppercase)"),
        (r'\\server3\share-data\Document\仕入れ\002_手配リスト\手配発注_ALL.xlsx', "share-data (lowercase)"),
    ]
    
    results = []
    
    for path, description in test_paths:
        print(f"\n📝 {description}")
        print(f"   パス: {path}")
        
        # os.path.exists
        exists_os = os.path.exists(path)
        print(f"   os.path.exists: {'✅' if exists_os else '❌'} {exists_os}")
        
        # Path.exists
        try:
            path_obj = Path(path)
            exists_path = path_obj.exists()
            print(f"   Path.exists: {'✅' if exists_path else '❌'} {exists_path}")
        except Exception as e:
            exists_path = False
            print(f"   Path.exists: ❌ エラー: {e}")
        
        # os.access
        if exists_os:
            readable = os.access(path, os.R_OK)
            print(f"   読み取り可能: {'✅' if readable else '❌'} {readable}")
        
        results.append({
            'description': description,
            'path': path,
            'exists': exists_os or exists_path
        })
    
    # 成功したパスを表示
    print("\n" + "=" * 60)
    print("結果サマリー")
    print("=" * 60)
    
    successful = [r for r in results if r['exists']]
    if successful:
        print("\n✅ アクセス可能なパス形式:")
        for r in successful:
            print(f"  - {r['description']}")
            print(f"    {r['path']}")
    else:
        print("\n❌ アクセス可能なパスがありません")
    
    return results

def test_direct_access():
    """直接的なファイルアクセステスト"""
    
    print("\n" + "=" * 60)
    print("直接アクセステスト")
    print("=" * 60)
    
    path = r'\\server3\Share-data\Document\仕入れ\002_手配リスト\手配発注_ALL.xlsx'
    
    # ファイルの読み込みテスト
    print(f"\n📄 ファイル: {path}")
    
    if os.path.exists(path):
        print("✅ ファイルが存在します")
        
        # サイズ取得
        try:
            size = os.path.getsize(path)
            print(f"✅ ファイルサイズ: {size / (1024*1024):.2f} MB")
        except Exception as e:
            print(f"❌ サイズ取得エラー: {e}")
        
        # pandas での読み込みテスト
        try:
            import pandas as pd
            print("\npandasでの読み込みテスト...")
            
            # シート名のみ取得（軽量）
            xl_file = pd.ExcelFile(path)
            sheet_names = xl_file.sheet_names
            print(f"✅ シート数: {len(sheet_names)}")
            print(f"   シート名: {', '.join(sheet_names[:5])}")  # 最初の5シート
            
        except Exception as e:
            print(f"❌ pandas読み込みエラー: {e}")
        
        # openpyxl での読み込みテスト
        try:
            from openpyxl import load_workbook
            print("\nopenpyxlでの読み込みテスト...")
            
            wb = load_workbook(path, read_only=True)
            sheet_names = wb.sheetnames
            print(f"✅ シート数: {len(sheet_names)}")
            wb.close()
            
        except Exception as e:
            print(f"❌ openpyxl読み込みエラー: {e}")
            
    else:
        print("❌ ファイルが存在しません")

if __name__ == "__main__":
    # パス形式テスト
    test_path_formats()
    
    # 直接アクセステスト
    test_direct_access()
    
    print("\n" + "=" * 60)
    print("テスト完了")
    print("=" * 60)
    
    input("\nEnterキーを押して終了...")